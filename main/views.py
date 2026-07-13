from io import BytesIO
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import InstitutionRegisterForm, LoginForm, ClassHeadLoginForm, SubjectHeadLoginForm, StudentLoginForm, ParentLoginForm, AddClassForm, AddSubjectForm, AddStudentForm, ClassUploadForm, SubjectUploadForm, StudentUploadForm
from .models import Institution, Classes, Subjects, Students, Users, Parents, Chat, Announcements, Attendance, StudentEvaluation, QuizResponse, Quizzes, AuditLog
from django.db import IntegrityError, transaction
from django.db import connection
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse
from django.db.models import Q, Max
from django.views.decorators.csrf import csrf_exempt
import json
from django.utils.timezone import now, localtime
from django.utils import timezone
from django.http import HttpResponse
from django.core.mail import send_mail
from django.core.cache import cache
from django.utils.crypto import get_random_string
import pandas as pd
from django.core.files.storage import FileSystemStorage
import os
from datetime import datetime
import openpyxl
from decimal import Decimal, ROUND_HALF_UP
import random
import re
from textblob import TextBlob
import joblib
import numpy as np
from django.conf import settings
from ml.predict import predict_performance
from django.contrib.auth.hashers import make_password
from django.core.cache import cache  # Use Django cache
import textwrap
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Quizzes, QuizQuestions, Classes # Ensure correct imports
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import base64
from cryptography.fernet import Fernet
from django.conf import settings
import hashlib
from django.core.mail import send_mail
from django.conf import settings
import requests

# Password Encryption/Decryption Helper Functions
def get_encryption_key():
    """Generate a consistent encryption key from Django SECRET_KEY"""
    # Create a 32-byte key from SECRET_KEY for Fernet encryption
    return base64.urlsafe_b64encode(hashlib.sha256(settings.SECRET_KEY.encode()).digest())

def encrypt_password(password):
    """Encrypt password using Fernet symmetric encryption"""
    try:
        f = Fernet(get_encryption_key())
        encrypted = f.encrypt(password.encode())
        return encrypted.decode()
    except Exception as e:
        print(f"❌ Encryption error: {str(e)}")
        return password  # Fallback to plain password if encryption fails

def decrypt_password(encrypted_password):
    """Decrypt password using Fernet symmetric encryption"""
    try:
        f = Fernet(get_encryption_key())
        decrypted = f.decrypt(encrypted_password.encode())
        return decrypted.decode()
    except Exception as e:
        print(f"❌ Decryption error: {str(e)}")
        return encrypted_password  # Return as-is if decryption fails

def verify_password(plain_password, encrypted_password):
    """Verify a plain password against an encrypted password"""
    try:
        decrypted = decrypt_password(encrypted_password)
        return plain_password == decrypted
    except Exception as e:
        print(f"❌ Password verification error: {str(e)}")
        return False


# Index page
def index(request):
    return render(request, 'index.html')

from django.urls import reverse
def send_account_creation_email(email, password, role, name, institution_email, studId=None):
    subject = "Welcome to Eduke – Your Account is Ready"
    userId = email if studId is None else studId
    role_name = role.replace('_', ' ').title()

    # Context-specific messages
    role_messages = {
        "class_head": "As a <strong>Class Head</strong>, you now have administrative access to manage student rosters and classroom activities.",
        "subject_head": "As a <strong>Subject Head</strong>, you can now curate curriculum content and monitor subject-specific performance.",
        "student": "Your student portal is live! Access your study materials and stay in touch with your mentors.",
        "parent": "Welcome! You can now stay updated on your child's academic progress and communicate directly with teachers."
    }

    specific_content = role_messages.get(role, "Your account has been successfully provisioned.")
    portal_url = f"{settings.BASE_URL}{reverse('user_portal')}"
    # High-End UI Design in Purple and Red
    html_content = f"""
    <div style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #333; max-width: 600px; margin: 20px auto; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 15px rgba(0,0,0,0.1); border: 1px solid #e0e0e0;">
        
        <div style="background: linear-gradient(135deg, #6B32A1 0%, #4B2075 100%); padding: 30px; text-align: center;">
            <h1 style="color: white; margin: 0; font-size: 28px; letter-spacing: 2px; text-transform: uppercase;">Eduke</h1>
            <p style="color: #d1b9f0; margin: 10px 0 0 0; font-size: 14px;">Empowering the future of education</p>
        </div>

        <div style="padding: 40px 30px; background-color: #ffffff;">
            <h2 style="color: #4B2075; margin-top: 0;">Welcome, {name}!</h2>
            <p style="font-size: 16px; color: #555;">Your account has been created successfully. Below are your login credentials for the portal:</p>
            
            <div style="background-color: #f4f0fa; border-left: 5px solid #6B32A1; padding: 20px; margin: 25px 0; border-radius: 4px;">
                <p style="margin: 8px 0;"><strong>Role:</strong> <span style="color: #6B32A1;">{role_name}</span></p>
                <p style="margin: 8px 0;"><strong>Username:</strong> <code>{userId}</code></p>
                <p style="margin: 8px 0;"><strong>Temporary Password:</strong> <span style="background: #fff; border: 1px dashed #6B32A1; padding: 2px 8px; font-weight: bold; color: #4B2075; border-radius: 4px;">{password}</span></p>
            </div>

            <p style="font-size: 15px;">{specific_content}</p>
            
            <div style="text-align: center; margin: 35px 0;">
                <a href="{portal_url}" style="background-color: #6B32A1; color: white; padding: 14px 35px; text-decoration: none; border-radius: 50px; font-weight: bold; font-size: 16px; box-shadow: 0 4px 10px rgba(107, 50, 161, 0.3);">
                    Go to Dashboard
                </a>
            </div>

            <div style="background-color: #fff5f5; border: 1px solid #feb2b2; padding: 15px; border-radius: 8px; text-align: center;">
                <p style="margin: 0; color: #c53030; font-weight: bold; font-size: 14px;">
                    ⚠️ ACTION REQUIRED: For your security, you must change your password immediately after your first login.
                </p>
            </div>
        </div>

        <div style="background-color: #f8f9fa; padding: 20px; text-align: center; font-size: 12px; color: #888; border-top: 1px solid #eee;">
            <p style="margin: 0;">This email was sent on behalf of <strong>{institution_email}</strong></p>
            <p style="margin: 5px 0 0 0;">&copy; 2026 Eduke Education Platform. All rights reserved.</p>
        </div>
    </div>
    """

    # Using your existing helper function in a thread
    url = os.environ.get('EMAIL_URL')
    # Create and start the thread
    
    epayload = {
        "email": email,
        "subject": subject,
        "html_content": html_content
    }
    print(f"DEBUG: Payload created: {epayload}")
    
    # Send the request to Pipedream
    try:
        response = requests.post(url, json=epayload)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Error sending email: {e}")



######################################################################################################################
import threading
from django.core.mail import EmailMultiAlternatives
from django.utils.html import strip_tags

def send_styled_email_thread(subject, html_content, recipient_list):
    """Internal function to handle the actual sending in a background thread."""
    from_email = 'Eduke Team <noreply@eduke.com>'
    text_content = strip_tags(html_content)
    
    msg = EmailMultiAlternatives(subject, text_content, from_email, recipient_list)
    msg.attach_alternative(html_content, "text/html")
    msg.send()

def send_institution_welcome_email(institution_name, email, password):
    """Prepares content and starts a new thread."""
    subject = "Welcome to Eduke - Registration Successful"
    
    violet_primary = "#7c3aed"
    slate_bg = "#f8fafc"
    slate_text = "#1e293b"
    slate_muted = "#64748b"
    danger_red = "#ef4444"
    portal_url = f"{settings.BASE_URL}{reverse('user_portal')}"

    html_content = f"""
    <div style="background-color: {slate_bg}; padding: 40px 10px; font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; line-height: 1.5;">
        <div style="max-width: 600px; margin: auto; background: #ffffff; border-radius: 12px; border: 1px solid #e2e8f0; overflow: hidden; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);">
            
            <div style="background-color: {violet_primary}; padding: 30px; text-align: center;">
                <h2 style="color: #ffffff; margin: 0; font-size: 24px; font-weight: 700; letter-spacing: -0.025em;">Welcome to Eduke</h2>
                <p style="color: #ddd6fe; margin: 5px 0 0 0; font-size: 14px;">Institutional Account Activated</p>
            </div>
            
            <div style="padding: 30px;">
                <p style="color: {slate_text}; font-size: 16px; margin-bottom: 25px;">
                    Hello <strong>{institution_name}</strong>, <br>
                    Your registration has been processed successfully. Below are your administrative login credentials.
                </p>

                <div style="margin-bottom: 25px;">
                    <label style="color: {slate_muted}; font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em;">Access Details</label>
                    <div style="margin-top: 10px; padding: 20px; background-color: #f1f5f9; border-left: 4px solid {violet_primary}; border-radius: 4px;">
                        <p style="margin: 0 0 8px 0; color: {slate_text};"><strong>Portal Email:</strong> {email}</p>
                        <p style="margin: 0; color: {slate_text};"><strong>Temporary Password:</strong> <code style="background: #e2e8f0; padding: 2px 6px; border-radius: 4px; font-family: monospace;">{password}</code></p>
                    </div>
                </div>

                <div style="margin-bottom: 30px;">
                    <label style="color: {danger_red}; font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em;">Security Requirement</label>
                    <div style="margin-top: 10px; padding: 15px; background-color: #fef2f2; border: 1px solid #fee2e2; border-left: 4px solid {danger_red}; border-radius: 4px; color: {danger_red}; font-size: 14px; font-weight: 500;">
                        For your security, you must change this password immediately after your first login.
                    </div>
                </div>

                <div style="text-align: center; margin-top: 35px;">
                    <a href="{portal_url}" style="background-color: {violet_primary}; color: #ffffff; padding: 14px 28px; border-radius: 8px; text-decoration: none; font-weight: 600; display: inline-block; font-size: 15px;">
                        Login to Dashboard
                    </a>
                </div>
            </div>

            <div style="background-color: {slate_bg}; padding: 20px; text-align: center; border-top: 1px solid #e2e8f0;">
                <p style="font-size: 12px; color: {slate_muted}; margin: 0;">
                    Regards, <br>
                    <strong>The Eduke Team</strong>
                </p>
                <p style="font-size: 11px; color: #94a3b8; margin-top: 10px;">
                    &copy; 2026 Eduke Platform. All rights reserved.
                </p>
            </div>
        </div>
    </div>
    """
    
    url = os.environ.get('EMAIL_URL')
    # Create and start the thread
    
    epayload = {
        "email": email,
        "subject": subject,
        "html_content": html_content
    }
    print(f"DEBUG: Payload created: {epayload}")
    
    # Send the request to Pipedream
    try:
        response = requests.post(url, json=epayload)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Error sending email: {e}")
    # email_thread = threading.Thread(
    #     target=send_styled_email_thread, 
    #     args=(subject, html_content, [email])
    # )
    # email_thread.start()

def register_institution(request):
    if request.method == 'POST':
        form = InstitutionRegisterForm(request.POST)
        
        if form.is_valid():
            institution_name = form.cleaned_data['institution_name']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            institution_abbrevation = form.cleaned_data['institution_abbreviation']

            # Check if email already exists
            if Institution.objects.filter(email=email).exists():
                messages.error(request, 'An institution with this email already exists.', extra_tags="reg_error_unique")
                # No redirect; stays on page and shows error
            else:
                try:
                    # Save the record with encrypted password
                    Institution.objects.create(
                        institution_name=institution_name,
                        email=email,
                        password=encrypt_password(password),  # Encrypt the password before saving
                        institution_abbreviation = institution_abbrevation
                    )

                    # Send the styled HTML email via helper function
                    send_institution_welcome_email(institution_name, email, password)

                    messages.success(
                        request, 
                        'Institution registered successfully! A confirmation email has been sent.', 
                        extra_tags='should_redirect'
                    )
                    # Optional: redirect to login or clear form

                except Exception as e:
                    print(f"DEBUG Error: {e}")
                    messages.error(request, 'An error occurred during registration. Please try again.')
        else:
            messages.error(request, 'Please correct the errors in the form.')
            
    else:
        form = InstitutionRegisterForm()

    return render(request, 'registration/institution_register.html', {'form': form})

# Institution Login
def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            # Authenticate the institution  
            try:
                institution = Institution.objects.get(email=email)
                
                # Check if the provided password matches the encrypted password
                if not verify_password(password, institution.password):
                    raise Institution.DoesNotExist('Invalid credentials')

                # Set session variables
                request.session['institution_id'] = institution.institution_id
                request.session['institution_name'] = institution.institution_name

                # Log the access
                try:
                    client_ip = get_client_ip(request)
                    AuditLog.objects.create(
                        institution=institution,
                        action='login',
                        ip_address=client_ip,
                        user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
                    )
                except Exception as e:
                    print(f"Logging failed: {e}")

                # Redirect to the admin dashboard
                return redirect('admin_dashboard')  # Adjust the URL name accordingly
            except Institution.DoesNotExist:
                messages.error(request, 'Invalid email or password.', extra_tags="login_error")
        else:
            messages.error(request, 'Please correct the errors.', extra_tags="login_error")
    else:
        form = LoginForm()

    return render(request, 'login.html', {'form': form})


# Admin Logout
def logout(request):
    # Handle institution logout
    institution_id = request.session.get('institution_id')
    if institution_id:
        try:
            institution = Institution.objects.get(institution_id=institution_id)
            client_ip = get_client_ip(request)
            AuditLog.objects.create(
                institution=institution,
                action='logout',
                ip_address=client_ip,
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
            )
        except Exception as e:
            print(f"Institution logout logging failed: {e}")
    
    # Handle student logout
    student_id = request.session.get('student_id')
    if student_id:
        try:
            from .models import Students, Classes
            student = Students.objects.get(id=student_id)
            # Get institution through student's class
            class_obj = Classes.objects.get(id=student.class_obj_id)
            institution = Institution.objects.get(institution_id=class_obj.institution_id)
            client_ip = get_client_ip(request)
            AuditLog.objects.create(
                user=student.user,
                institution=institution,
                action='logout',
                ip_address=client_ip,
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
                details=f'Student logout: {student.name}',
                related_object_id=student_id,
                related_object_type='Students'
            )
        except Exception as e:
            print(f"Student logout logging failed: {e}")
    
    # Handle class head logout
    class_id = request.session.get('class_id')
    if class_id:
        try:
            from .models import Classes
            class_obj = Classes.objects.get(id=class_id)
            institution = Institution.objects.get(institution_id=class_obj.institution_id)
            client_ip = get_client_ip(request)
            AuditLog.objects.create(
                user=class_obj.user,
                institution=institution,
                action='logout',
                ip_address=client_ip,
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
                details=f'Class Head logout: {class_obj.class_name}',
                related_object_id=class_id,
                related_object_type='Classes'
            )
        except Exception as e:
            print(f"Class head logout logging failed: {e}")
    
    # Handle subject head logout
    subject_id = request.session.get('subject_id')
    if subject_id:
        try:
            from .models import Subjects
            subject_obj = Subjects.objects.get(id=subject_id)
            # Get institution through subject's class
            class_obj = Classes.objects.get(id=subject_obj.class_obj_id)
            institution = Institution.objects.get(institution_id=class_obj.institution_id)
            client_ip = get_client_ip(request)
            AuditLog.objects.create(
                user=subject_obj.user,
                institution=institution,
                action='logout',
                ip_address=client_ip,
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
                details=f'Subject Head logout: {subject_obj.subject_name}',
                related_object_id=subject_id,
                related_object_type='Subjects'
            )
        except Exception as e:
            print(f"Subject head logout logging failed: {e}")
    
    # Handle parent logout
    parent_id = request.session.get('parent_id')
    if parent_id:
        try:
            from .models import Parents, Students
            parent_obj = Parents.objects.get(id=parent_id)
            # Get institution through parent's student's class
            student_obj = Students.objects.get(id=parent_obj.student_id)
            class_obj = Classes.objects.get(id=student_obj.class_obj_id)
            institution = Institution.objects.get(institution_id=class_obj.institution_id)
            client_ip = get_client_ip(request)
            AuditLog.objects.create(
                user=parent_obj.user,
                institution=institution,
                action='logout',
                ip_address=client_ip,
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
                details=f'Parent logout: {parent_obj.name or "Parent of " + str(student_obj.roll_no)}',
                related_object_id=parent_id,
                related_object_type='Parents'
            )
        except Exception as e:
            print(f"Parent logout logging failed: {e}")
    
    # Clear session and logout
    request.session.flush()
    return redirect('/')

######################################################################################################################

# Admin Dashboard view
def admin_dashboard(request):
    # 1. Session Security Check
    if 'institution_id' not in request.session:
        messages.error(request, 'Please login to access the dashboard')
        return redirect('login')

    institution_id = request.session['institution_id']
    
    
    try:
        # 2. Fetch the institution object
        # Note: Using .get() returns a model instance
        inst_obj = Institution.objects.get(institution_id=institution_id)
        
        # We pack this into a dictionary so the template can access {{ institution.name }}
        institution_data = {
            'institution_name': inst_obj.institution_name,
            'institution_abbreviation': inst_obj.institution_abbreviation,
            'email': inst_obj.email,
        }
    except Institution.DoesNotExist:
        messages.error(request, "Institution not found.")
        return redirect('login')
    
    # 3. Efficient Queries for Counts
    # We define the queryset for classes once to reuse it
    institution_classes = Classes.objects.filter(institution_id=institution_id)
    class_ids = institution_classes.values_list('id', flat=True)

    total_classes = institution_classes.count()
    
    student_count = Students.objects.filter(class_obj_id__in=class_ids).count()
    
    subject_count = Subjects.objects.filter(class_obj_id__in=class_ids).count()

    # Calculate Total Users (Class Users + Student Users)
    class_user_count = Users.objects.filter(id__in=institution_classes.values('user')).count()
    student_user_count = Users.objects.filter(id__in=Students.objects.filter(class_obj_id__in=class_ids).values('user')).count()
    total_users = class_user_count + student_user_count

    # 4. Correct Context Structure
    context = {
        'institution': institution_data, # This fixes your HTML issue
        'total_users': total_users,
        'student_count': student_count,
        'total_classes': total_classes,
        'subject_count': subject_count, 
    }

    return render(request, 'admin/admin_dashboard.html', context)

def admin_reports(request):
    if 'institution_id' not in request.session:
        return redirect('login')
    
    inst_id = request.session['institution_id']
    institution = get_object_or_404(Institution, institution_id=inst_id)
    classes = Classes.objects.filter(institution=institution)
    
    context = {
        'institution': institution,
        'classes': classes,
    }
    return render(request, 'admin/admin_reports_hub.html', context)

def admin_report_attendance(request, class_id):
    if 'institution_id' not in request.session:
        return redirect('login')
    
    inst_id = request.session['institution_id']
    institution = get_object_or_404(Institution, institution_id=inst_id)
    selected_class = get_object_or_404(Classes, id=class_id, institution=institution)
    
    # Get all students in this class
    students = Students.objects.filter(class_obj=selected_class)
    
    # Fetch attendance for these students
    attendance_records = Attendance.objects.filter(student__in=students).select_related('student', 'subject').order_by('-attendance_date')
    
    context = {
        'institution': institution,
        'selected_class': selected_class,
        'records': attendance_records,
        'report_type': 'Attendance',
    }
    return render(request, 'admin/admin_report_view.html', context)



def admin_report_quiz(request, class_id):
    if 'institution_id' not in request.session:
        return redirect('login')
    
    inst_id = request.session['institution_id']
    institution = get_object_or_404(Institution, institution_id=inst_id)
    selected_class = get_object_or_404(Classes, id=class_id, institution=institution)
    
    # Fetch all students in the class
    all_students = Students.objects.filter(class_obj=selected_class).order_by('name')
    
    # Fetch all quizzes assigned to this class - chronological order
    class_quizzes = Quizzes.objects.filter(class_obj=selected_class).order_by('id')
    
    report_records = []

    for quiz in class_quizzes:
        for student in all_students:
            # 1. Get all questions for this quiz
            questions = QuizQuestions.objects.filter(quiz=quiz)
            total_questions = questions.count()
            
            # 2. Check for student responses to any question in this quiz
            responses = QuizResponse.objects.filter(
                student=student, 
                question__in=questions
            )
            
            if responses.exists():
                # Calculate marks: count responses where student choice matches correct_option
                score = 0
                for resp in responses:
                    if resp.student_response == resp.question.correct_option:
                        score += 1
                
                # Calculate marks out of 100
                marks_out_of_100 = round((score / total_questions) * 100, 2) if total_questions > 0 else 0
                
                # Determine status: Passed if score is >= half of total questions, otherwise Failed
                if score >= (total_questions / 2) and total_questions > 0:
                    quiz_status = 'Pass'
                    raw_status = 'pass'
                else:
                    quiz_status = 'Failed'
                    raw_status = 'failed'
                
                report_records.append({
                    'student_name': student.name,
                    'quiz_name': quiz.name,
                    'subject': quiz.subject.subject_name,
                    'status': quiz_status,
                    'marks': f"{marks_out_of_100}/100",
                    'correct_answers': f"{score}/{total_questions}",
                    'raw_status': raw_status
                })
            else:
                # No responses found = Not Attended
                report_records.append({
                    'student_name': student.name,
                    'quiz_name': quiz.name,
                    'subject': quiz.subject.subject_name,
                    'status': 'Not Attended',
                    'marks': "0/100",
                    'correct_answers': f"0/{total_questions}",
                    'raw_status': 'not attended'
                })

    context = {
        'institution': institution,
        'selected_class': selected_class,
        'records': report_records,
        'all_students': all_students,
        'all_quizzes': class_quizzes,
        'report_type': 'Quiz Results',
    }
    return render(request, 'admin/admin_report_view.html', context)



def admin_report_logs(request, class_id):
    if 'institution_id' not in request.session:
        return redirect('login')
    
    inst_id = request.session['institution_id']
    institution = get_object_or_404(Institution, institution_id=inst_id)
    selected_class = get_object_or_404(Classes, id=class_id, institution=institution)
    
    # Get users related to this class
    class_head_user_id = selected_class.user_id
    student_user_ids = Students.objects.filter(class_obj=selected_class).values_list('user_id', flat=True)
    
    all_target_users = list(student_user_ids)
    all_target_users.append(class_head_user_id)
    
    # Fetch AuditLogs for these users
    logs = AuditLog.objects.filter(user_id__in=all_target_users).select_related('user').order_by('-timestamp')
    
    context = {
        'institution': institution,
        'selected_class': selected_class,
        'records': logs,
        'report_type': 'System Logs',
    }
    return render(request, 'admin/admin_report_view.html', context)
# def admin_classes(request):
#     print("🟢 View Loaded: admin_classes()")  

#     if 'institution_id' not in request.session:
#         messages.error(request, 'Please log in to access this page.')
#         print("🔴 User not logged in. Redirecting to login.")
#         return redirect('login')

#     institution_id = request.session['institution_id']
#     print(f"🟢 Institution ID: {institution_id}")

#     try:
#         institution = Institution.objects.get(institution_id=institution_id)
#         print(f"🟢 Institution Found: {institution.institution_name}")
#     except Institution.DoesNotExist:
#         messages.error(request, "Institution not found.")
#         print("🔴 Institution not found. Redirecting to login.")
#         return redirect('login')

#     classes = Classes.objects.filter(institution_id=institution.institution_id)
#     print(f"🟢 Retrieved {classes.count()} classes.")
#     for class_instance in classes:
#         print(f"   Class ID: {class_instance.id}, Name: {class_instance.class_name}, Head: {class_instance.class_head}, Email: {class_instance.email}")

#     if request.method == 'POST':
#         print("🟢 POST Request Received.")
#         print(f"🔍 Request Data: {request.POST}")

#         print("🟢 Handling New Class Addition")
#         form = AddClassForm(request.POST)

#         if form.is_valid():
#             class_name = form.cleaned_data['class_name']
#             class_head = form.cleaned_data['class_head']
#             email = form.cleaned_data['email']
#             password = form.cleaned_data['password']

#             # 🔥 Check if class name already exists for this institution
#             if Classes.objects.filter(class_name=class_name, institution_id=institution.institution_id).exists():
#                 messages.error(request, "This class name already exists for this institution. Please choose a different name.")
#                 print(f"🔴 Duplicate class name '{class_name}' detected for Institution ID {institution.institution_id}.")
#                 return redirect('admin_classes')

#             with transaction.atomic():
#                 print(f"🟢 New Class Data - Name: {class_name}, Head: {class_head}, Email: {email}")

#                 user = Users.objects.create(role='class_head')
#                 print("🟢 New User Created with ID:", user.id)

#                 new_class = Classes(
#                     class_name=class_name,
#                     class_head=class_head,
#                     email=email,
#                     password=password,
#                     institution=institution,
#                     user=user
#                 )
#                 new_class.save()
#                 print(f"🟢 New Class Added with ID: {new_class.id}. Data: Name: {class_name}, Head: {class_head}, Email: {email}")

#                 send_account_creation_email(email, password, "class_head", class_head, institution.email)
#                 print("🟢 Account creation email sent.")

#                 messages.success(request, 'New class added successfully!')
#                 return redirect('admin_classes')
#         else:
#             messages.error(request, 'Error adding class. Please try again.')
#             print("🔴 Error: Form validation failed.")

#     else:
#         print("🟢 Rendering Page with GET Request.")
#         form = AddClassForm()

#     context = {
#         'institution_name': institution.institution_name,
#         'classes': classes,
#         'form': form,
#     }

#     print(f"🟢 Sending context to template: {context}")
#     print("🟢 Rendering admin/admin_classes.html")
#     return render(request, 'admin/admin_classes.html', context)
def admin_classes(request):
    print("🟢 View Loaded: admin_classes()")

    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        print("🔴 User not logged in. Redirecting to login.")
        return redirect('login')

    institution_id = request.session['institution_id']
    print(f"🟢 Institution ID: {institution_id}")

    try:
        institution = Institution.objects.get(institution_id=institution_id)
        print(f"🟢 Institution Found: {institution.institution_name}")
    except Institution.DoesNotExist:
        messages.error(request, "Institution not found.")
        print("🔴 Institution not found. Redirecting to login.")
        return redirect('login')

    classes = Classes.objects.filter(institution=institution)

    if request.method == 'POST':
        form = AddClassForm(request.POST)

        print(f"\n🔵 POST request received")
        print(f"   - Form is valid: {form.is_valid()}")
        
        if not form.is_valid():
            print(f"   - Form errors: {form.errors}")
            for field, errors in form.errors.items():
                for error in errors:
                    print(f"   - {field}: {error}")
                    messages.error(request, f"{field.title()}: {error}", extra_tags="form_error")
        else:
            print(f"   - Form data: {form.cleaned_data}")

        if form.is_valid():
            class_name = form.cleaned_data['class_name']
            class_abbr = form.cleaned_data.get('class_abbreviation')
            class_head = form.cleaned_data['class_head']
            email      = form.cleaned_data['email']
            password   = form.cleaned_data['password']

            if Classes.objects.filter(
                    class_name=class_name,
                    institution=institution
                ).exists():
                    messages.error(request, f"Class '{class_name}' already exists.")
                    return redirect('admin_classes')

            try:
                with transaction.atomic():
                    user = Users.objects.create(role='class_head')

                    new_class = Classes(
                        class_name=class_name,
                            class_abbreviation=class_abbr if class_abbr else None,
                            class_head=class_head,
                            email=email,
                            password=encrypt_password(password),  # Encrypt the password for display
                            institution=institution,
                            user=user
                        )
                    new_class.save()
                    
                    print(f"✅ Class created successfully: {new_class.id}")
                    print(f"   - Class Name: {class_name}")
                    print(f"   - Class Head: {class_head}")
                    print(f"   - Email: {email}")
                    print(f"   - Password Encrypted: {not new_class.password.startswith('pbkdf2_')}")
                    
                    send_account_creation_email(
                        email, password, "class_head", class_head, institution.email
                    )

                    messages.success(request, 'New class added successfully!')
                    return redirect('admin_classes')

            except Exception as e:
                print(f"❌ Error creating class: {str(e)}")
                print(f"   - Exception Type: {type(e).__name__}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error: {str(e)}")
                return redirect('admin_classes')

    else:
        form = AddClassForm()



    # Create the dictionary that the template {{ institution.xxx }} expects
    institution_data = {
        'institution_name': institution.institution_name,
        'institution_abbreviation': institution.institution_abbreviation,
    }

    # Decrypt passwords for display in the template
    classes_with_decrypted_passwords = []
    for cls in classes:
        cls.password = decrypt_password(cls.password)
        classes_with_decrypted_passwords.append(cls)

    context = {
        # This is the key that fixes your HTML display
        'institution': institution_data, 
        
        # Keeping these for any other parts of your template
        'institution_name': institution.institution_name,
        'institution_abbreviation': institution.institution_abbreviation,
        'classes': classes_with_decrypted_passwords,  # Use decrypted passwords
        'form': form,
    }

    return render(request, 'admin/admin_classes.html', context)

def admin_class_edit(request, class_id):

    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        print("🔴 User not logged in. Redirecting to login.")
        return redirect('login')

    try:
        class_obj = Classes.objects.get(id=class_id)

        if request.method == "POST":
            new_class_name = request.POST.get('class_name')
            new_class_abbr = request.POST.get('class_abbreviation')
            new_class_head = request.POST.get('class_head')
            new_email = request.POST.get('email')
            # FIX: Get the password from the POST data
            new_password = request.POST.get('password') 

            # Uniqueness check
            existing_class = Classes.objects.filter(
                institution_id=class_obj.institution_id,
                class_name=new_class_name
            ).exclude(id=class_obj.id).first()

            if existing_class:
                messages.error(request, "This class name already exists for the institution.")
                # ... (your existing error handling logic)
                return render(request, 'admin/admin_classes.html', context)

            # Save updates
            class_obj.class_name = new_class_name
            class_obj.class_abbreviation = new_class_abbr if new_class_abbr else None
            class_obj.class_head = new_class_head
            class_obj.email = new_email
            # FIX: Encrypt the new password before saving
            if new_password:
                class_obj.password = encrypt_password(new_password)
            
            class_obj.save()

            messages.success(request, "Class updated successfully!")
            return redirect('admin_classes')

        return redirect('admin_classes')

    except Classes.DoesNotExist:
        messages.error(request, "Class not found.")
        return redirect('admin_classes')
    
from django.db import connection
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages


def admin_class_detail(request, class_id):
    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        return redirect('login')

    institution_id = request.session['institution_id']

    with connection.cursor() as cursor:
        # 1. Fetch institution details
        cursor.execute("""
            SELECT institution_name, institution_abbreviation 
            FROM main_institution WHERE institution_id = %s
        """, [institution_id])
        inst_row = cursor.fetchone()
        if not inst_row:
            return redirect('login')
        
        institution_data = {
            'institution_name': inst_row[0],
            'institution_abbreviation': inst_row[1],
        }

        # 2. Fetch class details
        cursor.execute("""
            SELECT id, class_name, class_head, email 
            FROM main_classes WHERE id = %s
        """, [class_id])
        class_row = cursor.fetchone()
        if not class_row:
            messages.error(request, "Class not found.")
            return redirect('admin_students')

        class_obj = {
            'id': class_row[0],
            'class_name': class_row[1],
            'class_head': class_row[2],
            'email': class_row[3],
        }

        # 3. Dynamic Students List
        cursor.execute("SELECT id, roll_no, name, email FROM main_students WHERE class_obj_id = %s", [class_id])
        students = [{'id': row[0], 'roll_no': row[1], 'name': row[2], 'email': row[3]} for row in cursor.fetchall()]
        total_students = len(students)

        # 4. Dynamic Performance (Mastery) & Grade Distribution
        # We calculate the average based on actual quiz responses vs correct answers
        cursor.execute("""
            SELECT 
                COUNT(CASE WHEN score_pct >= 85 THEN 1 END) as distinction,
                COUNT(CASE WHEN score_pct >= 70 AND score_pct < 85 THEN 1 END) as merit,
                COUNT(CASE WHEN score_pct >= 50 AND score_pct < 70 THEN 1 END) as pass,
                COUNT(CASE WHEN score_pct < 50 THEN 1 END) as fail,
                AVG(score_pct) as avg_score
            FROM (
                SELECT 
                    (COUNT(CASE WHEN r.student_response = q.correct_option THEN 1 END) * 100.0 / NULLIF(COUNT(r.id), 0)) as score_pct
                FROM main_quizresponse r
                JOIN main_quizquestions q ON r.question_id = q.id
                WHERE r.student_id IN (SELECT id FROM main_students WHERE class_obj_id = %s)
                GROUP BY r.student_id
            ) as student_averages
        """, [class_id])
        grade_stats = cursor.fetchone()

        # 5. Dynamic Attendance Rate
        # Calculating the actual percentage from the attendance table for this class
        cursor.execute("""
            SELECT 
                COALESCE(COUNT(CASE WHEN status = 'present' THEN 1 END) * 100.0 / NULLIF(COUNT(*), 0), 0)
            FROM main_attendance 
            WHERE student_id IN (SELECT id FROM main_students WHERE class_obj_id = %s)
        """, [class_id])
        attendance_rate = round(cursor.fetchone()[0] or 0, 1)

        # 6. Quiz Count & Completion Rate
        cursor.execute("SELECT COUNT(*) FROM main_quizzes WHERE class_obj_id = %s", [class_id])
        quiz_count = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT COUNT(*) FROM main_quizresponse 
            WHERE student_id IN (SELECT id FROM main_students WHERE class_obj_id = %s)
        """, [class_id])
        total_completions = cursor.fetchone()[0] or 0

        possible_submissions = total_students * quiz_count
        completion_rate = round((total_completions / possible_submissions) * 100, 1) if possible_submissions > 0 else 0
        
        
        # 7. Get number of unique students who participated
        cursor.execute("""
            SELECT COUNT(DISTINCT student_id) 
            FROM main_quizresponse 
            WHERE student_id IN (SELECT id FROM main_students WHERE class_obj_id = %s)
        """, [class_id])
        participating_students_count = cursor.fetchone()[0] or 0

        # 3. Fetch students with individual submission rates
        cursor.execute("""
            SELECT 
                s.id, 
                s.roll_no, 
                s.name, 
                s.email,
                COUNT(r.id) as completed_quizzes,
                CASE 
                    WHEN %s > 0 THEN ROUND((COUNT(r.id) * 100.0) / %s, 1)
                    ELSE 0 
                END as individual_rate
            FROM main_students s
            LEFT JOIN main_quizresponse r ON s.id = r.student_id
            WHERE s.class_obj_id = %s
            GROUP BY s.id, s.roll_no, s.name, s.email
        """, [quiz_count, quiz_count, class_id])

        students = [
            {
                'id': row[0], 
                'roll_no': row[1], 
                'name': row[2], 
                'email': row[3],
                'completed': row[4],
                'rate': row[5]
            } for row in cursor.fetchall()
        ]

    # Final Context Mapping
    context = {
        'institution': institution_data,
        'class_obj': class_obj,
        'students': students,
        'student_count': total_students,
        'total_completions':total_completions,        
        
        # Performance Analytics
        'distinction_count': grade_stats[0] or 0,
        'merit_count': grade_stats[1] or 0,
        'pass_count': grade_stats[2] or 0,
        'fail_count': grade_stats[3] or 0,
        'avg_score': round(grade_stats[4] or 0, 1),
        'students_at_risk': grade_stats[3] or 0,
        'participating_students_count':participating_students_count,
        
        # Activity Metrics
        'attendance_rate': attendance_rate,
        'quiz_count': quiz_count,
        'completion_rate': completion_rate,
        
        # Ratings (Derived from performance/attendance)
        'participation_rating': min(5, int(completion_rate / 20)),
        'academic_rating': min(5, int((grade_stats[4] or 0) / 20)),
    }

    return render(request, 'admin/admin_class_detail.html', context)





# def admin_subjects(request):
#     # Check if the user is logged in
#     if 'institution_id' not in request.session:
#         messages.error(request, 'Please log in to access this page.')
#         return redirect('login')
    
#     # Get the institution_id from the session
#     institution_id = request.session['institution_id']
#     print(f"DEBUG: Institution ID from session - {institution_id}")

#     try:
#         # Fetch the institution based on the session institution_id
#         institution = Institution.objects.get(institution_id=institution_id)
#         print(f"DEBUG: Institution found - {institution.institution_name}")
#     except Institution.DoesNotExist:
#         messages.error(request, "Institution not found.")
#         return redirect('login')

#     # Fetch all classes for the dropdown
#     classes = Classes.objects.filter(institution_id=institution_id)
#     print(f"DEBUG: Classes retrieved - {list(classes.values('id', 'class_name'))}")

#     # Fetch subjects using ORM
#     subjects = Subjects.objects.filter(class_obj__institution_id=institution_id).select_related('class_obj')
#     print(f"DEBUG: Retrieved subjects - {list(subjects.values('id', 'subject_name', 'subject_head', 'email', 'class_obj__class_name'))}")

#     if request.method == "POST":
#         form = AddSubjectForm(request.POST, institution_id=institution_id)
#         if form.is_valid():
#             subject_name = form.cleaned_data['subject_name']
#             subject_head = form.cleaned_data['subject_head']
#             email = form.cleaned_data['email']
#             password = form.cleaned_data['password']
#             class_obj = form.cleaned_data['class_obj']

#             print(f"DEBUG: Form Data - Subject: {subject_name}, Head: {subject_head}, Email: {email}, Class ID: {class_obj.id}")

#             # Check if subject name already exists under the same class
#             if Subjects.objects.filter(subject_name=subject_name, class_obj=class_obj).exists():
#                 messages.error(request, "This subject name already exists for the selected class. Please choose a different name.")
#                 print(f"🔴 Duplicate subject '{subject_name}' detected for Class ID {class_obj.id}.")
#                 return redirect('admin_subjects')

#             try:
#                 # Step 1: Create user
#                 user = Users.objects.create(role='subject_head')

#                 # Step 2: Save subject
#                 subject = form.save(commit=False)
#                 subject.user = user
#                 subject.save()

#                 print(f"DEBUG: Inserted subject - ID: {subject.id}, Class ID: {class_obj.id}")

#                 # Step 3: Send Account Creation Email
#                 send_account_creation_email(email, password, "subject_head", subject_head, institution.email)

#                 messages.success(request, "Subject added successfully!")
#                 return redirect('admin_subjects')

#             except Exception as e:
#                 print(f"ERROR: {e}")
#                 messages.error(request, f"An error occurred: {e}")
#         else:
#             print("DEBUG: Form is invalid", form.errors)
#     else:
#         form = AddSubjectForm(institution_id=institution_id)

#     context = {
#         'institution': institution,
#         'form': form,
#         'classes': classes,
#         'subjects': subjects  
#     }

#     return render(request, 'admin/admin_subjects.html', context)


def admin_subjects(request):
    """
    Admin view for managing subjects:
    - Manual single subject creation
    """
    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        return redirect('login')

    institution_id = request.session['institution_id']

    try:
        institution = Institution.objects.get(institution_id=institution_id)
    except Institution.DoesNotExist:
        messages.error(request, "Institution not found.")
        return redirect('login')

    # All classes of this institution (used for form dropdown)
    classes = Classes.objects.filter(institution=institution)

    # All existing subjects (for display)
    subjects = Subjects.objects.filter(
        class_obj__institution=institution
    ).select_related('class_obj', 'user')
    
    # Decrypt passwords for display in template
    for subject in subjects:
        subject.password = decrypt_password(subject.password)

    form = AddSubjectForm(institution_id=institution_id)

    if request.method == "POST":
        # ────────────────────────────────────────────────
        #             MANUAL SINGLE SUBJECT
        # ────────────────────────────────────────────────
        form = AddSubjectForm(request.POST, institution_id=institution_id)

        if form.is_valid():

            subject_name = form.cleaned_data['subject_name']
            subject_head = form.cleaned_data['subject_head']
            email        = form.cleaned_data['email']
            password     = form.cleaned_data['password']
            class_obj    = form.cleaned_data['class_obj']

            if Subjects.objects.filter(
                subject_name=subject_name,
                class_obj=class_obj
            ).exists():
                messages.error(request,
                    f"Subject '{subject_name}' already exists in class {class_obj}."
                )
                return redirect('admin_subjects')

            try:
                with transaction.atomic():
                    user = Users.objects.create(role='subject_head')

                    subject = form.save(commit=False)
                    subject.user = user
                    # Encrypt the password before saving
                    subject.password = encrypt_password(password)
                    subject.save()

                    try:
                        send_account_creation_email(
                            email, password, "subject_head",
                            subject_head, institution.email
                        )
                    except Exception as mail_err:
                        print(f"Welcome email failed: {mail_err}")

                messages.success(request, "Subject created successfully.")
                return redirect('admin_subjects')

            except Exception as e:
                messages.error(request, f"Error creating subject: {str(e)}")

        else:
            messages.error(request, "Please correct the errors below.")

    # GET or failed POST → show form + list
    context = {
        'institution': institution,
        'form': form,
        'classes': classes,
        'subjects': subjects,
    }

    return render(request, 'admin/admin_subjects.html', context)

# def admin_subjects(request):
#     if 'institution_id' not in request.session:
#         messages.error(request, 'Please log in to access this page.')
#         return redirect('login')

#     institution_id = request.session['institution_id']
#     institution = Institution.objects.get(institution_id=institution_id)

#     classes = Classes.objects.filter(institution_id=institution_id)
#     subjects = Subjects.objects.filter(class_obj__institution=institution).select_related('class_obj')

#     if request.method == "POST":
#         if request.POST.get("bulk_upload") == "true":  # Bulk upload
#             file = request.FILES.get('file')
#             if not file:
#                 messages.error(request, "No file uploaded.")
#                 return redirect('admin_subjects')

#             # Read Excel/CSV
#             try:
#                 if file.name.endswith('.csv'):
#                     df = pd.read_csv(file)
#                 elif file.name.endswith(('.xls', '.xlsx')):
#                     df = pd.read_excel(file)
#                 else:
#                     messages.error(request, "Unsupported file format. Use CSV or XLSX.")
#                     return redirect('admin_subjects')

#                 # Expected columns (case-insensitive)
#                 expected_cols = {'subject_name', 'subject_head', 'email', 'password', 'class_name'}
#                 df.columns = df.columns.str.strip().str.lower()
#                 if not expected_cols.issubset(df.columns):
#                     missing = expected_cols - set(df.columns)
#                     messages.error(request, f"Missing columns: {', '.join(missing)}")
#                     return redirect('admin_subjects')

#                 success_count = 0
#                 errors = []

#                 with transaction.atomic():
#                     for idx, row in df.iterrows():
#                         try:
#                             # Get class by name (case-insensitive)
#                             class_name = str(row['class_name']).strip()
#                             class_obj = Classes.objects.filter(
#                                 institution=institution,
#                                 class_name__iexact=class_name
#                             ).first()

#                             if not class_obj:
#                                 errors.append(f"Row {idx+2}: Class '{class_name}' not found.")
#                                 continue

#                             # Check for duplicate subject in same class
#                             if Subjects.objects.filter(
#                                 subject_name__iexact=row['subject_name'].strip(),
#                                 class_obj=class_obj
#                             ).exists():
#                                 errors.append(f"Row {idx+2}: Subject '{row['subject_name']}' already exists in class '{class_name}'.")
#                                 continue

#                             # Check email uniqueness
#                             email = str(row['email']).strip()
#                             if Subjects.objects.filter(email=email).exists():
#                                 errors.append(f"Row {idx+2}: Email '{email}' already in use.")
#                                 continue

#                             # Create user
#                             user = Users.objects.create(role='subject_head')

#                             # Create subject
#                             subject = Subjects.objects.create(
#                                 subject_name=row['subject_name'].strip(),
#                                 subject_head=row['subject_head'].strip(),
#                                 email=email,
#                                 password=str(row['password']).strip(),  # Consider hashing later
#                                 class_obj=class_obj,
#                                 user=user
#                             )

#                             success_count += 1

#                         except Exception as e:
#                             errors.append(f"Row {idx+2}: {str(e)}")

#                 if success_count > 0:
#                     messages.success(request, f"{success_count} subjects added successfully!")

#                 if errors:
#                     for err in errors:
#                         messages.error(request, err)

#             except Exception as e:
#                 messages.error(request, f"File processing error: {str(e)}")

#             return redirect('admin_subjects')

#         else:  # Manual form submission
#             form = AddSubjectForm(request.POST, institution_id=institution_id)
#             if form.is_valid():
#                 subject_name = form.cleaned_data['subject_name']
#                 class_obj = form.cleaned_data['class_obj']

#                 # Duplicate check (same as before)
#                 if Subjects.objects.filter(subject_name=subject_name, class_obj=class_obj).exists():
#                     messages.error(request, "This subject name already exists for the selected class.")
#                     return redirect('admin_subjects')

#                 try:
#                     with transaction.atomic():
#                         user = Users.objects.create(role='subject_head')
#                         subject = form.save(commit=False)
#                         subject.user = user
#                         subject.save()

                    

#                     messages.success(request, "Subject added successfully!")
#                     return redirect('admin_subjects')

#                 except IntegrityError:
#                     messages.error(request, "Email already in use or database conflict.")
#                 except Exception as e:
#                     messages.error(request, f"Error: {str(e)}")
#             else:
#                 messages.error(request, "Please correct the form errors.")

#     else:
#         form = AddSubjectForm(institution_id=institution_id)

#     context = {
#         'institution': institution,
#         'form': form,
#         'classes': classes,
#         'subjects': subjects
#     }

#     return render(request, 'admin/admin_subjects.html', context)



def admin_subject_edit(request, subject_id):

    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        print("🔴 User not logged in. Redirecting to login.")
        return redirect('login')
    try:
        subject_obj = Subjects.objects.get(id=subject_id)
    except Subjects.DoesNotExist:
        messages.error(request, "Subject not found.")
        return redirect('admin_subjects')

    if request.method == "POST":
        new_subject_name = request.POST.get('subject_name')
        new_subject_head = request.POST.get('subject_head')
        new_email = request.POST.get('email')
        # FIX 1: Capture the password from the POST request
        new_password = request.POST.get('password')

        duplicate = Subjects.objects.filter(
            class_obj=subject_obj.class_obj,
            subject_name=new_subject_name
        ).exclude(id=subject_obj.id).exists()

        if duplicate:
            messages.error(request, "This subject name already exists under the same class.")
        else:
            subject_obj.subject_name = new_subject_name
            subject_obj.subject_head = new_subject_head
            subject_obj.email = new_email
            # FIX 2: Encrypt the new password before saving
            if new_password:
                subject_obj.password = encrypt_password(new_password)
            
            subject_obj.save()
            messages.success(request, "Subject updated successfully!")
            # Use redirect instead of render to ensure fresh data
            return redirect('admin_subjects')

    # Reload context for the render (only if not redirected)
    institution_id = request.session.get('institution_id')
    institution = Institution.objects.get(institution_id=institution_id)
    classes = Classes.objects.filter(institution_id=institution_id).values('id', 'class_name')

    with connection.cursor() as cursor:
        # FIX 3: Add ms.password to your SELECT statement
        cursor.execute("""
            SELECT 
                ms.id, ms.subject_name, ms.subject_head, ms.email, 
                c.class_name, ms.password
            FROM main_subjects ms
            JOIN main_classes c ON ms.class_obj_id = c.id
            WHERE c.institution_id = %s
        """, [institution_id])
        subjects = cursor.fetchall()

    from main.forms import AddSubjectForm
    form = AddSubjectForm()

    context = {
        'institution_name': institution.institution_name,
        'form': form,
        'classes': classes,
        'subjects': subjects
    }
    return render(request, 'admin/admin_subjects.html', context)
def admin_subject_detail(request, subject_id):

    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        print("🔴 User not logged in. Redirecting to login.")
        return redirect('login')
    print(f"Received subject_id: {subject_id}, Type: {type(subject_id)}")

    # Validate subject_id
    if not subject_id or not str(subject_id).isdigit():
        print("Invalid subject_id, redirecting to admin_students.")
        messages.error(request, "Invalid subject ID.")
        return redirect('admin_students')

    # Check if the user is logged in
    if 'institution_id' not in request.session:
        print("User not logged in, redirecting to login page.")
        messages.error(request, 'Please log in to access this page.')
        return redirect('login')

    # Get the institution_id from the session
    institution_id = request.session['institution_id']
    print(f"Institution ID from session: {institution_id}")

    with connection.cursor() as cursor:
        # --- FIX START ---
        # Fetch institution details (Added abbreviation so the sidebar works)
        cursor.execute("SELECT institution_name, institution_abbreviation FROM main_institution WHERE institution_id = %s", [institution_id])
        institution_row = cursor.fetchone()
        
        if not institution_row:
            print("Institution not found, redirecting to login.")
            messages.error(request, "Institution not found.")
            return redirect('login')
        
        # Create the dictionary object the template expects ({{ institution.institution_name }})
        institution = {
            'institution_name': institution_row[0],
            'institution_abbreviation': institution_row[1]
        }
        print(f"Institution: {institution['institution_name']}")
        # --- FIX END ---

        # Fetch subject details
        cursor.execute("SELECT id, subject_name, subject_head, email FROM main_subjects WHERE id = %s", [subject_id])
        subject_row = cursor.fetchone()
        if not subject_row:
            print("Subject not found, redirecting to admin_students.")
            messages.error(request, "Subject not found.")
            return redirect('admin_students')

        subject_obj = {
            'id': subject_row[0],
            'subject_name': subject_row[1],
            'subject_head': subject_row[2],
            'email': subject_row[3],
        }
        print(f"Subject Name: {subject_obj['subject_name']}")

    # Prepare context for the template
    context = {
        'institution': institution,          # Needed for sidebar: {{ institution.institution_name }}
        'institution_name': institution['institution_name'], # Needed for main body: {{ institution_name }}
        'subject_obj': subject_obj,
    }

    return render(request, 'admin/admin_subject_detail.html', context)

def admin_students(request):
    # Check if the user is logged in
    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        return redirect('login')

    institution_id = request.session['institution_id']

    try:
        institution = Institution.objects.get(institution_id=institution_id)
    except Institution.DoesNotExist:
        messages.error(request, "Institution not found.")
        return redirect('login')

    classes = Classes.objects.filter(institution_id=institution_id)
    print(f"DEBUG: institution_id={institution_id}")
    print(f"DEBUG: Classes filter count={classes.count()}")
    for c in classes:
        print(f"DEBUG: Class ID={c.id}, Name={c.class_name}, Inst={c.institution_id}")

    if request.method == "POST":
        # Since bulk upload is now handled by its own view, we only handle manual entry here
        form = AddStudentForm(request.POST, institution_id=institution_id)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            parent_email = form.cleaned_data['parent_email']
            roll_no = form.cleaned_data['roll_no']
            password = form.cleaned_data['password']
            class_obj = form.cleaned_data['class_obj']

            # Duplicate check
            if Students.objects.filter(roll_no=roll_no, class_obj=class_obj).exists():
                messages.error(request, f"Roll No '{roll_no}' already exists for the selected class.")
                return redirect('admin_students')

            try:
                with transaction.atomic():
                    # Create a user for the student
                    u_student = Users.objects.create(role='student')
                    
                    # Create the student
                    student = Students.objects.create(
                        user=u_student,
                        name=name,
                        roll_no=roll_no,
                        email=email,
                        password=encrypt_password(password),  # Encrypt the password
                        class_obj=class_obj
                    )

                    # Create a user for the parent and link to student
                    u_parent = Users.objects.create(role='parent')
                    Parents.objects.create(
                        user=u_parent,
                        student=student,
                        password=encrypt_password(password), # Encrypt the password before saving
                        name=None, # Parent name can be added later
                        email=parent_email
                    )
                
                messages.success(request, "Student added successfully!")
                # Send email to student
                try:
                    send_account_creation_email(email, password, "student", name, institution.email, roll_no)
                except Exception as e:
                    messages.warning(request, f"Student added, but failed to send email: {e}")

                # Send email to parent
                try:
                    if parent_email:
                        # For parents, password is currently roll_no, and name is student name (can be customized if needed)
                        send_account_creation_email(parent_email, password, "parent", f"Parent of {name}", institution.email, roll_no)
                except Exception as e:
                    messages.warning(request, f"Student added, but failed to send email to parent: {e}")

                return redirect('admin_students')

            except IntegrityError:
                messages.error(request, "Email already in use or database conflict.")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
        else:
            messages.error(request, "Please correct the form errors.")

    else: # This 'else' handles GET requests
        form = AddStudentForm(institution_id=institution_id)

    # Fetch students for display
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                s.id, s.name AS student_name, s.roll_no, s.email, c.class_abbreviation, 
                c.class_head, s.class_obj_id, s.password, p.email AS parent_email
            FROM main_students s
            LEFT JOIN main_classes c ON s.class_obj_id = c.id
            LEFT JOIN main_parents p ON s.id = p.student_id
            WHERE c.institution_id = %s
        """, [institution_id])
        students_raw = cursor.fetchall()
        
        # Decrypt passwords for each student
        students = []
        for student in students_raw:
            student_list = list(student)
            # Decrypt password (index 7 in the tuple)
            student_list[7] = decrypt_password(student[7]) if student[7] else ""
            students.append(tuple(student_list))
        
        
        

    institution_data = {
        'institution_name': institution.institution_name,
        'institution_abbreviation': institution.institution_abbreviation,
    }

    context = {
        'institution': institution_data,
        'institution_id': institution_id,
        'form': form,
        'students': students,
        'classes': classes,
    }

    return render(request, 'admin/admin_students.html', context)

#-------------------------------------------------------------
from django.views.decorators.http import require_http_methods

from django.views.decorators.http import require_GET
from datetime import datetime
@require_http_methods(["GET"])
@require_GET
def generate_roll_number(request, class_id):
    if 'institution_id' not in request.session:
        return JsonResponse({'error': 'Not authenticated'}, status=401)

    institution_id = request.session['institution_id']

    try:
        class_obj = Classes.objects.get(id=class_id, institution=institution_id)

        current_year = datetime.now().year                     # 2026
        year_short = str(current_year)[-2:]                     # 26

        # Use the explicit abbreviation if provided, otherwise fallback to class name
        if class_obj.class_abbreviation:
            class_abbr = class_obj.class_abbreviation.replace(" ", "").upper()
        else:
            # Better class abbreviation (fallback)
            class_abbr = ''.join(word[0].upper() for word in class_obj.class_name.split())
            if not class_abbr:
                class_abbr = class_obj.class_name[:2].upper()

        # Limit length
        class_abbr = class_abbr[:8]

        prefix = f"{institution_id}S{class_abbr}{year_short}"

        # Find highest sequence number for this exact prefix this year
        existing_rolls = Students.objects.filter(
            class_obj=class_obj,
            roll_no__startswith=prefix
        ).values_list('roll_no', flat=True)

        max_seq = 0
        for roll in existing_rolls:
            try:
                seq_part = roll[-3:] 
                
                if seq_part.isdigit():
                    seq = int(seq_part)
                    if seq > max_seq:
                        max_seq = seq
            except (ValueError, IndexError):
                continue

        next_seq = max_seq + 1

        roll_number = f"{prefix}{next_seq:03d}"  # :03d = 001, 002, 003...

        return JsonResponse({
            'success': True,
            'roll_number': roll_number,
            'prefix': prefix,
            'sequence': next_seq
        })

    except Classes.DoesNotExist:
        return JsonResponse({'error': 'Class not found or access denied'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)

    #----------------------------------------------------------------





def admin_student_edit(request, student_id):

    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        print("🔴 User not logged in. Redirecting to login.")
        return redirect('login')
    institution_id = request.session.get('institution_id')
    institution = Institution.objects.get(institution_id=institution_id)
    
    try:
        student_obj = Students.objects.get(id=student_id)
        parent_obj = Parents.objects.filter(student=student_obj).first()

        if request.method == "POST":
            form = AddStudentForm(request.POST, instance=student_obj, institution_id=institution_id)
            if form.is_valid():
                student_obj = form.save()
                
                # Update parent details
                parent_email = form.cleaned_data.get('parent_email')
                if parent_obj:
                    parent_obj.email = parent_email
                    parent_obj.password = student_obj.password # Keep synced
                    parent_obj.save()

                messages.success(request, "Student Metadata Synchronized Successfully.", extra_tags="students_update_success")
                return redirect('admin_students')
            else:
                for field, errors in form.errors.items():
                    messages.error(request, f"Update Failed | {field.title()}: {', '.join(errors)}")
                return redirect('admin_students')
        
        # If it's a GET request, we still redirect because edit is now in-page
        return redirect('admin_students')
    
    except Students.DoesNotExist:
        messages.error(request, "Student not found.")
        return redirect('admin_students')
    except Exception as e:
        messages.error(request, f"Error updating student: {e}")
        return redirect('admin_students')



from django.shortcuts import render, redirect
from django.db import connection
from django.contrib import messages

def admin_student_performance(request, student_id):
    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        return redirect('login')
    
    institution_id = request.session['institution_id']
    try:
        institution = Institution.objects.get(institution_id=institution_id)
    except Institution.DoesNotExist:
        messages.error(request, 'Institution data not found.')
        return redirect('login')
    
    # Validate student_id
    if not student_id or not str(student_id).isdigit():
        messages.error(request, "Invalid student ID.")
        return redirect('admin_students')

    try:
        with connection.cursor() as cursor:
            # 1. Fetch student details
            cursor.execute("""
                SELECT s.id, s.roll_no, s.name, s.email, s.class_obj_id, c.class_abbreviation
                FROM main_students s
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s
            """, [student_id])
            student_row = cursor.fetchone()
            
            if not student_row:
                messages.error(request, "Student not found.")
                return redirect('admin_students')

            student = {
                'id': student_row[0],
                'roll_no': student_row[1],
                'name': student_row[2],
                'email': student_row[3],
                'class_id': student_row[4],
                'class_name': student_row[5],
            }

            # 2. Fetch subjects for dropdown
            cursor.execute("SELECT id, subject_name FROM main_subjects WHERE class_obj_id = %s", [student["class_id"]])
            subjects = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]
           
        subject_id = request.GET.get('subject_id', None)
        
        # Default evaluation data
        evaluations = {
            "marks_percentage": 0, "attendance_percentage": 0,
            "study_time_rating": 0, "sleep_time_rating": 0,
            "class_participation_rating": 0, "academic_activity_rating": 0,
        }
        quiz_percentage = 0
        completion_rate = 0 # NEW: Dynamic Completion Rate
        total_completed = 0 # NEW: Raw count

        # 3. Fetch Evaluations if subject is selected
        if subject_id:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT marks_percentage, attendance_percentage,
                           study_time_rating, sleep_time_rating, 
                           class_participation_rating, academic_activity_rating
                    FROM main_studentevaluation
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()
                if row:
                    evaluations = {
                        "marks_percentage": round(row[0] or 0, 2),
                        "attendance_percentage": round(row[1] or 0, 2),
                        "study_time_rating": round(row[2] or 0, 2),
                        "sleep_time_rating": round(row[3] or 0, 2),
                        "class_participation_rating": round(row[4] or 0, 2),
                        "academic_activity_rating": round(row[5] or 0, 2),
                    }

                # 4. Quiz Performance (Accuracy)
                cursor.execute("""
                    SELECT COUNT(*) AS total_attempted,
                           SUM(CASE WHEN q.correct_option = r.student_response THEN 1 ELSE 0 END) AS correct_answers
                    FROM main_quizresponse r
                    JOIN main_quizquestions q ON r.question_id = q.id
                    JOIN main_quizzes mq ON q.quiz_id = mq.id
                    WHERE r.student_id = %s AND mq.subject_id = %s
                """, [student_id, subject_id])
                q_row = cursor.fetchone()
                if q_row and q_row[0] > 0:
                    quiz_percentage = round(((q_row[1] or 0) / q_row[0]) * 100, 2)

                # 5. Submission Rate (Total Quizzes vs This Student's Responses)
                # Count total quizzes for this subject
                cursor.execute("SELECT COUNT(*) FROM main_quizzes WHERE subject_id = %s", [subject_id])
                total_quizzes = cursor.fetchone()[0] or 0
                
                # Count unique quizzes this student has responded to
                cursor.execute("""
                    SELECT COUNT(DISTINCT mq.id) 
                    FROM main_quizresponse r
                    JOIN main_quizquestions q ON r.question_id = q.id
                    JOIN main_quizzes mq ON q.quiz_id = mq.id
                    WHERE r.student_id = %s AND mq.subject_id = %s
                """, [student_id, subject_id])
                total_completed = cursor.fetchone()[0] or 0
                
                if total_quizzes > 0:
                    completion_rate = round((total_completed / total_quizzes) * 100, 1)

        # --- PREPARE DATA FOR FRONTEND ---
        ratings = [
            ("Study Time", evaluations["study_time_rating"]),
            ("Sleep Cycle", evaluations["sleep_time_rating"]),
            ("Participation", evaluations["class_participation_rating"]),
            ("Activity", evaluations["academic_activity_rating"]),
        ]

        graph_data = [
            float(evaluations["marks_percentage"]),
            float(evaluations["attendance_percentage"]),
            float(evaluations["study_time_rating"]),
            float(evaluations["sleep_time_rating"]),
            float(evaluations["class_participation_rating"]),
            float(evaluations["academic_activity_rating"]),
            float(quiz_percentage)
        ]

    except Exception as e:
        print(f"[ERROR] {e}")
        messages.error(request, "An internal error occurred.")
        return redirect('admin_students')

    context = {
        'student': student,
        'subjects': subjects,
        'selected_subject_id': int(subject_id) if subject_id else None,
        'evaluations': evaluations,
        'quiz_percentage': quiz_percentage,
        'completion_rate': completion_rate,  # Use in your HTML card
        'total_completed': total_completed,  # Raw count for label
        'ratings': ratings,
        'graph_data': json.dumps(graph_data),
        'institution_name': institution.institution_name,
        'institution_abbreviation': institution.institution_abbreviation,
        'institution': institution,
    }

    return render(request, 'admin/admin_student_performance.html', context)

def admin_profile(request):
    # 1. Auth Check & Fetch
    inst_id = request.session.get('institution_id')
    if not inst_id:
        return redirect('login')
    
    institution = Institution.objects.get(institution_id=inst_id)
    
    # Decrypt password for display in template
    institution.password = decrypt_password(institution.password)

    # 2. Update logic (The "Short" Way)
    if request.method == 'POST':
        # List all fields you want to allow updates for
        fields = ['institution_name', 'institution_abbreviation', 'email']
        
        for field in fields:
            value = request.POST.get(field)
            if value: # Only update if the field isn't empty
                setattr(institution, field, value)
        
        # Handle password separately with encryption
        password = request.POST.get('password')
        if password:
            institution.password = encrypt_password(password)
        
        institution.save()
        messages.success(request, 'Profile updated successfully!')
        return redirect('admin_profile')

    return render(request, 'admin/admin_profile.html', {'institution': institution})


def delete_class(request, class_id):
    if request.method == "POST":
        class_obj = get_object_or_404(Classes, id=class_id)
        
        # Collect all associated User IDs to be deleted
        user_ids_to_delete = set()
        
        # 1. The class head user
        if class_obj.user_id:
            user_ids_to_delete.add(class_obj.user_id)
            
        # 2. Subject head users tied to this class
        subject_users = Subjects.objects.filter(class_obj=class_obj).values_list('user_id', flat=True)
        user_ids_to_delete.update(subject_users)
        
        # 3. Students tied to this class
        # Note: We fetch students before deleting the class because class_obj is SET_NULL
        students = Students.objects.filter(class_obj=class_obj)
        student_users = students.values_list('user_id', flat=True)
        user_ids_to_delete.update(student_users)
        
        # 4. Parents tied to these students
        parent_users = Parents.objects.filter(student__in=students).values_list('user_id', flat=True)
        user_ids_to_delete.update(parent_users)
        
        # Perform deletion
        # Deleting User records will CASCADE and delete corresponding Classes, Subjects, Students, and Parents
        if user_ids_to_delete:
            Users.objects.filter(id__in=user_ids_to_delete).delete()
        
        # Ensure the class record itself is gone (in case user was null or other issues)
        if Classes.objects.filter(id=class_id).exists():
            class_obj.delete()

        # Success message
        messages.success(request, "Class and associated users deleted successfully.", extra_tags="classes_success")

    return redirect('admin_classes')

def delete_subject(request, subject_id):
    if request.method == "POST":
        with connection.cursor() as cursor:
            # 1. Fetch user_id of the subject head
            cursor.execute("SELECT user_id FROM main_subjects WHERE id = %s", [subject_id])
            row = cursor.fetchone()
            user_id = row[0] if row else None

            # 2. Delete Student Evaluations linked to this subject
            cursor.execute("DELETE FROM main_studentevaluation WHERE subject_id = %s", [subject_id])

            # 3. Delete Chat Messages where this user is the sender (or receiver)
            if user_id:
                cursor.execute("DELETE FROM main_chat WHERE sender_id = %s", [user_id])
                # If your chat model also has a receiver_id, uncomment the next line:
                # cursor.execute("DELETE FROM main_chat WHERE receiver_id = %s", [user_id])

            # 4. Delete the Subject
            cursor.execute("DELETE FROM main_subjects WHERE id = %s", [subject_id])

            # 5. Finally, delete the User
            if user_id:
                cursor.execute("DELETE FROM main_users WHERE id = %s", [user_id])

        messages.success(request, "Subject, associated evaluations, chats, and user deleted successfully.", extra_tags="subjects_success")

    return redirect('admin_subjects')


def delete_student(request, student_id):
    if request.method == "POST":
        try:
            from django.db import transaction
            
            with transaction.atomic():
                # 1. Fetch student object
                student = Students.objects.get(id=student_id)
                student_user_id = student.user_id
                
                # 2. Fetch parent user_id (Parent is linked to student)
                parent_user_id = None
                try:
                    parent = Parents.objects.get(student=student)
                    parent_user_id = parent.user_id
                except Parents.DoesNotExist:
                    pass

                # 3. Delete student (Triggers CASCADE for Parents, Attendance, Marks, Evaluations, etc.)
                student.delete()

                # 4. Cleanup associated user records
                users_to_delete = [student_user_id]
                if parent_user_id:
                    users_to_delete.append(parent_user_id)
                
                Users.objects.filter(id__in=users_to_delete).delete()

            messages.success(request, "Student record successfully expunged from registry.", extra_tags="students_success")
        except Students.DoesNotExist:
            messages.error(request, "Student not found.")
        except Exception as e:
            messages.error(request, f"Failure in Deletion Protocol: {str(e)}")

    return redirect('admin_students')


######################################################################################################################


def upload_classes(request):
    """
    Dedicated view for bulk class uploads.
    Supports CSV, XLSX, and JSON formats.
    """
    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        return redirect('login')

    institution_id = request.session['institution_id']
    try:
        institution = Institution.objects.get(institution_id=institution_id)
    except Institution.DoesNotExist:
        messages.error(request, "Institution not found.")
        return redirect('login')

    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        filename = file.name.lower()
        success_count = 0
        errors = []

        try:
            # Read file based on extension
            if filename.endswith('.csv'):
                df = pd.read_csv(file)
            elif filename.endswith(('.xlsx', '.xls')):
                excel_buffer = BytesIO(file.read())
                df = pd.read_excel(excel_buffer)
            elif filename.endswith('.json'):
                json_data = json.load(file)
                df = pd.DataFrame(json_data)
            else:
                messages.error(request, "Only .csv, .xlsx, .json files are supported.")
                return redirect('admin_classes')

            # Normalize column names
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

            required = {'class_name', 'class_abbreviation', 'class_head', 'email', 'password'}
            found = set(df.columns)

            missing = required - found
            if missing:
                messages.error(request, f"Missing required columns: {', '.join(missing)}")
                return redirect('admin_classes')

            # Rename to internal names
            df = df.rename(columns={
                'class_name': 'class_name',
                'class_abbreviation': 'class_abbreviation',
                'class_head': 'class_head',
                'email': 'email',
                'password': 'password',
            })

            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    name   = str(row.get('class_name', '')).strip()
                    abbr   = str(row.get('class_abbreviation', '')).strip() if 'class_abbreviation' in df.columns else ''
                    head   = str(row.get('class_head', '')).strip()
                    email  = str(row.get('email', '')).strip()
                    pwd    = str(row.get('password', '')).strip()

                    row_errors = []

                    # Detailed missing field check
                    required_fields = ['class_name', 'class_abbreviation', 'class_head', 'email', 'password']
                    missing_row_fields = [f for f in required_fields if not str(row.get(f, '')).strip() or str(row.get(f, '')).strip().lower() == 'nan']
                    if missing_row_fields:
                        row_errors.append(f"Missing fields: {', '.join(missing_row_fields)}")

                    # Email format check
                    if email and '@' not in email:
                        row_errors.append(f"Invalid email: {email}")

                    # Duplicate email check (Pre-check to avoid transaction failure)
                    if email:
                        if Classes.objects.filter(email=email).exists() or Subjects.objects.filter(email=email).exists():
                            row_errors.append(f"Email '{email}' is already taken")

                    # Duplicate class name check
                    if name and Classes.objects.filter(class_name=name, institution=institution).exists():
                        row_errors.append(f"Class '{name}' already exists")

                    # Strict password check
                    pwd_requirements = []
                    if len(pwd) < 6:
                        pwd_requirements.append("at least 6 characters")
                    if not any(c.isdigit() for c in pwd):
                        pwd_requirements.append("at least one digit")
                    if not any(c.isupper() for c in pwd):
                        pwd_requirements.append("at least one uppercase letter")
                    if not any(c.islower() for c in pwd):
                        pwd_requirements.append("at least one lowercase letter")
                    if not any(c in "@$!%*%s&" for c in pwd):
                        pwd_requirements.append("at least one special character (@$!%*%s&)")
                    
                    if pwd_requirements:
                        row_errors.append(f"Weak password (requires {', '.join(pwd_requirements)})")

                    if row_errors:
                        errors.append(f"Row {row_num}: " + " | ".join(row_errors))
                        continue

                    # Atomic block per row to safely handle potential DB errors
                    with transaction.atomic():
                        user = Users.objects.create(role='class_head')

                        Classes.objects.create(
                            class_name=name,
                            class_abbreviation=abbr if abbr else None,
                            class_head=head,
                            email=email,
                            password=encrypt_password(pwd),  # Encrypt the password before saving
                            institution=institution,
                            user=user
                        )

                        # Optional: send email
                        try:
                            send_account_creation_email(email, pwd, "class_head", head, institution.email)
                        except Exception:
                            pass

                        success_count += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: Unexpected error: {str(e)}")

            if success_count > 0:
                messages.success(request, f"Imported {success_count} class{'es' if success_count > 1 else ''} successfully!", extra_tags="classes_success")

            if errors:
                for err in errors:
                    messages.error(request, f"❌ {err}", extra_tags="classes_error")

        except Exception as e:
            messages.error(request, f"Bulk upload failed: {str(e)}", extra_tags="classes_error")

    return redirect('admin_classes')



def upload_subjects(request):

    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        print("🔴 User not logged in. Redirecting to login.")
        return redirect('login')
    print("Request received. Method:", request.method)  # Debugging

    if request.method == 'POST':
        form = SubjectUploadForm(request.POST, request.FILES)
        if form.is_valid():
            print("Form is valid.")  # Debugging
        else:
            print("Form errors:", form.errors)  # Debugging
            messages.add_message(request, messages.ERROR, "❌ Invalid form submission.", extra_tags="subject_error")
            return render(request, 'admin/admin_subjects.html', {'form': form})

        file = request.FILES['file']
        print("File received:", file.name)  # Debugging

        try:
            # Fetch the currently logged-in institution's ID
            institution_id = request.session.get('institution_id')  # Assuming it's stored in session
            if not institution_id:
                messages.add_message(request, messages.ERROR, "❌ Institution not found. Please log in again.", extra_tags="subject_error")
                print("Error: Institution ID not found.")  # Debugging
                return redirect('admin_subjects')

            print("Logged-in Institution ID:", institution_id)  # Debugging
            
            try:
                institution = Institution.objects.get(institution_id=institution_id)
            except Institution.DoesNotExist:
                messages.add_message(request, messages.ERROR, "❌ Institution record not found.", extra_tags="subject_error")
                return render(request, 'admin/admin_subjects.html', {'form': form})

            # Read Excel file
            try:
                df = pd.read_excel(file)
                print("File read successfully.")  # Debugging
                print("Columns in file:", df.columns.tolist())  # Debugging
            except Exception as e:
                messages.add_message(request, messages.ERROR, f"❌ Error reading the file: {e}", extra_tags="subject_error")
                print("Error reading the file:", e)  # Debugging
                return redirect('admin_subjects')

            # Validate required columns
            required_columns = ['Subject Name', 'Subject Head', 'Email', 'Password', 'Class Abbreviation']
            if not all(col in df.columns for col in required_columns):
                messages.add_message(request, messages.ERROR, "❌ Invalid file format! Ensure the correct columns are present.", extra_tags="subject_error")
                print("Error: Missing required columns.")  # Debugging
                return redirect('admin_subjects')

            success_count = 0
            error_details = []
            
            # Password validation pattern (same as upload_students)
            password_pattern = r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{6,}$"
            email_pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"

            with connection.cursor() as cursor:
                for index, row in df.iterrows():
                    excel_row_num = index + 2
                    subject_name = str(row['Subject Name']).strip()
                    subject_head = str(row['Subject Head']).strip()
                    email = str(row['Email']).strip()
                    password = str(row['Password']).strip()
                    class_abbreviation = str(row['Class Abbreviation']).strip()

                    print(f"Processing row {index}: {subject_name}, {subject_head}, {email}, {class_abbreviation}")  # Debugging

                    # Validate required fields
                    if not all([subject_name, subject_head, email, password, class_abbreviation]):
                        error_details.append(f"Row {excel_row_num}: Missing required fields")
                        continue

                    # Validate email format
                    if not re.match(email_pattern, email):
                        error_details.append(f"Row {excel_row_num}: Invalid email format '{email}'")
                        continue

                    # Validate password strength
                    if not re.match(password_pattern, password):
                        error_details.append(f"Row {excel_row_num}: Password for '{subject_name}' is too weak. (Must have: 1 Upper, 1 Lower, 1 Digit, 1 Special Char, and 6+ length)")
                        continue

                    # Fetch the class_obj_id based on Class Abbreviation
                    cursor.execute("SELECT id FROM main_classes WHERE class_abbreviation = %s AND institution_id = %s", [class_abbreviation, institution_id])
                    class_row = cursor.fetchone()

                    if not class_row:
                        error_details.append(f"Row {excel_row_num}: Class with abbreviation '{class_abbreviation}' does not exist")
                        continue

                    class_obj_id = class_row[0]
                    print(f"Class with abbreviation '{class_abbreviation}' found with ID: {class_obj_id}")  # Debugging

                    # Check for duplicate subject name in the same class
                    cursor.execute("SELECT id FROM main_subjects WHERE subject_name = %s AND class_obj_id = %s", [subject_name, class_obj_id])
                    if cursor.fetchone():
                        error_details.append(f"Row {excel_row_num}: Subject '{subject_name}' already exists for class with abbreviation '{class_abbreviation}'")
                        continue

                    # Check for duplicate email
                    cursor.execute("SELECT id FROM main_subjects WHERE email = %s", [email])
                    if cursor.fetchone():
                        error_details.append(f"Row {excel_row_num}: Email '{email}' is already taken")
                        continue

                    # Check for duplicate subject head name in the same class
                    cursor.execute("SELECT id FROM main_subjects WHERE subject_head = %s AND class_obj_id = %s", [subject_head, class_obj_id])
                    if cursor.fetchone():
                        error_details.append(f"Row {excel_row_num}: Subject Head '{subject_head}' already exists for class with abbreviation '{class_abbreviation}'")
                        continue

                    try:
                        # Insert into Users table
                        cursor.execute("INSERT INTO main_users (role) VALUES (%s)", ['subject_head'])
                        user_id = cursor.lastrowid  # Get the last inserted ID
                        print(f"Inserted user with ID: {user_id}")  # Debugging

                        # Encrypt the password before storing
                        encrypted_password = encrypt_password(password)
                        
                        # Insert into main_subjects table using class_obj_id
                        cursor.execute("""
                            INSERT INTO main_subjects (class_obj_id, subject_name, subject_head, email, password, user_id)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, [class_obj_id, subject_name, subject_head, email, encrypted_password, user_id])
                        print(f"Inserted subject: {subject_name}")  # Debugging
                        success_count += 1

                        # Try to send welcome email (don't fail whole transaction if email fails)
                        try:
                            send_account_creation_email(email, password, "subject_head", subject_head, institution.email)
                        except Exception as mail_err:
                            print(f"Email failed for {email}: {mail_err}")

                    except Exception as e:
                        error_details.append(f"Row {excel_row_num}: Database error - {str(e)}")
                        continue  # Skip this row

            # Provide Feedback
            if success_count > 0:
                messages.add_message(request, messages.SUCCESS, f"✅ {success_count} subjects uploaded successfully!", extra_tags="subject_success")
            
            if error_details:
                for err in error_details:
                    messages.add_message(request, messages.ERROR, f"❌ {err}", extra_tags="subject_error")

            print("Upload process completed! Redirecting to admin_subjects.")  # Debugging
            return redirect('admin_subjects')

        except Exception as e:
            messages.add_message(request, messages.ERROR, f"❌ Error processing file: {e}", extra_tags="subject_error")
            print("Error occurred:", e)  # Debugging
            return redirect('admin_subjects')

    else:
        form = SubjectUploadForm()
        print("GET request received. Rendering form.")  # Debugging

    return render(request, 'admin/admin_subjects.html', {'form': form})


def upload_students(request):

    if 'institution_id' not in request.session:
        messages.error(request, 'Please log in to access this page.')
        print("🔴 User not logged in. Redirecting to login.")
        return redirect('login')
    form = StudentUploadForm()

    if request.method == 'POST':
        form = StudentUploadForm(request.POST, request.FILES)
        
        if not form.is_valid():
            messages.add_message(request, messages.ERROR, "❌ Invalid form submission.", extra_tags="student_error")
            return render(request, 'admin/admin_students.html', {'form': form})

        file = request.FILES['file']

        try:
            # 1. Get Institution Info
            institution_id = request.session.get('institution_id')
            if not institution_id:
                messages.add_message(request, messages.ERROR, "❌ Session Error: Institution ID not found.", extra_tags="student_error")
                return render(request, 'admin/admin_students.html', {'form': form})
            
            # Fetch institution object to get the institution email for the sender field
            try:
                institution = Institution.objects.get(institution_id=institution_id)
            except Institution.DoesNotExist:
                messages.add_message(request, messages.ERROR, "❌ Institution record not found.", extra_tags="student_error")
                return render(request, 'admin/admin_students.html', {'form': form})

            # 2. Read Excel
            df = pd.read_excel(file)
            df.dropna(how='all', inplace=True)
            
            current_yy = datetime.now().strftime('%y')
            success_count = 0
            error_details = []
            emails_to_send = [] # Queue to hold data for successful entries

            password_pattern = r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{6,}$"

            with connection.cursor() as cursor:
                for index, row in df.iterrows():
                    excel_row_num = index + 2 
                    
                    # Clean data
                    student_name = str(row['Student Name']).strip()
                    email = str(row['Email']).strip()
                    roll_no = str(row['Roll No']).strip()
                    password = str(row['Password']).strip()
                    class_abbr = str(row['Class Abbreviation']).strip()
                    parent_email = str(row.get('Parent Email', '')).strip()

                    # --- VALIDATION ---
                    pattern = rf"^{re.escape(str(institution_id))}S{re.escape(class_abbr)}{current_yy}\d+$"
                    
                    #1. Check for Roll No Format
                    if not re.match(pattern, roll_no):
                        error_msg = f"Row {excel_row_num}: '{roll_no}' rejected. Required format: {institution_id}S{class_abbr}{current_yy}XXX"
                        error_details.append(error_msg)
                        continue 
                    
                    #2. Check for Password Strength
                    if not re.match(password_pattern, password):
                        error_details.append(f"Row {excel_row_num}: Password for '{student_name}' is too weak. (Must have: 1 Upper, 1 Lower, 1 Digit, 1 Special Char, and 6+ length)")
                        continue

                    try:
                        #3. Check if Class exists
                        cursor.execute("SELECT id FROM main_classes WHERE class_abbreviation = %s AND institution_id = %s", [class_abbr, institution_id])
                        class_row = cursor.fetchone()

                        if not class_row:
                            error_details.append(f"Row {excel_row_num}: Class Abbreviation '{class_abbr}' not found.")
                            continue
                        
                        class_obj_id = class_row[0]

                        #4 Check for Duplicates
                        cursor.execute("SELECT id FROM main_students WHERE roll_no = %s AND class_obj_id = %s", [roll_no, class_obj_id])
                        if cursor.fetchone():
                            error_details.append(f"Row {excel_row_num}: Roll No already exists.")
                            continue

                        # 5. Check for Duplicate Email (NEW)
                        cursor.execute("SELECT id FROM main_students WHERE email = %s", [email])
                        if cursor.fetchone():
                            error_details.append(f"Row {excel_row_num}: Email '{email}' is already registered.")
                            continue

                        # Execute Insertion in Transaction
                        with transaction.atomic():
                            # Create Student User Profile
                            cursor.execute("INSERT INTO main_users (role) VALUES (%s)", ['student'])
                            student_user_id = cursor.lastrowid

                            # Encrypt the password before storing
                            encrypted_student_password = encrypt_password(password)
                            
                            # Create Student
                            cursor.execute("""
                                INSERT INTO main_students (user_id, name, roll_no, password, class_obj_id, email)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            """, [student_user_id, student_name, roll_no, encrypted_student_password, class_obj_id, email])
                            student_id = cursor.lastrowid

                            # Create Parent User Profile
                            cursor.execute("INSERT INTO main_users (role) VALUES (%s)", ['parent'])
                            parent_user_id = cursor.lastrowid

                            # Parent password is roll_no, encrypt it
                            encrypted_parent_password = encrypted_student_password
                            
                            # Create Parent
                            cursor.execute("""
                                INSERT INTO main_parents (user_id, student_id, password, name, email)
                                VALUES (%s, %s, %s, NULL, %s)
                            """, [parent_user_id, student_id, encrypted_parent_password, parent_email if parent_email else None])

                        # If we reached here, DB insertion was successful
                        success_count += 1
                        # Add to email queue
                        emails_to_send.append({
                            'student_email': email,
                            'student_password': password,
                            'student_name': student_name,
                            'parent_email': parent_email,
                            'roll_no': roll_no
                        })

                    except Exception as row_error:
                        error_details.append(f"Row {excel_row_num}: {str(row_error)}")

            # --- SEND EMAILS FOR SUCCESSFUL UPLOADS ---
            for data in emails_to_send:
                # Student Notification
                try:
                    send_account_creation_email(
                        data['student_email'], 
                        data['student_password'], 
                        "student", 
                        data['student_name'], 
                        institution.email, 
                        data['roll_no']
                    )
                except Exception as e:
                    print(f"Student email failed for {data['student_email']}: {e}")

                # Parent Notification
                try:
                    if data['parent_email']:
                        send_account_creation_email(
                            data['parent_email'],
                            data['student_password'], # Parent password is roll_no
                            "parent",
                            f"Parent of {data['student_name']}",
                            institution.email,
                            data['roll_no']
                        )
                except Exception as e:
                    print(f"Parent email failed for {data['parent_email']}: {e}")

            # Provide Feedback
            if success_count > 0:
                messages.add_message(request, messages.SUCCESS, f"✅ {success_count} students uploaded and notified.", extra_tags="student_success")
            
            if error_details:
                for err in error_details:
                    messages.add_message(request, messages.ERROR, f"❌ {err}", extra_tags="student_error")

            return redirect('admin_students')

        except Exception as e:
            messages.add_message(request, messages.ERROR, f"❌ System Error: {e}", extra_tags="student_error")
            return redirect('admin_students')

    return redirect('admin_students')



######################################################################################################################


def class_head_login(request):
    form = ClassHeadLoginForm()

    if request.method == 'POST':
        form = ClassHeadLoginForm(request.POST)

        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            # Fetch class head record
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT c.id, c.class_name, c.password, u.role, c.institution_id
                    FROM main_classes c
                    JOIN main_users u ON c.user_id = u.id
                    WHERE c.email = %s
                """, [email])

                class_head = cursor.fetchone()

            if class_head:
                class_id, class_name, db_password, role, institution_id = class_head

                # Validate role and password using decryption for encrypted passwords
                if role == 'class_head' and verify_password(password, db_password):
                    request.session['class_id'] = class_id  # Store class_id in session
                    
                    # Create audit log
                    try:
                        from .models import Institution, Classes
                        client_ip = get_client_ip(request)
                        institution = Institution.objects.get(institution_id=institution_id)
                        class_obj = Classes.objects.get(id=class_id)
                        AuditLog.objects.create(
                            user=class_obj.user,
                            institution=institution,
                            action='login',
                            ip_address=client_ip,
                            user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
                            details=f'Class Head login: {class_name}',
                            related_object_id=class_id,
                            related_object_type='Classes'
                        )
                    except Exception as e:
                        print(f"Class head login audit logging failed: {e}")
                    
                    return redirect('class_head_dashboard')  # Redirect to the dashboard
                else:
                    messages.error(request, 'Invalid credentials', extra_tags="ch_login_error")
            else:
                messages.error(request, 'No Class Head found with this email!', extra_tags="ch_no_email")

    return render(request, 'class_head/class_head_login.html', {'form': form})



def class_head_dashboard(request):
    class_id = request.session.get('class_id')

    if not class_id:
        messages.error(request, 'Session expired or not logged in. Please log in again.')
        return redirect('class_head_login')

    try:
        with connection.cursor() as cursor:
            # 1. Fetch class details
            cursor.execute("""
                SELECT c.class_name, c.email, c.class_head
                FROM main_classes c
                WHERE c.id = %s
            """, [class_id])
            class_details = cursor.fetchone()

            if not class_details:
                messages.error(request, 'Class details not found.')
                return redirect('class_head_login')

            # 2. Fetch student count
            cursor.execute("""
                SELECT COUNT(*)
                FROM main_students s
                WHERE s.class_obj_id = %s
            """, [class_id])
            student_count = cursor.fetchone()[0]

            # 3. Fetch attendance count
            cursor.execute("""
                SELECT COUNT(*) 
                FROM main_attendance a
                JOIN main_students s ON a.student_id = s.id
                WHERE s.class_obj_id = %s AND a.status = 'present'
            """, [class_id])
            present_count = cursor.fetchone()[0]

            attendance_percentage = (present_count / student_count * 100) if student_count > 0 else 0

            # 4. Fetch average performance (FIXED JOIN HERE)
            cursor.execute("""
                SELECT AVG(ev.academic_activity_rating)
                FROM main_studentevaluation ev
                JOIN main_students s ON ev.student_id = s.id
                WHERE s.class_obj_id = %s
            """, [class_id])
            avg_performance = cursor.fetchone()[0] or 0

            # 5. Fetch subject count
            cursor.execute("""
                SELECT COUNT(*)
                FROM main_subjects sub
                WHERE sub.class_obj_id = %s
            """, [class_id])
            subject_count = cursor.fetchone()[0]

    except Exception as e:
        print(f"Database Error: {e}")
        messages.error(request, "An error occurred while loading the dashboard.")
        return redirect('class_head_login')

    return render(request, 'class_head/class_head_dashboard.html', {
        'class_details': class_details,
        'student_count': student_count,
        'subject_count': subject_count,
        'attendance_percentage': round(attendance_percentage, 2),
        'avg_performance': round(avg_performance, 2),
    })
from django.shortcuts import render, redirect
from django.db import connection
from django.http import JsonResponse
from django.contrib import messages
from .models import Classes, Subjects, Students, QuizQuestions, QuizResponse

def class_head_class(request):
    class_id = request.session.get('class_id')
    if not class_id:
        return redirect('class_head_login')

    try:
        class_details = Classes.objects.get(id=class_id)
    except Classes.DoesNotExist:
        messages.error(request, 'Class not found.')
        return redirect('class_head_login')

    # --- 1. POST HANDLER (Announcements) ---
    if request.method == "POST" and request.POST.get("announcementText"):
        announcement_text = request.POST.get("announcementText", "").strip()
        if announcement_text:
            with connection.cursor() as cursor:
                cursor.execute(
                    "INSERT INTO main_announcements (message, created_at, class_obj_id, is_read) VALUES (%s, %s, %s, FALSE)",
                    [announcement_text, now(), class_id]
                )
            return JsonResponse({"success": True, "message": "Announcement added successfully!"})
        return JsonResponse({"success": False, "message": "Announcement cannot be empty!"})

    # --- 2. DATA RETRIEVAL ---
    subjects = Subjects.objects.filter(class_obj=class_details)
    students = Students.objects.filter(class_obj=class_details)
    total_students = students.count()

    with connection.cursor() as cursor:
        # Quiz Count
        cursor.execute("SELECT COUNT(*) FROM main_quizzes WHERE class_obj_id = %s", [class_id])
        quiz_count = cursor.fetchone()[0] or 0

        # Total individual responses (completions)
        cursor.execute("""
            SELECT COUNT(*) FROM main_quizresponse 
            WHERE student_id IN (SELECT id FROM main_students WHERE class_obj_id = %s)""", [class_id])
        total_completions = cursor.fetchone()[0] or 0

        # Attendance Rate
        cursor.execute("""
            SELECT COALESCE(COUNT(CASE WHEN status = 'present' THEN 1 END) * 100.0 / NULLIF(COUNT(*), 0), 0)
            FROM main_attendance 
            WHERE student_id IN (SELECT id FROM main_students WHERE class_obj_id = %s)
        """, [class_id])
        attendance_percentage = round(cursor.fetchone()[0], 1)

        # Performance Rate (Mastery)
        # We join QuizResponse with QuizQuestions to compare student_response with correct_option
        cursor.execute("""
            SELECT 
                COALESCE(COUNT(CASE WHEN r.student_response = q.correct_option THEN 1 END) * 100.0 / NULLIF(COUNT(r.id), 0), 0)
            FROM main_quizresponse r
            JOIN main_quizquestions q ON r.question_id = q.id
            WHERE r.student_id IN (SELECT id FROM main_students WHERE class_obj_id = %s)
        """, [class_id])
        performance_rate = round(cursor.fetchone()[0], 1)

        # Announcements
        cursor.execute("""
            SELECT id, message, created_at FROM main_announcements 
            WHERE class_obj_id = %s ORDER BY created_at DESC
        """, [class_id])
        announcements = cursor.fetchall()

    # --- 3. COMPLETION RATE CALCULATION ---
    # This is "How many quizzes did they actually take vs how many were assigned"
    possible_submissions = total_students * quiz_count
    completion_rate = round((total_completions / possible_submissions) * 100, 1) if possible_submissions > 0 else 0

    return render(request, 'class_head/class_head_class.html', {
        'class_head_name': class_details.class_head,
        'class_name': class_details.class_name,
        'students': students,
        'student_count': total_students,
        'subjects': subjects,
        'subject_count': subjects.count(),
        'announcements': announcements,
        'quiz_count': quiz_count,
        'total_completions': total_completions,
        'total_students': total_students,
        'completion_rate': completion_rate,
        'attendance_percentage': attendance_percentage,
        'performance_rate': performance_rate,
    })  
    

    
def class_head_profile(request):
    class_head_id = request.session.get('class_id')  # Use the correct session key
    
    if not class_head_id:
        return redirect('class_head_login')  # Redirect if not logged in

    # Fetch the class (which represents the class head) along with its institution
    class_head = get_object_or_404(Classes, id=class_head_id)
    
    # Decrypt password for display in template
    decrypted_password = decrypt_password(class_head.password)
    
    # Get institution object using the foreign key relationship (assuming it's a ForeignKey)
    institution = class_head.institution_id  # This is the integer (ID) for institution

    if institution:
        # Fetch the institution object to get the institution_name
        institution_obj = get_object_or_404(Institution, institution_id=institution)
        institution_name = institution_obj.institution_name
    else:
        institution_name = "Unknown Institution"

    if request.method == 'POST':
        class_head.class_head = request.POST.get('name')  # Update class head name
        password = request.POST.get('password')
        if password:  # Only update password if provided
            class_head.password = encrypt_password(password)  # Encrypt the password
        class_head.save()
        messages.success(request, 'Profile updated successfully.')

   # Keep all your existing logic above exactly as it is...


    return render(request, 'class_head/class_head_profile.html', {
        'class_head': class_head,
        'institution_name': institution_name,
        'class_head_name': class_head.class_head,
        'decrypted_password': decrypted_password,  # Pass decrypted password separately
    })


def class_head_reports(request):
    
    class_head_id = request.session.get('class_id')
    
    if not class_head_id:
        return redirect('class_head_login')

    # Fetch the class and institution
    class_head = get_object_or_404(Classes, id=class_head_id)
    institution = class_head.institution_id
    
    if institution:
        institution_obj = get_object_or_404(Institution, institution_id=institution)
        institution_name = institution_obj.institution_name
    else:
        institution_name = "Unknown Institution"
    
    # Get students and subjects for this class
    students = Students.objects.filter(class_obj=class_head)
    subjects = Subjects.objects.filter(class_obj=class_head)
    
    # Get report type from URL parameter
    report_type = request.GET.get('type', 'attendance')
    valid_report_types = ['attendance', 'quiz', 'marks']
    if report_type not in valid_report_types:
        report_type = 'attendance'
    
    # --- FIXED ATTENDANCE LOGIC ---
    # Get date filter from request if available
    selected_date_str = request.GET.get('date')
    if selected_date_str:
        try:
            from datetime import datetime as dt
            selected_date = dt.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()  # fallback to today if invalid date
    else:
        selected_date = timezone.now().date()  # default to today
        
    attendance_data = []
    
    attendance_query = """
    SELECT 
        s.id, 
        s.name, 
        s.roll_no,
        COALESCE(attendance_stats.present_count, 0) as present_count,
        COALESCE(attendance_stats.total_count, 0) as total_count,
        CASE 
            WHEN COALESCE(attendance_stats.total_count, 0) > 0 THEN 
                ROUND((COALESCE(attendance_stats.present_count, 0) * 100.0) / attendance_stats.total_count, 2)
            ELSE 0
        END as percentage
    FROM main_students s
    LEFT JOIN (
        SELECT 
            student_id,
            SUM(CASE WHEN status = 'present' THEN 1 ELSE 0 END) as present_count,
            COUNT(*) as total_count
        FROM main_attendance
        WHERE attendance_date = %s
        GROUP BY student_id
    ) attendance_stats ON s.id = attendance_stats.student_id
    WHERE s.class_obj_id = %s
    ORDER BY s.roll_no
    """
        
    # The order MUST match the '%s' in the query: 1. date, 2. class_id
    attendance_params = [selected_date, class_head_id]
        
    try:
        with connection.cursor() as cursor:
            cursor.execute(attendance_query, attendance_params)
            rows = cursor.fetchall()
                
        for row in rows:
            attendance_data.append({
                'id': row[0],
                'name': row[1],
                'roll_no': row[2],
                'present_count': row[3],
                'total_count': row[4],
                'percentage': row[5]
            })
    except Exception as e:
        print(f"Attendance Error: {e}")
        attendance_data = []

    # --- QUIZ LOGIC ---
    quiz_data = []
    
    # First get total quizzes available for the class
    total_class_quizzes = 0
    with connection.cursor() as cursor:
        cursor.execute("SELECT COUNT(*) FROM main_quizzes WHERE class_obj_id = %s", [class_head_id])
        total_class_quizzes = cursor.fetchone()[0] or 0
    
    quiz_query = """
        SELECT s.id, s.name, s.roll_no,
               COALESCE(quiz_stats.total_quizzes, 0) as total_quizzes,
               COALESCE(quiz_stats.total_correct, 0) as total_correct,
               COALESCE(quiz_stats.total_questions, 0) as total_questions,
               CASE 
                   WHEN COALESCE(quiz_stats.total_questions, 0) > 0 THEN 
                       ROUND((COALESCE(quiz_stats.total_correct, 0) * 100.0) / quiz_stats.total_questions, 2)
                   ELSE 0
               END as overall_percentage
        FROM main_students s
        LEFT JOIN (
            SELECT qr.student_id,
                   COUNT(DISTINCT q.id) as total_quizzes,
                   SUM(CASE WHEN qr.student_response = qq.correct_option THEN 1 ELSE 0 END) as total_correct,
                   COUNT(qq.id) as total_questions
            FROM main_quizresponse qr
            JOIN main_quizquestions qq ON qr.question_id = qq.id
            JOIN main_quizzes q ON qq.quiz_id = q.id
            GROUP BY qr.student_id
        ) quiz_stats ON s.id = quiz_stats.student_id
        WHERE s.class_obj_id = %s
        ORDER BY s.roll_no
    """
    
    try:
        with connection.cursor() as cursor:
            cursor.execute(quiz_query, [class_head_id])
            quiz_rows = cursor.fetchall()
            
        for row in quiz_rows:
            quiz_data.append({
                'student_id': row[0],
                'student_name': row[1],
                'roll_no': row[2],
                'attended_quizzes': row[3],
                'total_class_quizzes': total_class_quizzes,
                'total_questions': row[5],
                'overall_percentage': row[6]
            })
    except Exception as e:
        quiz_data = []

    # --- MARKS LOGIC ---
    # --- MARKS LOGIC ---
    marks_data = []
    try:
        for student in students:
            student_marks = {
                'id': student.id,
                'name': student.name,
                'roll_no': student.roll_no,
                'subjects': []
            }
            
            for subject in subjects:
                marks_query = """
                    SELECT m.marks_percentage
                    FROM main_studentevaluation m
                    WHERE m.student_id = %s AND m.subject_id = %s
                    ORDER BY m.id DESC LIMIT 1
                """
                with connection.cursor() as marks_cursor:
                    marks_cursor.execute(marks_query, [student.id, subject.id])
                    marks_record = marks_cursor.fetchone()
                    
                if marks_record:
                    internal_marks = float(marks_record[0]) if marks_record[0] is not None else 0
                    exam_type = "N/A"
                else:
                    internal_marks = 0
                    exam_type = "N/A"
                
                student_marks['subjects'].append({
                    'subject_id': subject.id,
                    'subject_name': subject.subject_name,
                    'internal_marks': internal_marks,
                    'exam_type': exam_type,
                    'max_marks': 100
                })
            
            # Calculate totals for marks
            total_obtained = sum(sub['internal_marks'] for sub in student_marks['subjects'])
            student_marks['total_obtained'] = total_obtained
            student_marks['average'] = round(total_obtained / len(student_marks['subjects']), 2) if student_marks['subjects'] else 0
            
            # --- FIXED INDENTATION HERE ---
            # This must be inside the student loop to save EACH student
            marks_data.append(student_marks) 

    except Exception as e:
        print(f"Error: {e}")
        marks_data = []
    
    total_possible_marks = subjects.count() * 100 if subjects.exists() else 0
    hours_range = range(1, 11)
    
    context = {
        'class_head': class_head,
        'institution_name': institution_name,
        'class_head_name': class_head.class_head,
        'students': students,
        'subjects': subjects,
        'report_type': report_type,
        'attendance_data': attendance_data,
        'quiz_data': quiz_data,
        'marks_data': marks_data,
        'total_possible_marks': total_possible_marks,
        'hours_range': hours_range,
        'selected_date': selected_date_str,
        'today_date': selected_date,
    }
    
    return render(request, 'class_head/class_head_reports.html', context)


def class_head_chat(request):
    """Handles the chat functionality for the class head."""
    
    # Retrieve the class_id from session
    class_id = request.session.get('class_id')

    # Redirect to login if class_id is not found in session
    if not class_id:
        return redirect('class_head_login')

    # Retrieve class head details
    class_head = Classes.objects.get(id=class_id)
    class_head_user = class_head.user  # Associated user for the class head

    # Fetch users who have interacted with the class head
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT DISTINCT u.id, u.role, 
                CASE 
                    WHEN u.role = 'subject_head' THEN s.subject_head
                    WHEN u.role = 'student' THEN st.name
                    WHEN u.role = 'parent' THEN 
                        CASE 
                            WHEN p.student_id IS NOT NULL THEN 
                                (SELECT st.name FROM main_students AS st WHERE st.id = p.student_id)
                            ELSE p.name 
                        END 
                END AS user_name
            FROM main_chat AS chat
            JOIN main_users AS u ON (chat.sender_id = u.id OR chat.receiver_id = u.id)
            LEFT JOIN main_subjects AS s ON s.user_id = u.id AND u.role = 'subject_head'
            LEFT JOIN main_students AS st ON st.user_id = u.id AND u.role = 'student'
            LEFT JOIN main_parents AS p ON p.user_id = u.id AND u.role = 'parent'
            WHERE (chat.sender_id = %s OR chat.receiver_id = %s)
            AND u.role != 'class_head'
            ORDER BY user_name;
        """, [class_head_user.id, class_head_user.id])

        involved_users = cursor.fetchall()

        # Beautified debug output
        print("\n📌 **Involved Users:**")
        for user in involved_users:
            print(f"🔹 ID: {user[0]} | Role: {user[1]} | Name: {user[2]}")
        print("======================================\n")

        # Fetch all users (subject heads, students, parents) for the class
        cursor.execute("""
            SELECT u.id, u.role, 
                CASE 
                    WHEN u.role = 'subject_head' THEN s.subject_head
                    WHEN u.role = 'student' THEN st.name
                    WHEN u.role = 'parent' THEN linked_st.name  -- Always show linked student name for parents
                END AS user_name
            FROM main_users AS u
            LEFT JOIN main_subjects AS s ON s.user_id = u.id AND u.role = 'subject_head'
            LEFT JOIN main_students AS st ON st.user_id = u.id AND u.role = 'student'
            LEFT JOIN main_parents AS p ON p.user_id = u.id AND u.role = 'parent'
            LEFT JOIN main_students AS linked_st ON linked_st.id = p.student_id  -- Always fetch student's name for parents
            WHERE u.role IN ('subject_head', 'student', 'parent')
            AND (
                st.class_obj_id = %s OR 
                linked_st.class_obj_id = %s OR  -- Ensures parents are included based on their child's class
                s.class_obj_id = %s
            )
            ORDER BY user_name;
        """, [class_id, class_id, class_id])




        all_users_for_class = cursor.fetchall()

        # Beautified debug output
        print("\n📌 **All Users for Class:**")
        for user in all_users_for_class:
            print(f"🔹 ID: {user[0]} | Role: {user[1]} | Name: {user[2]}")
        print("======================================\n")

    # Prepare context for rendering
    context = {
        'class_head': class_head,
        'involved_users': involved_users,
        'all_users_for_class': all_users_for_class,  
    }

    # ... your existing code above ...
    
    # Extract the name string from the object to satisfy the template variable
    class_head_name = class_head.class_head 

    # Prepare context for rendering
    context = {
        'class_head': class_head,
        'class_head_name': class_head_name,  # ADD THIS LINE
        'involved_users': involved_users,
        'all_users_for_class': all_users_for_class,  
    }

    return render(request, "class_head/class_head_chat.html", context)





def class_head_chat_user(request, user_id):
    from django.utils import timezone
    # Retrieve the logged-in class head's ID from session
    class_head_id = request.session.get('class_id')
    class_data = Classes.objects.get(id=class_head_id)
    class_head_name = class_data.class_head
    if not class_head_id:
        return redirect('class_head_login')

    # Fetch the actual user_id from main_users table based on the session user_id (class_head_id)
    logged_in_user_id = None
    with connection.cursor() as cursor:
        cursor.execute("SELECT user_id FROM main_classes WHERE id = %s", [class_head_id])
        logged_in_user = cursor.fetchone()
        if logged_in_user:
            logged_in_user_id = logged_in_user[0]  # Corrected user ID

    if not logged_in_user_id:
        return redirect('class_head_login')  # Prevents further errors

    # Initialize the user name and role
    selected_user_name = "Unknown"
    selected_user_role = "Unknown"

    # Fetch the selected user's name and role
    with connection.cursor() as cursor:
        cursor.execute("SELECT name FROM main_students WHERE user_id = %s", [user_id])
        student_name = cursor.fetchone()
        if student_name and student_name[0]:
            selected_user_name = student_name[0]
            selected_user_role = 'Student'
        else:
            cursor.execute("SELECT student_id FROM main_parents WHERE user_id = %s", [user_id])
            parent_student_roll_no = cursor.fetchone()
            if parent_student_roll_no and parent_student_roll_no[0]:
                cursor.execute("SELECT name FROM main_students WHERE id = %s", [parent_student_roll_no[0]])
                student_name_for_parent = cursor.fetchone()
                if student_name_for_parent and student_name_for_parent[0]:
                    selected_user_name = f"Parent of {student_name_for_parent[0]}"
                    selected_user_role = 'Parent'
            else:
                cursor.execute("SELECT subject_head FROM main_subjects WHERE user_id = %s", [user_id])
                subject_head_name = cursor.fetchone()
                if subject_head_name and subject_head_name[0]:
                    selected_user_name = subject_head_name[0]
                    selected_user_role = 'Subject Head'
                else:
                    cursor.execute("SELECT role FROM main_users WHERE id = %s", [user_id])
                    selected_user = cursor.fetchone()
                    if selected_user and selected_user[0]:
                        selected_user_role = selected_user[0]
                        selected_user_name = selected_user_role.capitalize()

    print(f"Chatting with {selected_user_name} ({selected_user_role})")

    # Mark all messages from this user as read
    Chat.objects.filter(
        sender_id=user_id,
        receiver_id=logged_in_user_id,
        is_read=False
    ).update(is_read=True)

    # Fetch messages between logged-in user and selected user
    messages = []
    with connection.cursor() as cursor:
        cursor.execute(""" 
            SELECT message, sender_id, receiver_id, created_at 
            FROM main_chat 
            WHERE (sender_id = %s AND receiver_id = %s) 
            OR (sender_id = %s AND receiver_id = %s) 
            ORDER BY created_at;
        """, [user_id, logged_in_user_id, logged_in_user_id, user_id])  
        messages = cursor.fetchall() or []  # Prevent NoneType errors

    print(f"Messages fetched between {logged_in_user_id} and {user_id}")

    # Handle message sending via AJAX
    if request.method == 'POST':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            message_text = request.POST.get('message', '').strip()
            if message_text:
                with connection.cursor() as cursor:
                    cursor.execute(""" 
                        INSERT INTO main_chat (sender_id, receiver_id, message, created_at, is_read) 
                        VALUES (%s, %s, %s, %s, FALSE);
                    """, [logged_in_user_id, user_id, message_text, timezone.now()])
                print(f"Message sent: '{message_text}' from {logged_in_user_id} to {user_id}")
                return JsonResponse({'success': True, 'message': 'Message sent'})
            return JsonResponse({'success': False, 'error': 'Empty message'})
        
    context = {
        'selected_user_name': selected_user_name,
        'selected_user_role': selected_user_role,
        'messages': messages,
        'logged_in_user_id': logged_in_user_id,
        'class_head_name' : class_head_name,
    }
    return render(request, 'class_head/class_head_chat_user.html', context)


def fetch_messages(request, user_id):
    """API endpoint to fetch new messages via AJAX"""
    if request.method == 'GET':
        class_head_id = request.session.get('class_id')
        if not class_head_id:
            return JsonResponse({'error': 'Unauthorized'}, status=401)
        
        # Get logged-in user ID
        with connection.cursor() as cursor:
            cursor.execute("SELECT user_id FROM main_classes WHERE id = %s", [class_head_id])
            logged_in_user = cursor.fetchone()
            if not logged_in_user:
                return JsonResponse({'error': 'Unauthorized'}, status=401)
            logged_in_user_id = logged_in_user[0]
        
        # Fetch messages
        with connection.cursor() as cursor:
            cursor.execute(""" 
                SELECT message, sender_id, receiver_id, created_at 
                FROM main_chat 
                WHERE (sender_id = %s AND receiver_id = %s) 
                OR (sender_id = %s AND receiver_id = %s) 
                ORDER BY created_at;
            """, [user_id, logged_in_user_id, logged_in_user_id, user_id])  
            messages = cursor.fetchall() or []
        
        # Mark messages as read
        Chat.objects.filter(
            sender_id=user_id,
            receiver_id=logged_in_user_id,
            is_read=False
        ).update(is_read=True)
        
        # Format messages for JSON response
        formatted_messages = []
        for msg in messages:
            formatted_messages.append({
                'message': msg[0],
                'sender_id': msg[1],
                'receiver_id': msg[2],
                'created_at': msg[3].isoformat() if msg[3] else ''  # Send ISO format for JS to convert to local time
            })
        
        return JsonResponse({'messages': formatted_messages})
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def student_fetch_messages(request, user_id):
    """API endpoint for students to fetch new messages via AJAX"""
    if request.method == 'GET':
        student_id = request.session.get('student_id')
        if not student_id:
            return JsonResponse({'error': 'Unauthorized'}, status=401)
        
        # Get logged-in student user ID using ORM
        student = get_object_or_404(Students, id=student_id)
        logged_in_user_id = student.user_id
        
        # Fetch messages using ORM
        chat_messages = Chat.objects.filter(
            (Q(sender_id=user_id) & Q(receiver_id=logged_in_user_id)) |
            (Q(sender_id=logged_in_user_id) & Q(receiver_id=user_id))
        ).order_by('created_at')
        
        # Mark messages as read
        Chat.objects.filter(
            sender_id=user_id,
            receiver_id=logged_in_user_id,
            is_read=False
        ).update(is_read=True)
        
        # Format messages for JSON response
        formatted_messages = []
        for msg in chat_messages:
            formatted_messages.append({
                'message': msg.message,
                'sender_id': msg.sender_id,
                'receiver_id': msg.receiver_id,
                'created_at': msg.created_at.isoformat() if msg.created_at else ''
            })
        
        return JsonResponse({'messages': formatted_messages})
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def student_get_unread_count(request):
    """API endpoint for students to get total unread message count and recent unread messages"""
    if request.method == 'GET':
        student_id = request.session.get('student_id')
        if not student_id:
            return JsonResponse({'error': 'Unauthorized'}, status=401)
        
        # Get logged-in student user ID using ORM
        student = get_object_or_404(Students, id=student_id)
        logged_in_user_id = student.user_id
        
        # Get all unread messages from teachers
        unread_messages = Chat.objects.filter(
            receiver_id=logged_in_user_id,
            is_read=False
        ).select_related('sender').order_by('-created_at')[:20]  # Last 20 messages
        
        # Group by sender and count
        from django.db.models import Count, Max
        unread_counts = Chat.objects.filter(
            receiver_id=logged_in_user_id,
            is_read=False
        ).values('sender_id').annotate(
            count=Count('id'),
            last_message=Max('created_at')
        ).order_by('-last_message')
        
        # Build response with sender details
        senders_info = []
        for item in unread_counts:
            sender_id = item['sender_id']
            count = item['count']
            
            # Get sender name based on role
            try:
                user = Users.objects.get(id=sender_id)
                if user.role == 'class_head':
                    sender_name = Classes.objects.filter(user_id=sender_id).values_list('class_head', flat=True).first()
                elif user.role == 'subject_head':
                    sender_name = Subjects.objects.filter(user_id=sender_id).values_list('subject_head', flat=True).first()
                else:
                    sender_name = 'Unknown'
                
                senders_info.append({
                    'sender_id': sender_id,
                    'sender_name': sender_name or 'Unknown',
                    'count': count,
                    'role': user.role
                })
            except Users.DoesNotExist:
                continue
        
        # Get unread announcements count
        from .models import Announcements, AnnouncementRead
        student_class = student.class_obj
        
        # Get all announcements for student's class
        all_announcements = Announcements.objects.filter(class_obj=student_class)
        
        # Get announcements student has already read
        read_announcement_ids = AnnouncementRead.objects.filter(student=student).values_list('announcement_id', flat=True)
        
        # Unread announcements are those not in read list
        unread_announcements = all_announcements.exclude(id__in=read_announcement_ids)
        unread_announcements_count = unread_announcements.count()
        
        total_unread = sum(item['count'] for item in unread_counts) + unread_announcements_count
        
        return JsonResponse({
            'total_unread': total_unread,
            'senders': senders_info,
            'unread_announcements': unread_announcements_count
        })
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@require_http_methods(["POST"])
def student_mark_announcements_read(request):
    """API endpoint to mark all announcements as read for a student"""
    if request.method == 'POST':
        student_id = request.session.get('student_id')
        if not student_id:
            return JsonResponse({'error': 'Unauthorized'}, status=401)
        
        try:
            student = Students.objects.get(id=student_id)
            from .models import Announcements, AnnouncementRead
            
            # Get all announcements for student's class
            all_announcements = Announcements.objects.filter(class_obj=student.class_obj)
            
            # Mark all unread announcements as read
            marked_count = 0
            for announcement in all_announcements:
                # Check if student already read this announcement
                read_record, created = AnnouncementRead.objects.get_or_create(
                    announcement=announcement,
                    student=student
                )
                if created:
                    marked_count += 1
            
            return JsonResponse({
                'success': True,
                'marked_count': marked_count,
                'message': f'{marked_count} announcements marked as read'
            })
        except Students.DoesNotExist:
            return JsonResponse({'error': 'Student not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


def class_head_get_unread_count(request):
    from django.db.models import Count, Max
    """API endpoint for class heads to get total unread message count and recent unread messages"""
    if request.method == 'GET':
        class_id = request.session.get('class_id')
        if not class_id:
            return JsonResponse({'error': 'Unauthorized'}, status=401)
        
        # Get class head user ID using ORM
        classes = get_object_or_404(Classes, id=class_id)
        logged_in_user_id = classes.user_id
        
        # Get all unread messages from students, parents, and subject heads
        unread_counts = Chat.objects.filter(
            receiver_id=logged_in_user_id,
            is_read=False
        ).values('sender_id').annotate(
            count=Count('id'),
            last_message=Max('created_at')
        ).order_by('-last_message')
        
        # Build response with sender details
        senders_info = []
        for item in unread_counts:
            sender_id = item['sender_id']
            count = item['count']
            
            # Get sender name based on role
            try:
                user = Users.objects.get(id=sender_id)
                if user.role == 'student':
                    sender_name = Students.objects.filter(user_id=sender_id).values_list('name', flat=True).first()
                elif user.role == 'subject_head':
                    sender_name = Subjects.objects.filter(user_id=sender_id).values_list('subject_head', flat=True).first()
                elif user.role == 'parent':
                    parent = Parents.objects.filter(user_id=sender_id).first()
                    if parent and parent.student_id:
                        sender_name = Students.objects.filter(id=parent.student_id).values_list('name', flat=True).first()
                    else:
                        sender_name = parent.name if parent else 'Unknown'
                else:
                    sender_name = 'Unknown'
                
                senders_info.append({
                    'sender_id': sender_id,
                    'sender_name': sender_name or 'Unknown',
                    'count': count,
                    'role': user.role
                })
            except Users.DoesNotExist:
                continue
        
        total_unread = sum(item['count'] for item in unread_counts)
        
        return JsonResponse({
            'total_unread': total_unread,
            'senders': senders_info
        })
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)






def subject_head_get_unread_count(request):
    """API endpoint for subject heads to get total unread message count and recent unread messages"""
    if request.method == 'GET':
        subject_id = request.session.get('subject_id')
        if not subject_id:
            return JsonResponse({'error': 'Unauthorized'}, status=401)
        
        # Get subject head user ID using ORM
        subject = get_object_or_404(Subjects, id=subject_id)
        logged_in_user_id = subject.user_id
        
        # Get all unread messages from students, class heads, and parents
        from django.db.models import Count, Max
        unread_counts = Chat.objects.filter(
            receiver_id=logged_in_user_id,
            is_read=False
        ).values('sender_id').annotate(
            count=Count('id'),
            last_message=Max('created_at')
        ).order_by('-last_message')
        
        # Build response with sender details
        senders_info = []
        for item in unread_counts:
            sender_id = item['sender_id']
            count = item['count']
            
            # Get sender name based on role
            try:
                user = Users.objects.get(id=sender_id)
                if user.role == 'student':
                    sender_name = Students.objects.filter(user_id=sender_id).values_list('name', flat=True).first()
                elif user.role == 'class_head':
                    sender_name = Classes.objects.filter(user_id=sender_id).values_list('class_head', flat=True).first()
                elif user.role == 'parent':
                    parent = Parents.objects.filter(user_id=sender_id).first()
                    if parent and parent.student_id:
                        sender_name = Students.objects.filter(id=parent.student_id).values_list('name', flat=True).first()
                    else:
                        sender_name = parent.name if parent else 'Unknown'
                elif user.role == 'subject_head':
                    sender_name = Subjects.objects.filter(user_id=sender_id).values_list('subject_head', flat=True).first()
                else:
                    sender_name = 'Unknown'
                
                senders_info.append({
                    'sender_id': sender_id,
                    'sender_name': sender_name or 'Unknown',
                    'count': count,
                    'role': user.role
                })
            except Users.DoesNotExist:
                continue
        
        total_unread = sum(item['count'] for item in unread_counts)
        
        return JsonResponse({
            'total_unread': total_unread,
            'senders': senders_info
        })
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


def parent_get_unread_count(request):
    """API endpoint for parents to get total unread message count and recent unread messages"""
    if request.method == 'GET':
        parent_id = request.session.get('parent_id')
        if not parent_id:
            return JsonResponse({'error': 'Unauthorized'}, status=401)
        
        # Get logged-in parent user ID using ORM
        parent = get_object_or_404(Parents, id=parent_id)
        logged_in_user_id = parent.user_id
        
        # Get all unread messages from teachers (class heads and subject heads)
        from django.db.models import Count, Max
        unread_counts = Chat.objects.filter(
            receiver_id=logged_in_user_id,
            is_read=False
        ).values('sender_id').annotate(
            count=Count('id'),
            last_message=Max('created_at')
        ).order_by('-last_message')
        
        # Build response with sender details
        senders_info = []
        for item in unread_counts:
            sender_id = item['sender_id']
            count = item['count']
            
            # Get sender name based on role
            try:
                user = Users.objects.get(id=sender_id)
                if user.role == 'class_head':
                    sender_name = Classes.objects.filter(user_id=sender_id).values_list('class_head', flat=True).first()
                elif user.role == 'subject_head':
                    sender_name = Subjects.objects.filter(user_id=sender_id).values_list('subject_head', flat=True).first()
                else:
                    sender_name = 'Unknown'
                
                senders_info.append({
                    'sender_id': sender_id,
                    'sender_name': sender_name or 'Unknown',
                    'count': count,
                    'role': user.role
                })
            except Users.DoesNotExist:
                continue
        
        total_unread = sum(item['count'] for item in unread_counts)
        
        return JsonResponse({
            'total_unread': total_unread,
            'senders': senders_info
        })
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)



def class_head_students(request):
    class_id = request.session.get('class_id')

    if not class_id:
        print("❌ Class ID not found in session. Redirecting to login.")
        return redirect('class_head_login')

    with connection.cursor() as cursor:
        # Fetch class name AND class head name
        cursor.execute("""
            SELECT class_abbreviation, class_head FROM main_classes WHERE id = %s
        """, [class_id])
        class_data = cursor.fetchone()

        if class_data:
            class_name = class_data[0]
            class_head_name = class_data[1]  # <--- FETCH THIS
            print(f"🏫 Class Name: {class_name} | Head: {class_head_name}")
        else:
            print(f"⚠️ No class found for class_id: {class_id}")
            class_name = "N/A"
            class_head_name = "Unknown User"

        # Fetch all students for the logged-in class
        cursor.execute("""
            SELECT id, name, roll_no, email FROM main_students
            WHERE class_obj_id = %s
        """, [class_id])
        
        students = cursor.fetchall()
        student_list = []
        for student in students:
            student_list.append({
                "id": student[0],
                "name": student[1],
                "roll_no": student[2],
                "email": student[3] if student[3] else "N/A"
            })

    # ADD class_head_name TO THE CONTEXT
    return render(request, 'class_head/class_head_students.html', {
        "class_name": class_name,
        "class_head_name": class_head_name,  # <--- NOW THE FRONTEND CAN SEE IT
        "students": student_list
    })


import json
from django.shortcuts import render, redirect
from django.db import connection
from ml.predict import predict_performance  # Importing the prediction function

def class_head_student_performance(request, student_id):
    # Retrieve the class_id from session
    class_id = request.session.get('class_id')


    if not class_id:
        print("[ERROR] No class_id found in session. Redirecting to class_head login.")
        return redirect('class_head_login')

    subject_id = request.GET.get('subject_id')  # Get selected subject from query params
    print(f"\n[INFO] Class Head Student Performance View Loaded for Student ID: {student_id}")
    if subject_id:
        print(f"[INFO] Selected Subject ID: {subject_id}")

    student_data = {}
    evaluations = {}
    quiz_percentage = 0  
    subjects = []
    selected_subject_name = None
    predicted_marks = None  # Placeholder for predicted marks

    try:
        with connection.cursor() as cursor:
            # Fetch student details
            cursor.execute("""
                SELECT s.id, s.name, s.roll_no, c.class_name, s.email, c.id AS class_id
                FROM main_students s
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s AND s.class_obj_id = %s
            """, [student_id, class_id])
            row = cursor.fetchone()

            if row:
                student_data = {
                    "id": row[0],
                    "name": row[1],
                    "roll_no": row[2],
                    "class_name": row[3],
                    "email": row[4] if row[4] else "--",
                    "class_id": row[5]
                }
                print(f"[DEBUG] Student Details: {student_data}")
            else:
                print("[ERROR] Student not found or does not belong to this class.")
                return redirect('class_head_dashboard')

        # Fetch subjects for the student's class
        if student_data.get("class_id"):
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, subject_name FROM main_subjects WHERE class_obj_id = %s
                """, [student_data["class_id"]])
                subjects = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]

            selected_subject_name = next((s["name"] for s in subjects if str(s["id"]) == subject_id), None)
            if selected_subject_name:
                print(f"[DEBUG] Selected Subject: {selected_subject_name} (ID: {subject_id})")
            else:
                print("[WARNING] Selected subject not found in student's class subjects.")

        # Fetch student evaluation details for the selected subject
        if subject_id:
            print(f"[DEBUG] Fetching evaluation data for Subject ID: {subject_id}...")
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT marks_percentage, attendance_percentage,
                           study_time_rating, sleep_time_rating, 
                           class_participation_rating, academic_activity_rating
                    FROM main_studentevaluation
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()
                print(f"[DEBUG] Raw database row: {row}")

                if row:
                    evaluations = {
                        "marks_percentage": round(row[0] or 0, 2),
                        "attendance_percentage": round(row[1] or 0, 2),
                        "study_time_rating": round(row[2] or 0, 2),
                        "sleep_time_rating": round(row[3] or 0, 2),
                        "class_participation_rating": round(row[4] or 0, 2),
                        "academic_activity_rating": round(row[5] or 0, 2),
                    }
                    print(f"[DEBUG] Evaluation Data: {evaluations}")

                    # Call prediction model
                    print("[DEBUG] Calling prediction function...")
                    predicted_marks = predict_performance(
                        evaluations["attendance_percentage"],
                        evaluations["marks_percentage"],
                        evaluations["class_participation_rating"],
                        evaluations["academic_activity_rating"],
                        evaluations["sleep_time_rating"],
                        evaluations["study_time_rating"]
                    )

                    # Ensure predicted marks is a float or None
                    predicted_marks = round(float(predicted_marks), 2) if predicted_marks is not None else None
                    print(f"[DEBUG] Predicted Marks: {predicted_marks}")

                else:
                    print(f"[WARNING] No evaluation data found for Student {student_id}, Subject {subject_id}.")
                    print(f"[INFO] Please ensure StudentEvaluation records exist for this student-subject combination.")

        # Fetch quiz performance for the selected subject
        if subject_id and selected_subject_name:
            print(f"[DEBUG] Fetching quiz performance for {selected_subject_name} (ID: {subject_id})...")
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT COUNT(*) AS total_attempted,
                           SUM(CASE WHEN q.correct_option = r.student_response THEN 1 ELSE 0 END) AS correct_answers
                    FROM main_quizresponse r
                    JOIN main_quizquestions q ON r.question_id = q.id
                    JOIN main_quizzes mq ON q.quiz_id = mq.id
                    WHERE r.student_id = %s AND mq.subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    total_attempted, correct_answers = row[0], row[1] or 0
                    quiz_percentage = round((correct_answers / total_attempted) * 100, 2) if total_attempted > 0 else 0
                    print(f"[DEBUG] Quiz Performance: Attempted = {total_attempted}, Correct = {correct_answers}, Percentage = {quiz_percentage}%")
                else:
                    print(f"[WARNING] No quiz data found for {selected_subject_name}.")

        # Prepare data for Graph (ensure all values are numeric)
        graph_data = [
            float(evaluations.get("marks_percentage", 0)),
            float(evaluations.get("attendance_percentage", 0)),
            float(evaluations.get("study_time_rating", 0)),
            float(evaluations.get("sleep_time_rating", 0)),
            float(evaluations.get("class_participation_rating", 0)),
            float(evaluations.get("academic_activity_rating", 0)),
            float(quiz_percentage),
            float(predicted_marks) if predicted_marks is not None else 0  # Ensure numeric value for predicted marks
        ]

    except Exception as e:
        print(f"[ERROR] An error occurred: {e}")
        return redirect('error_page')  

    context = {
        "student": student_data,
        "evaluations": evaluations,
        "quiz_percentage": quiz_percentage,
        "subjects": subjects,
        "selected_subject_id": int(subject_id) if subject_id else None,
        "predicted_marks": predicted_marks,  # Include predicted marks in context
        "graph_data": json.dumps(graph_data),
        "class_head": Classes.objects.get(id=class_id).class_head,  # Class head name
    }

    # Check if the request is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({"predicted_marks": predicted_marks})

    print("[INFO] Final Context Data Prepared for Rendering.")
    return render(request, "class_head/class_head_student_performance.html", context)




######################################################################################################################


def subject_head_login(request):
    form = SubjectHeadLoginForm()

    if request.method == 'POST':
        form = SubjectHeadLoginForm(request.POST)

        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            # Fetch subject head record
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT s.id, s.subject_name, s.password, u.role, c.institution_id
                    FROM main_subjects s
                    JOIN main_users u ON s.user_id = u.id
                    JOIN main_classes c ON s.class_obj_id = c.id
                    WHERE s.email = %s
                """, [email])

                subject_head = cursor.fetchone()

            if subject_head:
                subject_id, subject_name, db_password, role, institution_id = subject_head

                # Validate role and password using decryption for encrypted passwords
                if role == 'subject_head' and verify_password(password, db_password):
                    request.session['subject_id'] = subject_id  # Store subject_id in session
                    
                    # Create audit log
                    try:
                        from .models import Institution, Subjects
                        client_ip = get_client_ip(request)
                        institution = Institution.objects.get(institution_id=institution_id)
                        subject_obj = Subjects.objects.get(id=subject_id)
                        AuditLog.objects.create(
                            user=subject_obj.user,
                            institution=institution,
                            action='login',
                            ip_address=client_ip,
                            user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
                            details=f'Subject Head login: {subject_name}',
                            related_object_id=subject_id,
                            related_object_type='Subjects'
                        )
                    except Exception as e:
                        print(f"Subject head login audit logging failed: {e}")
                    
                    return redirect('subject_head_dashboard')  # Redirect to subject head dashboard
                else:
                    messages.error(request, 'Invalid credentials!', extra_tags="sh_login_error")
            else:
                messages.error(request, 'No Subject Head found with this email!', extra_tags="sh_no_email")

    return render(request, 'subject_head/subject_head_login.html', {'form': form})




def subject_head_dashboard(request):
    # Ensure subject head is logged in
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')  # Redirect to login page if not logged in


    #Fetch student Details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(*)
            FROM main_students s
            JOIN main_subjects sub ON s.class_obj_id = sub.class_obj_id
            WHERE sub.id = %s
        """, [subject_id])

        student_count = cursor.fetchone()[0]  # Fetch count of students under this subject's class

    # Fetch subject details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT s.subject_name, s.subject_head, c.class_name, c.class_head
            FROM main_subjects s
            JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [subject_id])

        subject_details = cursor.fetchone()  # Fetch one record

    # Prepare context for the template
    context = {
        'subject_name': subject_details[0] if subject_details else None,
        'subject_head': subject_details[1] if subject_details else None,
        'class_name': subject_details[2] if subject_details else None,
        'class_head': subject_details[3] if subject_details else None,
        'student_count': student_count,
    }

    return render(request, 'subject_head/subject_head_dashboard.html', context)


def subject_head_class(request):
    # Ensure subject head is logged in
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')

    # Fetch class and faculty details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT c.class_name, c.class_head
            FROM main_subjects s
            JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [subject_id])
        class_info = cursor.fetchone()

    if class_info:
        class_name, class_head_name = class_info
    else:
        class_name, class_head_name = "Unknown", "Unknown"

    # Fetch the logged-in subject's details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT subject_name, subject_head, email
            FROM main_subjects
            WHERE id = %s
        """, [subject_id])
        logged_in_subject = cursor.fetchone()

    if logged_in_subject:
        logged_in_subject_name, logged_in_subject_head, logged_in_subject_email = logged_in_subject
    else:
        logged_in_subject_name, logged_in_subject_head, logged_in_subject_email = "Unknown", "Unknown", "Unknown"

    # Fetch subjects under this class
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, subject_name, subject_head, email
            FROM main_subjects
            WHERE class_obj_id = (SELECT class_obj_id FROM main_subjects WHERE id = %s)
        """, [subject_id])
        subjects = [
            {
                'id': row[0],
                'subject_name': row[1],
                'subject_head': row[2],
                'email': row[3],
                'is_logged_in_subject': row[2] == logged_in_subject_head  # Check if the subject belongs to the logged-in user
            }
            for row in cursor.fetchall()
        ]

    # Fetch students under this class
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, roll_no, name, email
            FROM main_students
            WHERE class_obj_id = (SELECT class_obj_id FROM main_subjects WHERE id = %s)
        """, [subject_id])
        students = [
            {'id': row[0], 'roll_no': row[1], 'name': row[2], 'email': row[3]}
            for row in cursor.fetchall()
        ]

    # Get student count
    student_count = len(students)

    context = {
        'class_name': class_name,
        'class_head_name': class_head_name,
        'subjects': subjects,
        'students': students,
        'student_count': student_count,
        # This is the teacher's name (e.g., "John Doe")
        'logged_in_subject_head': logged_in_subject_head,
        # This is the name of the subject (e.g., "Mathematics")
        'subject_name': logged_in_subject_name, # Fixed variable name here
        # Optional: If your sidebar uses 'subject_head_name', add it here too
        'subject_head_name': logged_in_subject_head, 
    }


    return render(request, 'subject_head/subject_head_class.html', context)



def subject_head_subjects(request):
    # Ensure subject head is logged in
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')

    # Fetch class and faculty details (for the subject head)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT c.class_name, s.subject_head
            FROM main_subjects s
            JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [subject_id])
        class_info = cursor.fetchone()

    if class_info:
        class_name, subject_head = class_info
    else:
        class_name, subject_head = "Unknown", "Unknown"

    # Fetch assigned subject details for the specific subject head
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, subject_name, subject_head, email
            FROM main_subjects
            WHERE id = %s
        """, [subject_id])
        subjects = [
            {'id': row[0], 'subject_name': row[1], 'subject_head': row[2], 'email': row[3]}
            for row in cursor.fetchall()
        ]

    # Fetch students under this subject's class
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, roll_no, name, email
            FROM main_students
            WHERE class_obj_id = (SELECT class_obj_id FROM main_subjects WHERE id = %s)
        """, [subject_id])
        students = [
            {'id': row[0], 'roll_no': row[1], 'name': row[2], 'email': row[3]}
            for row in cursor.fetchall()
        ]

    # Get student count
    student_count = len(students)

    context = {
        'class_name': class_name,
        'subject_head': subject_head,  # Pass subject head information to the template
        'subjects': subjects,
        'students': students,
        'student_count': student_count
    }

    return render(request, 'subject_head/subject_head_subjects.html', context)


def subject_head_profile(request):
    # Ensure subject head is logged in
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login') 

    if request.method == 'POST':
        # Get updated data from the form
        updated_name = request.POST.get('name')
        updated_password = request.POST.get('password')

        # Update the subject head's details in the database
        with connection.cursor() as cursor:
            if updated_password:
                # Encrypt the password before updating
                encrypted_password = encrypt_password(updated_password)
                cursor.execute("""
                    UPDATE main_subjects
                    SET subject_head = %s, password = %s
                    WHERE id = %s
                """, [updated_name, encrypted_password, subject_id])
            else:
                cursor.execute("""
                    UPDATE main_subjects
                    SET subject_head = %s
                    WHERE id = %s
                """, [updated_name, subject_id])
        
        messages.success(request, "Profile updated successfully!")

        return redirect('subject_head_profile')  # Refresh the page after update

    # Fetch current subject head details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT s.subject_name, s.subject_head, s.email, s.password, c.class_name, c.class_head, i.institution_name
            FROM main_subjects s
            JOIN main_classes c ON s.class_obj_id = c.id
            JOIN main_institution i ON c.institution_id = i.institution_id
            WHERE s.id = %s
        """, [subject_id])

        subject_head = cursor.fetchone()  # Fetch one record

    if subject_head:
        # Decrypt password for display
        decrypted_password = decrypt_password(subject_head[3])
        context = {
            'class_head': {
                'subject_name': subject_head[0],  # Subject Name
                'class_head': subject_head[1],  # Subject Head Name
                'email': subject_head[2],
                'password': decrypted_password,  # Use decrypted password
                'class_name': subject_head[4],
            },
            'subject_head': subject_head[1],
            'institution_name': subject_head[6]
        }
    else:
        context = {'class_head': None, 'institution_name': None}

    return render(request, 'subject_head/subject_head_profile.html', context)


def subject_head_chat(request):
    # 1. Authentication Check
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')

    # 2. Fetch Logged-in Subject Head Info (Fixed for name visibility)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT user_id, class_obj_id, subject_head 
            FROM main_subjects 
            WHERE id = %s
        """, [subject_id])
        subject_data = cursor.fetchone()

    if not subject_data:
        return redirect('subject_head_login')

    subject_user_id, class_id, logged_in_name = subject_data

    # 3. Handle Active Chat Selection (The Fix for Message Visibility)
    # We get the ID of the person the Subject Head clicked on (e.g., the Class Head)
    active_chat_user_id = request.GET.get('user_id') 
    messages_list = []

    if active_chat_user_id:
        with connection.cursor() as cursor:
            # THE CRITICAL FIX: Only fetch messages belonging to this SPECIFIC pair
            cursor.execute("""
                SELECT sender_id, receiver_id, message, timestamp 
                FROM main_chat 
                WHERE (sender_id = %s AND receiver_id = %s)
                   OR (sender_id = %s AND receiver_id = %s)
                ORDER BY timestamp ASC
            """, [subject_user_id, active_chat_user_id, active_chat_user_id, subject_user_id])
            
            messages_list = [
                {'sender': row[0], 'receiver': row[1], 'text': row[2], 'time': row[3]} 
                for row in cursor.fetchall()
            ]

    # 4. Fetch User Lists (Sidebar logic - Unchanged)
    with connection.cursor() as cursor:
        # Users involved in past chats
        cursor.execute("""
            SELECT u.id, u.role, 
            COALESCE(c.class_head, s.subject_head, st.name, 
                CASE WHEN p.user_id IS NOT NULL THEN CONCAT('Parent of ', st2.name) ELSE NULL END
            ) AS name
            FROM main_users u
            LEFT JOIN main_classes c ON u.id = c.user_id AND u.role = 'class_head'
            LEFT JOIN main_subjects s ON u.id = s.user_id AND u.role = 'subject_head'
            LEFT JOIN main_students st ON u.id = st.user_id AND u.role = 'student'
            LEFT JOIN main_parents p ON u.id = p.user_id AND u.role = 'parent'
            LEFT JOIN main_students st2 ON p.student_id = st2.id AND u.role = 'parent'
            WHERE u.id IN (
                SELECT DISTINCT sender_id FROM main_chat WHERE receiver_id = %s
                UNION
                SELECT DISTINCT receiver_id FROM main_chat WHERE sender_id = %s
            );
        """, [subject_user_id, subject_user_id])
        chat_users = [{'id': row[0], 'role': row[1], 'name': row[2] or 'Unknown'} for row in cursor.fetchall()]

        # Users available to start new chat
        cursor.execute("""
            SELECT u.id, u.role, 
            COALESCE(st.name, CASE WHEN p.user_id IS NOT NULL THEN CONCAT('Parent of ', st2.name) ELSE NULL END, c.class_head, s.subject_head) AS name
            FROM main_users u
            LEFT JOIN main_students st ON u.id = st.user_id AND u.role = 'student' AND st.class_obj_id = %s
            LEFT JOIN main_parents p ON u.id = p.user_id AND u.role = 'parent'  
            LEFT JOIN main_students st2 ON p.student_id = st2.id AND u.role = 'parent' AND st2.class_obj_id = %s
            LEFT JOIN main_classes c ON u.id = c.user_id AND u.role = 'class_head' AND c.id = %s
            LEFT JOIN main_subjects s ON u.id = s.user_id AND u.role = 'subject_head' AND s.class_obj_id = %s
            WHERE (st.class_obj_id IS NOT NULL OR st2.class_obj_id IS NOT NULL OR c.id IS NOT NULL OR (s.class_obj_id IS NOT NULL AND u.id != %s));
        """, [class_id, class_id, class_id, class_id, subject_user_id])
        
        all_potential_users = cursor.fetchall()

    # --- Categorization Logic ---
    students = []
    parents = []
    subject_heads = []
    class_heads = []

    for row in all_potential_users:
        user_info = {'id': row[0], 'role': row[1], 'name': row[2] or 'Unknown'}
        role = row[1]
        
        if role == 'student':
            students.append(user_info)
        elif role == 'parent':
            parents.append(user_info)
        elif role == 'subject_head':
            subject_heads.append(user_info)
        elif role == 'class_head':
            class_heads.append(user_info)

    return render(request, 'subject_head/subject_head_chat.html', {
        'chat_users': chat_users, 
        'students': students,
        'parents': parents,
        'subject_heads': subject_heads,
        'class_heads': class_heads,
        'messages': messages_list,
        'subject_head': logged_in_name,
        'active_chat_id': active_chat_user_id
    })



from django.shortcuts import render, redirect
from django.db import connection
from django.utils.timezone import now

def subject_head_chat_user(request, user_id):
    # 1. Authentication Check
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')

    # 2. Fetch logged-in Subject Head's AUTH USER ID
    with connection.cursor() as cursor:
        cursor.execute("SELECT user_id, subject_head FROM main_subjects WHERE id = %s", [subject_id])
        subject_data = cursor.fetchone()

    if not subject_data:
        return redirect('subject_head_login')

    # me_id must be the ID from the main_users table to link with chat sender_id/receiver_id
    me_id = int(subject_data[0]) 
    subject_head_name = subject_data[1]

    # 3. Fetch selected user's name and role (Chat Target)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT u.role, 
            COALESCE(c.class_head, s.subject_head, st.name, 
                CASE WHEN p.user_id IS NOT NULL THEN CONCAT('Parent of ', st2.name) ELSE NULL END
            ) AS name
        FROM main_users u
        LEFT JOIN main_classes c ON u.id = c.user_id AND u.role = 'class_head'
        LEFT JOIN main_subjects s ON u.id = s.user_id AND u.role = 'subject_head'
        LEFT JOIN main_students st ON u.id = st.user_id AND u.role = 'student'
        LEFT JOIN main_parents p ON u.id = p.user_id AND u.role = 'parent'
        LEFT JOIN main_students st2 ON p.student_id = st2.id AND u.role = 'parent'
        WHERE u.id = %s
        """, [user_id])
        selected_user = cursor.fetchone()

    if not selected_user:
        return redirect('subject_head_chat')

    selected_user_role, selected_user_name = selected_user

    # Mark all messages from this user as read
    Chat.objects.filter(
        sender_id=user_id,
        receiver_id=me_id,
        is_read=False
    ).update(is_read=True)

    # 4. Handle Message Sending (POST)
    if request.method == "POST":
        message_text = request.POST.get("message", "").strip()
        if message_text:
            with connection.cursor() as cursor:
                # We save sender and receiver as IDs from main_users table
                cursor.execute("""
                    INSERT INTO main_chat (message, sender_id, receiver_id, created_at) 
                    VALUES (%s, %s, %s, %s)
                """, [message_text, me_id, user_id, now()])
            # Redirect back to the same user chat to see the new message
            return redirect('subject_head_chat_user', user_id=user_id)

    # 5. FETCH MESSAGES (Pairwise filtering ensures individual chats)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT message, sender_id, receiver_id, created_at 
            FROM main_chat 
            WHERE (sender_id = %s AND receiver_id = %s) 
               OR (sender_id = %s AND receiver_id = %s) 
            ORDER BY created_at ASC
        """, [me_id, user_id, user_id, me_id])
        raw_history = cursor.fetchall()
    
    # Process history for the frontend loop
    chat_history = []
    for row in raw_history:
        chat_history.append({
            'text': row[0],
            'sender_id': int(row[1]),
            'receiver_id': int(row[2]),
            'time': row[3]
        })

    # 6. Sidebar Lists (Populating users for the sidebar)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT u.id, 
            COALESCE(c.class_head, s.subject_head, st.name) AS name
            FROM main_users u
            LEFT JOIN main_classes c ON u.id = c.user_id
            LEFT JOIN main_subjects s ON u.id = s.user_id
            LEFT JOIN main_students st ON u.id = st.user_id
            WHERE u.id IN (
                SELECT DISTINCT sender_id FROM main_chat WHERE receiver_id = %s
                UNION
                SELECT DISTINCT receiver_id FROM main_chat WHERE sender_id = %s
            )
        """, [me_id, me_id])
        chat_users = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]

    return render(request, 'subject_head/subject_head_chat_user.html', {
        'selected_user_role': selected_user_role,
        'selected_user_name': selected_user_name,
        'chat_history': chat_history,
        'logged_in_user_id': me_id,
        'subject_head': subject_head_name,
        'chat_users': chat_users,
        'selected_user_id': int(user_id)
    })


def subject_head_fetch_messages(request, user_id):
    """API endpoint to fetch new messages via AJAX for Subject Head"""
    if request.method == 'GET':
        subject_id = request.session.get('subject_id')
        if not subject_id:
            return JsonResponse({'error': 'Unauthorized'}, status=401)
        
        # Get logged-in user ID from main_users table
        with connection.cursor() as cursor:
            cursor.execute("SELECT user_id FROM main_subjects WHERE id = %s", [subject_id])
            logged_in_user = cursor.fetchone()
            if not logged_in_user:
                return JsonResponse({'error': 'Unauthorized'}, status=401)
            logged_in_user_id = logged_in_user[0]
        
        # Fetch messages
        with connection.cursor() as cursor:
            cursor.execute(""" 
                SELECT message, sender_id, receiver_id, created_at 
                FROM main_chat 
                WHERE (sender_id = %s AND receiver_id = %s) 
                OR (sender_id = %s AND receiver_id = %s) 
                ORDER BY created_at ASC;
            """, [user_id, logged_in_user_id, logged_in_user_id, user_id])  
            messages = cursor.fetchall() or []
        
        # Mark messages as read
        Chat.objects.filter(
            sender_id=user_id,
            receiver_id=logged_in_user_id,
            is_read=False
        ).update(is_read=True)
        
        # Format messages for JSON response - match the template structure
        formatted_messages = []
        for msg in messages:
            formatted_messages.append({
                'text': msg[0],  # Use 'text' to match template variable
                'sender_id': int(msg[1]),
                'receiver_id': int(msg[2]),
                'time': msg[3].strftime('%Y-%m-%d %H:%M:%S') if msg[3] else '',
                'created_at': msg[3].strftime('%Y-%m-%dT%H:%M:%S.%fZ') if msg[3] else ''
            })
        
        return JsonResponse({'messages': formatted_messages})
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)



def subject_head_quiz(request):
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')

    # Fetch subject details
    subject_data = None
    subject_head_name = "Subject Head"  # Default value

    try:
        with connection.cursor() as cursor:
            # ✅ Updated Query: Now also fetches s.subject_head
            cursor.execute("""
                SELECT s.id, s.subject_name, c.class_name, s.class_obj_id, s.subject_head 
                FROM main_subjects s
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s
            """, [subject_id])
            subject = cursor.fetchone()

            if subject:
                subject_data = {
                    'id': subject[0],
                    'subject_name': subject[1],
                    'class_obj_name': subject[2],
                    'class_obj_id': subject[3],
                }
                # ✅ Set the subject head name
                subject_head_name = subject[4] if subject[4] else "Subject Head"
    except Exception as e:
        print(f"Database Error: {e}") # Added catch for the try block

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')

            if action == "create_quiz":
                quiz_name = data.get('quiz_name')
                class_id = data.get('class_id')
                quiz_id = data.get('quiz_id')
                questions = data.get('questions', [])

                with connection.cursor() as cursor:
                    if quiz_id:  # UPDATE existing
                        cursor.execute("UPDATE main_quizzes SET name=%s, class_obj_id=%s WHERE id=%s", [quiz_name, class_id, quiz_id])
                        cursor.execute("DELETE FROM main_quizquestions WHERE quiz_id = %s", [quiz_id])
                    else:  # INSERT new
                        cursor.execute("INSERT INTO main_quizzes (name, subject_id, class_obj_id, created_at) VALUES (%s, %s, %s, %s)", [quiz_name, subject_id, class_id, datetime.now()])
                        # SQLite specific call to get the ID of the quiz we just created
                        cursor.execute("SELECT last_insert_rowid()") 
                        quiz_id = cursor.fetchone()[0]

                    # Insert Questions
                    if questions:
                        for q in questions:
                            cursor.execute("""
                                INSERT INTO main_quizquestions (quiz_id, question, option_a, option_b, option_c, option_d, correct_option) 
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """, [quiz_id, q['question'], q['option_a'], q['option_b'], q['option_c'], q['option_d'], q['correct_option']])

                    # Create announcement for students
                    announcement_message = f"📢 New Quiz Alert: '{quiz_name}' has been posted for {subject_data['subject_name']}. Please attend it at your earliest convenience. Good luck!"
                    cursor.execute(
                        "INSERT INTO main_announcements (message, created_at, class_obj_id, is_read) VALUES (%s, %s, %s, FALSE)",
                        [announcement_message, datetime.now(), class_id]
                    )

                    cursor.execute("SELECT email FROM main_students WHERE class_obj_id = %s", [class_id])
                    student_emails = [row[0] for row in cursor.fetchall() if row[0]]

                    # Logic to notify students (Example)
                    if student_emails:
                        print(f"Notifying {len(student_emails)} students about quiz: {quiz_name}")
                        notify_students_new_quiz(
                            email_list=student_emails,
                            subject_name=subject_data['subject_name'], # From your fetch subject logic
                            quiz_name=quiz_name
                        )
                        #send mail notification fro studnets
                return JsonResponse({'quiz_id': quiz_id, 'message': 'Registry Synced Successfully'})

            elif action == "add_questions":
                questions = data.get('questions')
                with connection.cursor() as cursor:
                    for q in questions:
                        cursor.execute("""
                            INSERT INTO main_quizquestions (quiz_id, question, option_a, option_b, option_c, option_d, correct_option) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, [q['quiz_id'], q['question'], q['option_a'], q['option_b'], q['option_c'], q['option_d'], q['correct_option']])
                return JsonResponse({'message': 'Questions synced'})

            elif action == "delete_quiz":
                quiz_id = data.get('quiz_id')
                with connection.cursor() as cursor:
                    cursor.execute("DELETE FROM main_quizquestions WHERE quiz_id = %s", [quiz_id])
                    cursor.execute("DELETE FROM main_quizzes WHERE id = %s", [quiz_id])
                return JsonResponse({'message': 'Quiz deleted'})

            elif action == "get_quiz_data":
                quiz_id = data.get('quiz_id')
                with connection.cursor() as cursor:
                    cursor.execute("SELECT name, class_obj_id FROM main_quizzes WHERE id = %s", [quiz_id])
                    quiz = cursor.fetchone()
                    cursor.execute("SELECT question, option_a, option_b, option_c, option_d, correct_option FROM main_quizquestions WHERE quiz_id = %s", [quiz_id])
                    qs = cursor.fetchall()
                
                return JsonResponse({
                    'name': quiz[0], 'class_id': quiz[1],
                    'questions': [{'question': q[0], 'option_a': q[1], 'option_b': q[2], 'option_c': q[3], 'option_d': q[4], 'correct_option': q[5]} for q in qs]
                })

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    # Fetch assigned classes and quizzes for rendering
    with connection.cursor() as cursor:
        cursor.execute("SELECT c.id, c.class_abbreviation FROM main_classes c JOIN main_subjects s ON c.id = s.class_obj_id WHERE s.id = %s", [subject_id])
        classes = cursor.fetchall()
        
        cursor.execute("""
            SELECT mq.id, mq.name, c.class_abbreviation, COUNT(mqq.id),
                   (SELECT COUNT(DISTINCT ms.id) 
                    FROM main_students ms 
                    WHERE ms.class_obj_id = mq.class_obj_id) as total_students,
                   (SELECT COUNT(DISTINCT mr.student_id) 
                    FROM main_quizresponse mr
                    JOIN main_quizquestions mqq2 ON mr.question_id = mqq2.id
                    WHERE mqq2.quiz_id = mq.id) as attended_students
            FROM main_quizzes mq
            JOIN main_classes c ON mq.class_obj_id = c.id
            LEFT JOIN main_quizquestions mqq ON mq.id = mqq.quiz_id
            WHERE mq.subject_id = %s
            GROUP BY mq.id, mq.name, c.class_abbreviation
            ORDER BY mq.id DESC
        """, [subject_id])
        quizzes = cursor.fetchall()

    return render(request, 'subject_head/subject_head_quiz.html', {
        'classes': classes, 'quizzes': quizzes,
         'subject_head': subject_head_name
    })






# def subject_head_quiz_edit(request, quiz_id):
#     subject_id = request.session.get('subject_id')
#     if not subject_id:
#         return redirect('subject_head_login')

#     # Fetch quiz details
#     with connection.cursor() as cursor:
#         cursor.execute("""
#             SELECT id, name 
#             FROM main_quizzes 
#             WHERE id = %s AND subject_id = %s
#         """, [quiz_id, subject_id])
#         quiz = cursor.fetchone()

#     if not quiz:
#         messages.error(request, "Quiz not found or access denied.")
#         return redirect('subject_head_quiz')

#     # Fetch questions for the quiz
#     with connection.cursor() as cursor:
#         cursor.execute("""
#             SELECT id, question, option_a, option_b, option_c, option_d, correct_option 
#             FROM main_quizquestions 
#             WHERE quiz_id = %s
#         """, [quiz_id])
#         questions = cursor.fetchall()

#     return render(request, 'subject_head/subject_head_quiz_edit.html', {
#         'quiz': quiz,
#         'questions': questions
#     })



from datetime import datetime
import os  # Required for path operations
from datetime import datetime  # Required for datetime.now()
from django.db import connection
from django.shortcuts import render, redirect
from django.contrib import messages

def subject_head_studys(request):
    # Ensure the user is a subject head
    subject_id = request.session.get('subject_id')
    if not subject_id:
        print("Redirecting: No subject_id found in session.")
        return redirect('subject_head_login')

    # Fetch subject head name for display
    subject_data = None
    subject_head_name = "Subject Head"  # Default value

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT s.id, s.subject_name, c.class_name, s.class_obj_id, s.subject_head 
                FROM main_subjects s
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s
            """, [subject_id])
            subject = cursor.fetchone()

            if subject:
                subject_data = {
                    'id': subject[0],
                    'subject_name': subject[1],
                    'class_obj_name': subject[2],
                    'class_obj_id': subject[3],
                }
                subject_head_name = subject[4] if subject[4] else "Subject Head"

        # Fetch class_obj_id
        with connection.cursor() as cursor:
            cursor.execute("SELECT class_obj_id FROM main_subjects WHERE id = %s", [subject_id])
            class_obj = cursor.fetchone()
        
        class_obj_id = class_obj[0] if class_obj else None

        # Fetch uploaded documents
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, file_url, announcement, created_at 
                FROM main_studymaterials 
                WHERE subject_id = %s AND class_obj_id = %s
                ORDER BY created_at DESC
            """, [subject_id, class_obj_id])
            docs = cursor.fetchall()

        if request.method == "POST":
            announcement = request.POST.get('announcement', '')
            file = request.FILES.get('file_url') 

            file_url = None 

            if file and class_obj_id and subject_id:
                # Use os.path.join for cross-platform safety
                upload_dir = os.path.join("media", "uploads")
                if not os.path.exists(upload_dir):
                    os.makedirs(upload_dir)

                file_path = os.path.join(upload_dir, file.name)

                with open(file_path, 'wb+') as destination:
                    for chunk in file.chunks():
                        destination.write(chunk)

                # Store the relative URL for the browser (e.g., /media/uploads/file.pdf)
                file_url = f"media/uploads/{file.name}"

            # Insert data
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO main_studymaterials (file_url, created_at, class_obj_id, subject_id, announcement) 
                    VALUES (%s, %s, %s, %s, %s)
                """, [file_url, datetime.now(), class_obj_id, subject_id, announcement if announcement else None])

            messages.success(request, "Study material uploaded successfully!")
            return redirect('subject_head_studys')

    except Exception as e:
        print(f"Error: {e}")
        messages.error(request, "An error occurred.")

    return render(request, 'subject_head/subject_head_study.html', {
        'docs': docs,
        'subject_name': subject_head_name
    })



def subject_head_marks(request):
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')

    subject_data = None
    student_list = []
    subject_head_name = "Subject Head"  # Default value

    try:
        with connection.cursor() as cursor:
            # ✅ Updated Query: Now also fetches s.subject_head
            cursor.execute("""
                SELECT s.id, s.subject_name, c.class_name, s.class_obj_id, s.subject_head 
                FROM main_subjects s
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s
            """, [subject_id])
            subject = cursor.fetchone()

            if subject:
                subject_data = {
                    'id': subject[0],
                    'subject_name': subject[1],
                    'class_obj_name': subject[2],
                    'class_obj_id': subject[3],
                }
                # ✅ Set the subject head name
                subject_head_name = subject[4] if subject[4] else "Subject Head"

                # Handle form submission to update marks
                if request.method == "POST":
                    success = True
                    for key, value in request.POST.items():
                        if key.startswith("marks_"):
                            student_id = key.split("_")[1]
                            mark_value = value.strip()
                            try:
                                mark_value = float(mark_value)
                                cursor.execute("""
                                    SELECT id FROM main_studentevaluation
                                    WHERE student_id = %s AND subject_id = %s
                                """, [student_id, subject_data['id']])
                                evaluation = cursor.fetchone()

                                if evaluation:
                                    cursor.execute("""
                                        UPDATE main_studentevaluation 
                                        SET marks_percentage = %s
                                        WHERE student_id = %s AND subject_id = %s
                                    """, [mark_value, student_id, subject_data['id']])
                                else:
                                    cursor.execute("""
                                        INSERT INTO main_studentevaluation (student_id, subject_id, marks_percentage)
                                        VALUES (%s, %s, %s)
                                    """, [student_id, subject_data['id'], mark_value])
                            except ValueError:
                                success = False

                    if success:
                        messages.success(request, "✅ Marks updated successfully!", extra_tags='marks_success')
                    else:
                        messages.error(request, "❌ Failed to update marks. Please try again.", extra_tags='marks_error')

                # Fetch students of this class
                cursor.execute("SELECT id, roll_no, name FROM main_students WHERE class_obj_id = %s", [subject_data['class_obj_id']])
                students = cursor.fetchall()

                for student in students:
                    student_id = student[0]
                    cursor.execute("""
                        SELECT marks_percentage FROM main_studentevaluation 
                        WHERE student_id = %s AND subject_id = %s
                    """, [student_id, subject_data['id']])
                    mark = cursor.fetchone()

                    student_list.append({
                        'id': student_id,
                        'roll_no': student[1],
                        'name': student[2],
                        'mark_percentage': mark[0] if mark else None
                    })

    except Exception as e:
        print(f"Database error: {e}")
        messages.error(request, "An unexpected error occurred.", extra_tags='marks_error')

    return render(request, 'subject_head/subject_head_marks.html', {
        'subject': subject_data,
        'students': student_list,
        'subject_name': subject_data['subject_name'] if subject_data else 'default',
        'subject_head_name': subject_head_name,  # ✅ Added to context
    })


def download_marks_template(request, subject_id):
    print(f"🔍 Fetching details for Subject ID: {subject_id}")

    with connection.cursor() as cursor:
        # Fetch subject name and class ID
        cursor.execute("""
            SELECT subject_name, class_obj_id FROM main_subjects WHERE id = %s
        """, [subject_id])
        subject_data = cursor.fetchone()

        if not subject_data:
            print("❌ Invalid subject ID! No data found.")
            return HttpResponse("Invalid subject ID", status=400)

        subject_name, class_id = subject_data
        print(f"✅ Retrieved Subject Name: {subject_name}, Class ID: {class_id}")

        cursor.execute("""
            SELECT s.roll_no, s.name
            FROM main_students s
            WHERE s.class_obj_id = %s
        """, [class_id])

        students_data = cursor.fetchall()
        print(f"✅ Found {len(students_data)} students in Class ID: {class_id}")

    # Create a new Excel workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks Template"
    print("✅ Created new Excel workbook.")

    # Define column headers
    headers = ["Roll No", "Student Name", "Subject Name", "Mark Percentage"]
    ws.append(headers)
    print("✅ Added headers:", headers)

    # Store Excel data for debugging
    excel_data = [headers]

    # Write student data (mark_percentage column is left empty)
    for roll_no, name in students_data:
        row = [roll_no, name, subject_name, ""]
        ws.append(row)
        excel_data.append(row)

    print("✅ Populated Excel sheet with student details.")

    # Print final contents of the dynamically created Excel
    print("\n📋 Final Excel Content:")
    for row in excel_data:
        print(row)

    # Prepare the response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Marks_Template_{subject_name}.xlsx"'

    # Save the workbook to response
    wb.save(response)
    print("✅ Excel file generated and sent as response.")

    return response



import openpyxl
from django.shortcuts import redirect
from django.contrib import messages
from django.db import connection, transaction

def upload_marks(request):
    if request.method == "POST" and request.FILES.get("file"):
        excel_file = request.FILES["file"]
        print(f"📂 Received file: {excel_file.name}")

        try:
            # Open the uploaded Excel file
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            print(f"✅ Opened Excel file. Sheet: {ws.title}")

            # Read headers
            headers = [str(cell.value).strip() for cell in ws[1] if cell.value]
            print(f"📑 Extracted Headers: {headers}")

            expected_headers = ["Roll No", "Student Name", "Subject Name", "Mark Percentage"]

            # Case-insensitive header validation
            if [h.lower() for h in headers] != [eh.lower() for eh in expected_headers]:
                print(f"❌ Invalid file format. Headers do not match. Expected: {expected_headers}, Found: {headers}")
                messages.error(request, "❌ Invalid file format! Please use the correct template.", extra_tags="marks_error")
                return redirect("subject_head_marks")

            print("✅ Headers validated. Processing rows...")

            # Fetch subject_id from session
            subject_id = request.session.get("subject_id")
            if not subject_id:
                print("❌ Subject ID not found in session. Redirecting to login.")
                messages.error(request, "❌ Please log in again.", extra_tags="marks_error")
                return redirect("subject_head_login")

            print(f"🎯 Subject ID identified from session: {subject_id}")

            rows_processed = 0

            # Start transaction for atomic updates
            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_values = [cell for cell in row if cell is not None]  # Remove None values

                    # Ensure we have exactly 4 values (skip rows with extra/missing columns)
                    if len(row_values) != 4:
                        print(f"⚠️ Skipping row due to incorrect column count: {row_values}")
                        continue

                    roll_no, student_name, subject_name, mark_percentage = row_values
                    print(f"🔍 Processing: Roll No: {roll_no}, Student: {student_name}, Subject: {subject_name}, Marks: {mark_percentage}")

                    # Validate mark_percentage
                    if mark_percentage is None or not isinstance(mark_percentage, (int, float)):
                        print(f"⚠️ Skipping Roll No {roll_no} - Invalid marks data.")
                        continue

                    with connection.cursor() as cursor:
                        # Get student_id using roll_no
                        cursor.execute("SELECT id FROM main_students WHERE roll_no = %s", [roll_no])
                        student_data = cursor.fetchone()

                        if not student_data:
                            print(f"❌ Skipping Roll No {roll_no} - Student not found.")
                            continue

                        student_id = student_data[0]

                        # Update or insert marks into main_studentevaluation
                        cursor.execute("""
                            UPDATE main_studentevaluation 
                            SET marks_percentage = %s 
                            WHERE student_id = %s AND subject_id = %s
                        """, [mark_percentage, student_id, subject_id])

                        if cursor.rowcount > 0:
                            print(f"🔄 Updated StudentEvaluation for Roll No {roll_no}: {mark_percentage}%")
                        else:
                            cursor.execute("""
                                INSERT INTO main_studentevaluation (student_id, subject_id, marks_percentage) 
                                VALUES (%s, %s, %s)
                            """, [student_id, subject_id, mark_percentage])
                            print(f"🆕 Inserted StudentEvaluation for Roll No {roll_no}: {mark_percentage}%")

                    rows_processed += 1

            print(f"✅ Finished processing {rows_processed} rows.")
            messages.success(request, f"✅ Successfully updated {rows_processed} student marks.", extra_tags="marks_success")

        except Exception as e:
            print(f"❌ Error processing file: {e}")
            messages.error(request, "❌ An error occurred while processing the file. Please try again.", extra_tags="marks_error")

        return redirect("subject_head_marks")

    print("❌ No file uploaded or invalid request method.")
    messages.error(request, "Please upload a valid Excel file.")
    return redirect("subject_head_marks")



def subject_head_attendance(request):
    # Use the session to identify the teacher/subject head
    subject_id = request.session.get('subject_id')
    if not subject_id:
        return redirect('subject_head_login')

    with connection.cursor() as cursor:
        # Fetch initial faculty info
        cursor.execute("""
            SELECT s.subject_name, s.subject_head, s.class_obj_id 
            FROM main_subjects s WHERE s.id = %s
        """, [subject_id])
        subject_info = cursor.fetchone()
        subject_name, subject_head_name, assigned_class_id = subject_info

        # 2. Fetch the assigned class for this Subject Head
        cursor.execute("""
            SELECT c.id, c.class_name 
            FROM main_classes c
            INNER JOIN main_subjects s ON c.id = s.class_obj_id
            WHERE s.id = %s
        """, [subject_id])
        faculty_classes = cursor.fetchall()

        # 3. Use the assigned class for this Subject Head
        target_class_id = assigned_class_id

        # Fetch students for the target class
        cursor.execute("""
            SELECT id, roll_no, name FROM main_students 
            WHERE class_obj_id = %s ORDER BY roll_no ASC
        """, [target_class_id])
        students = cursor.fetchall()
        
        # Create a dictionary mapping student_id to their data for easy lookup
        students_dict = {student[0]: {'roll_no': student[1], 'name': student[2]} for student in students}
        
        # Get selected date and hour from request (if any)
        selected_date = request.GET.get('attendance_date')
        selected_hour = request.GET.get('hour')
        
        # Fetch existing attendance records for the selected date and hour
        students_with_attendance = set()
        if students and selected_date and selected_hour:
            student_ids = [s[0] for s in students]
            placeholders = ','.join(['%s'] * len(student_ids))
            cursor.execute(f"""
                SELECT DISTINCT student_id FROM main_attendance
                WHERE student_id IN ({placeholders})
                AND attendance_date = %s AND hour = %s
            """, student_ids + [selected_date, selected_hour])
            existing_attendance = cursor.fetchall()
            students_with_attendance = {row[0] for row in existing_attendance}

        # --- POST Logic: Save Attendance ---
        if request.method == "POST":
            attendance_date = request.POST.get('attendance_date')
            hour = request.POST.get('hour')
            
            if not attendance_date or not hour:
                messages.error(request, "Please select both date and hour.")
                return redirect('subject_head_attendance')
            
            # Process attendance for each student
            attendance_saved = 0
            affected_students = set()  # Track students whose attendance was updated
            duplicate_errors = []  # Track students with duplicate attendance
            
            with connection.cursor() as post_cursor:
                for student in students:
                    student_id = student[0]
                    attendance_key = f"attendance_{student_id}"
                    status = request.POST.get(attendance_key)
                    
                    if status in ['present', 'absent']:
                        # Check if attendance already exists for this combination
                        post_cursor.execute("""
                            SELECT id FROM main_attendance 
                            WHERE student_id = %s AND subject_id = %s 
                            AND attendance_date = %s AND hour = %s
                        """, [student_id, subject_id, attendance_date, hour])
                        
                        existing = post_cursor.fetchone()
                        
                        if existing:
                            # Attendance already marked - add to error list
                            duplicate_errors.append(f"{student[1]} - {student[2]} (Roll No: {student[1]}, Hour: {hour})")
                        else:
                            # Insert new attendance record
                            affected_students.add(student_id)  # Add to affected students set
                            post_cursor.execute("""
                                INSERT INTO main_attendance 
                                (attendance_date, hour, status, student_id, subject_id, created_at)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            """, [attendance_date, hour, status, student_id, subject_id, now()])
                            
                            attendance_saved += 1
            
            # Show error for duplicates
            if duplicate_errors:
                messages.error(request, f"⚠️ Already marked for {len(duplicate_errors)} student(s)")
            
            # Update StudentEvaluation with new attendance percentages for affected students
            if affected_students:
                print(f"\n📊 Updating StudentEvaluation for {len(affected_students)} students...")
                with connection.cursor() as eval_cursor:
                    for student_id in affected_students:
                        # Calculate attendance percentage for this student across ALL subjects
                        eval_cursor.execute("""
                            SELECT 
                                SUM(CASE WHEN status = 'present' THEN 1 ELSE 0 END) * 100.0 / COUNT(*)
                            FROM main_attendance
                            WHERE student_id = %s
                        """, [student_id])
                        attendance_percentage = eval_cursor.fetchone()[0] or 0
                        
                        print(f"   Student {student_id}: {attendance_percentage:.2f}%")
                        
                        # Update or create StudentEvaluation record
                        # First check if evaluation exists for this student and subject
                        eval_cursor.execute("""
                            SELECT id FROM main_studentevaluation
                            WHERE student_id = %s AND subject_id = %s
                        """, [student_id, subject_id])
                        
                        eval_record = eval_cursor.fetchone()
                        
                        if eval_record:
                            # Update existing evaluation
                            eval_cursor.execute("""
                                UPDATE main_studentevaluation
                                SET attendance_percentage = %s
                                WHERE student_id = %s AND subject_id = %s
                            """, [round(attendance_percentage, 2), student_id, subject_id])
                        else:
                            # Create new evaluation record with default values for other fields
                            eval_cursor.execute("""
                                INSERT INTO main_studentevaluation
                                (student_id, subject_id, attendance_percentage, study_time_rating, 
                                 sleep_time_rating, class_participation_rating, academic_activity_rating, 
                                 marks_percentage)
                                VALUES (%s, %s, %s, NULL, NULL, NULL, NULL, NULL)
                            """, [student_id, subject_id, round(attendance_percentage, 2)])
                        
                        print(f"✅ Updated StudentEvaluation for student {student_id}")
            
            if attendance_saved > 0:
                if duplicate_errors:
                    messages.warning(request, f"✅ Attendance saved for {attendance_saved} students (with {len(duplicate_errors)} duplicates skipped)")
                else:
                    messages.success(request, f"✅ Attendance saved successfully for {attendance_saved} students!")
            elif not duplicate_errors:
                messages.warning(request, "No attendance was marked. Please select Present or Absent for each student.")
            
            return redirect('subject_head_attendance')

    context = {
        'subject_name': subject_name,
        'subject_head_name': subject_head_name,
        'students': students,
        'faculty_classes': faculty_classes,
        'students_with_attendance': students_with_attendance,  # For visual indicators
    }
    
    # Check if the request is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Return only the table body for AJAX updates
        from django.template.loader import render_to_string
        table_html = render_to_string(
            'subject_head/partials/attendance_table_body.html',
            context
        )
        return HttpResponse(table_html)
    
    return render(request, 'subject_head/subject_head_attendance.html', context)



def subject_head_attendance_history(request):
    subject_id = request.session.get('subject_id')

    if not subject_id:
        print("❌ Subject ID not found in session. Redirecting to login.")
        return redirect('subject_head_login')

    print(f"✅ Subject ID found: {subject_id}")
    sub_data = Subjects.objects.get(id=subject_id)
    subject_head_name = sub_data.subject_head

    # Get filter parameters from request
    filter_date = request.GET.get('date', '').strip()
    
    # Set default date to today if not provided
    if not filter_date:
        from datetime import datetime
        filter_date = datetime.now().strftime('%Y-%m-%d')
    
    filter_student = request.GET.get('student', '').strip()
    selected_hour = request.GET.get('hour', '').strip()
    
    # Pagination parameters
    page = int(request.GET.get('page', 1))
    per_page = 10

    print(f"🔍 Filter Date: {filter_date}, Filter Student: {filter_student}, Selected Hour: {selected_hour}")

    # Base SQL query with correct table name, ordered by date
    sql_query = """
        SELECT 
            a.id, 
            a.attendance_date AS date, 
            a.hour, 
            s.roll_no, 
            COALESCE(s.name, '') AS student_name,  -- Handle NULL names
            a.status
        FROM main_attendance a
        JOIN main_students s ON a.student_id = s.id
        JOIN main_subjects sub ON a.subject_id = sub.id
        WHERE sub.id = %s
    """
    
    sql_params = [subject_id]

    # Apply filters dynamically
    if filter_date:
        sql_query += " AND a.attendance_date = %s"
        sql_params.append(filter_date)

    if filter_student:
        sql_query += " AND COALESCE(s.name, '') LIKE %s"  # Avoid NULL issues
        sql_params.append(f"%{filter_student}%")
    
    # Filter by selected hour if provided
    if selected_hour:
        sql_query += " AND a.hour = %s"
        sql_params.append(selected_hour)

    # Order by date (newest first) and hour (ascending)
    sql_query += " ORDER BY a.attendance_date DESC, a.hour ASC"

    # Execute raw SQL query
    with connection.cursor() as cursor:
        cursor.execute(sql_query, sql_params)
        all_rows = cursor.fetchall()

    print(f"✅ Retrieved {len(all_rows)} attendance records.")

    # Get available hours for the radio button selection (from today's date or filtered date)
    hours_query = """
        SELECT DISTINCT a.hour
        FROM main_attendance a
        JOIN main_subjects sub ON a.subject_id = sub.id
        WHERE sub.id = %s
    """
    hours_params = [subject_id]
    
    if filter_date:
        hours_query += " AND a.attendance_date = %s"
        hours_params.append(filter_date)
    
    hours_query += " ORDER BY a.hour ASC"
    
    with connection.cursor() as cursor:
        cursor.execute(hours_query, hours_params)
        available_hours = [row[0] for row in cursor.fetchall()]

    # Check if there are any records for today/date filter
    today_attendance = len(available_hours) > 0

    # Convert query result to a list of dictionaries
    all_records = [
        {"id": row[0], "date": row[1], "hour": row[2], "roll_no": row[3], "student_name": row[4], "status": row[5]}
        for row in all_rows
    ]
    
    # Pagination logic
    total_records = len(all_records)
    total_pages = (total_records + per_page - 1) // per_page
    
    # Ensure page is within valid range
    page = max(1, min(page, total_pages)) if total_pages > 0 else 1
    
    # Calculate start and end indices
    start_index = (page - 1) * per_page
    end_index = min(start_index + per_page, total_records)
    
    # Get records for current page
    attendance_records = all_records[start_index:end_index]
    
    # Generate page range for pagination controls
    page_range = []
    if total_pages <= 7:
        page_range = list(range(1, total_pages + 1))
    else:
        if page <= 4:
            page_range = list(range(1, 6)) + ['...', total_pages]
        elif page >= total_pages - 3:
            page_range = [1, '...'] + list(range(total_pages - 3, total_pages + 1))
        else:
            page_range = [1, '...', page - 1, page, page + 1, '...', total_pages]

    context = {
        "attendance_records": attendance_records,
        "filter_date": filter_date,
        "filter_student": filter_student,
        "selected_hour": selected_hour,
        "available_hours": available_hours,
        "today_attendance": today_attendance,
        "total_records": total_records,
        "total_pages": total_pages,
        "current_page": page,
        "start_index": start_index + 1,
        "end_index": end_index,
        "has_previous": page > 1,
        "has_next": page < total_pages,
        "previous_page": page - 1,
        "next_page": page + 1,
        "page_range": page_range,
        "subject_head_name": subject_head_name,
    }

    return render(request, 'subject_head/subject_head_attendance_history.html', context)



def update_attendance(request):
    if request.method == "POST":
        attendance_ids = request.POST.getlist("attendance_ids[]")  # List of attendance record IDs
        statuses = request.POST.getlist("status[]")  # List of new statuses

        print("\n--- DEBUG: Incoming Attendance Update Request ---")
        print(f"Attendance IDs: {attendance_ids}")
        print(f"New Statuses: {statuses}")

        try:
            with connection.cursor() as cursor:
                for att_id, status in zip(attendance_ids, statuses):
                    # Update attendance status in the database
                    cursor.execute("""
                        UPDATE main_attendance
                        SET status = %s
                        WHERE id = %s
                    """, [status, att_id])
                    print(f"✅ Updated attendance ID {att_id} to status: {status}")

                # Step 1: Get all unique student IDs that were updated
                if attendance_ids:
                    placeholders = ','.join(['%s'] * len(attendance_ids))
                    cursor.execute(f"""
                        SELECT DISTINCT student_id FROM main_attendance WHERE id IN ({placeholders})
                    """, attendance_ids)
                    student_ids = [row[0] for row in cursor.fetchall()]
                else:
                    student_ids = []

                print(f"\n🔍 Students affected: {student_ids}")

                for student_id in student_ids:
                    # Step 2: Recalculate attendance percentage
                    cursor.execute("""
                        SELECT 
                            SUM(CASE WHEN status = 'present' THEN 1 ELSE 0 END) * 100.0 / COUNT(*)
                        FROM main_attendance
                        WHERE student_id = %s
                    """, [student_id])
                    attendance_percentage = cursor.fetchone()[0] or 0  # Ensure default value if no records found

                    print(f"📊 Student {student_id} new attendance percentage: {attendance_percentage:.2f}%")

                    # Step 3: Update `student_evaluations` table
                    cursor.execute("""
                        UPDATE main_studentevaluation
                        SET attendance_percentage = %s
                        WHERE student_id = %s
                    """, [attendance_percentage, student_id])

                    print(f"✅ Updated student_evaluations for student {student_id}")

            messages.success(request, "✅ Attendance updated successfully!")
        except Exception as e:
            messages.error(request, f"❌ Error updating attendance: {str(e)}")
            print(f"\n❌ ERROR: {str(e)}")

        return redirect("subject_head_attendance_history")

    messages.error(request, "❌ Invalid request.")
    return redirect("update_attendance")



def delete_attendance(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            record_id = data.get("id")

            if not record_id:
                return JsonResponse({"success": False, "error": "Invalid record ID."})

            print(f"Deleting attendance record ID: {record_id}")

            with connection.cursor() as cursor:
                # Fetch student_id and subject_id before deleting
                cursor.execute("SELECT student_id, subject_id FROM main_attendance WHERE id = %s", [record_id])
                result = cursor.fetchone()

                if not result:
                    return JsonResponse({"success": False, "error": "Attendance record not found."})

                student_id, subject_id = result
                print(f"Found record - Student ID: {student_id}, Subject ID: {subject_id}")

                # Delete the attendance record
                cursor.execute("DELETE FROM main_attendance WHERE id = %s", [record_id])
                print(f"Attendance record {record_id} deleted successfully.")

                # Calculate new attendance percentage
                cursor.execute("""
                    SELECT COUNT(*) FROM main_attendance
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                total_classes = cursor.fetchone()[0]

                cursor.execute("""
                    SELECT COUNT(*) FROM main_attendance
                    WHERE student_id = %s AND subject_id = %s AND status = 'present'
                """, [student_id, subject_id])
                total_present = cursor.fetchone()[0]

                if total_classes > 0:
                    new_attendance_percentage = (total_present / total_classes) * 100
                else:
                    new_attendance_percentage = 0  # No attendance records left

                print(f"Updated Attendance Percentage: {new_attendance_percentage}%")

                # Update the student_evaluations table
                cursor.execute("""
                    UPDATE main_studentevaluation
                    SET attendance_percentage = %s
                    WHERE student_id = %s AND subject_id = %s
                """, [new_attendance_percentage, student_id, subject_id])

                print(f"Attendance percentage updated for Student ID {student_id}.")

            return JsonResponse({"success": True})

        except Exception as e:
            print(f"Error deleting attendance: {e}")
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method."})



def subject_head_evaluation(request):
    subject_id = request.session.get('subject_id')

    if not subject_id:
        print("❌ Subject ID not found in session. Redirecting to login.")
        return redirect('subject_head_login')

    # 1. Fetch Subject Head Name
    with connection.cursor() as cursor:
        cursor.execute("SELECT subject_head FROM main_subjects WHERE id = %s", [subject_id])
        subject_head_row = cursor.fetchone()
        subject_head_name = subject_head_row[0] if subject_head_row else "Subject Head"

    # 2. Fetch students for the given subject
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT main_students.id, main_students.name, main_students.roll_no
            FROM main_students
            JOIN main_classes ON main_students.class_obj_id = main_classes.id
            JOIN main_subjects ON main_classes.id = main_subjects.class_obj_id
            WHERE main_subjects.id = %s
        """, [subject_id])
        students = cursor.fetchall()

    student_list = []
    student_ids = [row[0] for row in students]
    ratings_dict = {}

    # 3. Fetch existing ratings (Fixing the Tuple Error here)
    if student_ids:
        # Generate placeholders: "%s, %s, %s" based on number of IDs
        placeholders = ', '.join(['%s'] * len(student_ids))
        query = f"""
            SELECT student_id, academic_activity_rating, class_participation_rating
            FROM main_studentevaluation
            WHERE subject_id = %s AND student_id IN ({placeholders})
        """
        
        with connection.cursor() as cursor:
            # Pass subject_id and the flattened list of student_ids
            cursor.execute(query, [subject_id] + student_ids)
            ratings = cursor.fetchall()
            
            for rating in ratings:
                ratings_dict[rating[0]] = {
                    "academic_activity_rating": rating[1] if rating[1] is not None else 0,
                    "class_participation_rating": rating[2] if rating[2] is not None else 0
                }

    for row in students:
        student_id, name, roll_no = row
        student_list.append({
            'id': student_id,
            'name': name,
            'roll_no': roll_no,
            'academic_activity_rating': ratings_dict.get(student_id, {}).get("academic_activity_rating", 0),
            'class_participation_rating': ratings_dict.get(student_id, {}).get("class_participation_rating", 0),
        })

    # 4. Handle Rating Submission
    if request.method == "POST":
        rating_type = request.POST.get("rating_type")
        print("Rating type :", rating_type)
        
        # Ensure only allowed columns are used to prevent SQL injection
        if rating_type not in ['academic_activity_rating', 'class_participation_rating']:
            messages.error(request, "Invalid rating type.")
            return redirect('subject_head_evaluation')

        # Map roll numbers to ratings from the POST data
        # The form submits fields as 'rating_{roll_no}', so we need to extract those
        student_ratings_input = {}
        for key, value in request.POST.items():
            if key.startswith('rating_'):
                roll_no = key.split('_')[1]  # Extract roll number from 'rating_{roll_no}'
                student_ratings_input[roll_no] = value

        with connection.cursor() as cursor:
            for student in student_list:
                s_id = student['id']
                r_no = str(student['roll_no'])
                
                # Get the new rating from POST, or keep current value if not provided
                new_rating = student_ratings_input.get(r_no)
                if new_rating is None:
                    continue  # Skip if this student wasn't in the form

                # Upsert Logic (Check if exists, then Update or Insert)
                cursor.execute("""
                    SELECT id FROM main_studentevaluation
                    WHERE student_id = %s AND subject_id = %s
                """, [s_id, subject_id])
                record = cursor.fetchone()

                if record:
                    cursor.execute(f"""
                        UPDATE main_studentevaluation
                        SET {rating_type} = %s
                        WHERE student_id = %s AND subject_id = %s
                    """, [new_rating, s_id, subject_id])
                else:
                    cursor.execute(f"""
                        INSERT INTO main_studentevaluation (student_id, subject_id, {rating_type})
                        VALUES (%s, %s, %s)
                    """, [s_id, subject_id, new_rating])

        messages.success(request, f"{rating_type.replace('_', ' ').title()} updated successfully!")
        return redirect('subject_head_evaluation')

    return render(request, 'subject_head/subject_head_evaluation.html', {
        'students': student_list,
        'subject_head_name': subject_head_name
    })


import json
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.db import connection
from ml.predict import predict_performance  # Import the performance prediction function

def subject_head_students(request, student_id):
    # Retrieve the subject_id from session
    subject_id = request.session.get('subject_id')
    subjects = Subjects.objects.get(id=subject_id)
    subject_head = subjects.subject_head

    if not subject_id:
        print("[ERROR] No subject_id found in session. Redirecting to subject_head login.")
        return redirect('subject_head_login')

    print(f"\n[INFO] Subject Head Student Performance View Loaded for Student ID: {student_id}")
    print(f"[INFO] Viewing for Subject ID: {subject_id}")

    student_data = {}
    evaluations = {}
    quiz_percentage = 0
    predicted_marks = None
    subject_name = "Unknown Subject"  # Default in case of error

    try:
        with connection.cursor() as cursor:
            # Fetch subject name from subject_id
            cursor.execute("""
                SELECT subject_name FROM main_subjects WHERE id = %s
            """, [subject_id])
            row = cursor.fetchone()
            
            if row:
                subject_name = row[0]
                print(f"[INFO] Subject Name Retrieved: {subject_name}")
            else:
                print("[ERROR] Subject not found for given ID.")

            # Fetch student details
            cursor.execute("""
                SELECT s.id, s.name, s.roll_no, c.class_abbreviation, s.email, c.id AS class_id
                FROM main_students s
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s
            """, [student_id])
            row = cursor.fetchone()

            if row:
                student_data = {
                    "id": row[0],
                    "name": row[1],
                    "roll_no": row[2],
                    "class_name": row[3],
                    "email": row[4] if row[4] else "--",
                    "class_id": row[5]
                }
                print(f"[DEBUG] Student Details: {student_data}")
            else:
                print("[ERROR] Student not found.")
                return redirect('subject_head_dashboard')

        # Fetch student evaluation details for the subject
        print(f"[DEBUG] Fetching evaluation data for Subject ID: {subject_id}...")
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT marks_percentage, attendance_percentage,
                       study_time_rating, sleep_time_rating, 
                       class_participation_rating, academic_activity_rating
                FROM main_studentevaluation
                WHERE student_id = %s AND subject_id = %s
            """, [student_id, subject_id])
            row = cursor.fetchone()

            if row:
                evaluations = {
                    "marks_percentage": round(row[0] or 0, 2),
                    "attendance_percentage": round(row[1] or 0, 2),
                    "study_time_rating": round(row[2] or 0, 2),
                    "sleep_time_rating": round(row[3] or 0, 2),
                    "class_participation_rating": round(row[4] or 0, 2),
                    "academic_activity_rating": round(row[5] or 0, 2),
                }
                print(f"[DEBUG] Evaluation Data: {evaluations}")

                # Call prediction model
                print("[DEBUG] Calling prediction function...")
                predicted_marks = predict_performance(
                    evaluations["attendance_percentage"],
                    evaluations["marks_percentage"],
                    evaluations["class_participation_rating"],
                    evaluations["academic_activity_rating"],
                    evaluations["sleep_time_rating"],
                    evaluations["study_time_rating"]
                )

                # Ensure predicted marks is a float or None
                predicted_marks = round(float(predicted_marks), 2) if predicted_marks is not None else None
                print(f"[DEBUG] Predicted Marks: {predicted_marks}")

            else:
                print("[WARNING] No evaluation data found.")

        # Fetch quiz performance for the subject
        print(f"[DEBUG] Fetching quiz performance for Subject ID: {subject_id}...")
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT COUNT(*) AS total_attempted,
                       SUM(CASE WHEN q.correct_option = r.student_response THEN 1 ELSE 0 END) AS correct_answers
                FROM main_quizresponse r
                JOIN main_quizquestions q ON r.question_id = q.id
                JOIN main_quizzes mq ON q.quiz_id = mq.id
                WHERE r.student_id = %s AND mq.subject_id = %s
            """, [student_id, subject_id])
            row = cursor.fetchone()

            if row:
                total_attempted, correct_answers = row[0], row[1] or 0
                quiz_percentage = round((correct_answers / total_attempted) * 100, 2) if total_attempted > 0 else 0
                print(f"[DEBUG] Quiz Performance: Attempted = {total_attempted}, Correct = {correct_answers}, Percentage = {quiz_percentage}%")
            else:
                print("[WARNING] No quiz data found.")

        # Prepare data for Graph (ensure all values are numeric)
        graph_data = [
            float(evaluations.get("marks_percentage", 0)),
            float(evaluations.get("attendance_percentage", 0)),
            float(evaluations.get("study_time_rating", 0)),
            float(evaluations.get("sleep_time_rating", 0)),
            float(evaluations.get("class_participation_rating", 0)),
            float(evaluations.get("academic_activity_rating", 0)),
            float(quiz_percentage),
            float(predicted_marks) if predicted_marks is not None else 0  # Ensure numeric value for predicted marks
        ]

    except Exception as e:
        print(f"[ERROR] An error occurred: {e}")
        return redirect('error_page')  

    context = {
        "student": student_data,
        "evaluations": evaluations,
        "quiz_percentage": quiz_percentage,
        "predicted_marks": predicted_marks,  # Include predicted marks in context
        "graph_data": json.dumps(graph_data),
        "subject_name": subject_name, # Pass subject name to the template
        'subject_head': subject_head  
    }

    # Check if the request is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({"predicted_marks": predicted_marks})

    print("[INFO] Final Context Data Prepared for Rendering.")
    return render(request, "subject_head/subject_head_students.html", context)





######################################################################################################################

def student_login(request):
    form = StudentLoginForm()

    if request.method == "POST":
        form = StudentLoginForm(request.POST)

        if form.is_valid():
            roll_no = form.cleaned_data.get("roll_no").upper()
            password = form.cleaned_data.get("password")

            # Validate credentials with raw SQL query
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT s.id, s.name, c.institution_id, s.password FROM main_students s
                    JOIN main_users u ON s.user_id = u.id
                    JOIN main_classes c ON s.class_obj_id = c.id
                    WHERE s.roll_no = %s AND u.role = 'student'
                """, [roll_no])
                student = cursor.fetchone()

            if student:
                student_id, student_name, institution_id, db_password = student
                
                # Check password using decryption for encrypted passwords
                if verify_password(password, db_password):
                    request.session['student_id'] = student_id  # Save student ID in session
                    
                    # Create audit log
                    try:
                        from .models import Institution, Students
                        client_ip = get_client_ip(request)
                        institution = Institution.objects.get(institution_id=institution_id)
                        # Get the actual student object to access the user foreign key
                        student_obj = Students.objects.get(id=student_id)
                        AuditLog.objects.create(
                            user=student_obj.user,  # Use the student's user foreign key
                            institution=institution,
                            action='login',
                            ip_address=client_ip,
                            user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
                            details=f'Student login: {student_name} (Roll: {roll_no})',
                            related_object_id=student_id,
                            related_object_type='Students'
                        )
                    except Exception as e:
                        print(f"Student login audit logging failed: {e}")
                    
                    return redirect('student_dashboard')
            else:
                messages.error(request, 'Invalid Roll number or Password!', extra_tags="st_login_error")

    return render(request, 'students/student_login.html', {'form': form})


def student_dashboard(request):
    # Ensure the student is logged in
    if 'student_id' not in request.session:
        return redirect('student_login')
    student_id = request.session['student_id']

    print(f"Using student_id: {student_id} for the query.")

    with connection.cursor() as cursor:
        # Fetch the student's name and class ID
        cursor.execute("""
            SELECT s.name, s.class_obj_id
            FROM main_students s 
            WHERE s.id = %s
        """, [student_id])
        student_data = cursor.fetchone()

        if not student_data:
            return redirect('student_login')  # Redirect if student data is invalid

        student_name, class_id = student_data

        class_name = None
        subject_count = 0  # Default to 0 if no subjects are assigned

        if class_id:  # If the student is assigned to a class
            # Fetch the class name
            cursor.execute("""
                SELECT c.class_abbreviation 
                FROM main_classes c
                WHERE c.id = %s
            """, [class_id])
            class_data = cursor.fetchone()

            if class_data:
                class_name = class_data[0]

            # Count the number of subjects for this class
            cursor.execute("""
                SELECT COUNT(*) 
                FROM main_subjects 
                WHERE class_obj_id = %s
            """, [class_id])
            subject_count = cursor.fetchone()[0]  # Get the count value

    # Pass data to the template
    return render(request, 'students/student_dashboard.html', {
        'student_name': student_name,
        'class_name': class_name if class_name else "Not Assigned",
        'subject_count': subject_count,
    })


def student_profile(request):
    # Ensure the student is logged in
    if 'student_id' not in request.session:
        return redirect('student_login')

    student_id = request.session['student_id']

    if request.method == 'POST':
        # Handle the form submission and update the student details
        name = request.POST.get('name')
        password = request.POST.get('password')

        with connection.cursor() as cursor:
            # Update student profile in the database
            cursor.execute("""
                UPDATE main_students
                SET name = %s
                WHERE id = %s
            """, [name, student_id])

            if password:  # Only update password if a new one is provided
                encrypted_password = encrypt_password(password)
                cursor.execute("""
                    UPDATE main_students
                    SET password = %s
                    WHERE id = %s
                """, [encrypted_password, student_id])

        # Redirect after updating the profile
        messages.success(request, 'Profile updated successfully!')
        return redirect('student_profile')

    with connection.cursor() as cursor:
        # Fetch student profile data including current password
        cursor.execute("""
            SELECT s.name, s.roll_no, c.class_abbreviation, c.class_head, c.email, i.institution_name, s.password
            FROM main_students s
            LEFT JOIN main_classes c ON s.class_obj_id = c.id
            LEFT JOIN main_institution i ON c.institution_id = i.institution_id
            WHERE s.id = %s
        """, [student_id])
        profile_data = cursor.fetchone()

    # If student data is missing (edge case), redirect to login
    if not profile_data:
        return redirect('student_login')

    student_name, roll_no, class_name, teacher_name, teacher_email, institution_name, current_password = profile_data
    
    # Decrypt password for display
    decrypted_password = decrypt_password(current_password)

    return render(request, 'students/student_profile.html', {
        'student_name': student_name,
        'roll_no': roll_no,
        'class_name': class_name if class_name else "Not Assigned",
        'teacher_name': teacher_name if teacher_name else "Not Assigned",
        'teacher_email': teacher_email if teacher_email else "Not Assigned",
        'institution_name': institution_name if institution_name else "Not Assigned",
        'current_password': decrypted_password,  # Use decrypted password
    })


def student_class(request):
    # Ensure the student is logged in
    if 'student_id' not in request.session:
        return redirect('student_login')

    student_id = request.session['student_id']

    with connection.cursor() as cursor:
        # Fetch student details and class head
        cursor.execute("""
            SELECT s.name AS student_name, c.class_name, c.class_head, c.id AS class_id
            FROM main_students s
            LEFT JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [student_id])
        student_details = cursor.fetchone()

        # Fetch list of students in the same class
        cursor.execute("""
            SELECT s.roll_no, s.name
            FROM main_students s
            WHERE s.class_obj_id = (
                SELECT class_obj_id
                FROM main_students
                WHERE id = %s
            )
        """, [student_id])
        students = cursor.fetchall()

        # Fetch subjects assigned to this class
        cursor.execute("""
            SELECT id, subject_name, subject_head, email
            FROM main_subjects
            WHERE class_obj_id = (
                SELECT class_obj_id
                FROM main_students
                WHERE id = %s
            )
        """, [student_id])
        subjects = cursor.fetchall()

        # Fetch announcements for this class
        cursor.execute("""
            SELECT id, message, created_at
            FROM main_announcements
            WHERE class_obj_id = (
                SELECT class_obj_id
                FROM main_students
                WHERE id = %s
            )
            ORDER BY created_at DESC
        """, [student_id])
        announcements = cursor.fetchall()

    # Handle missing student or class data
    if not student_details:
        student_name, class_name, class_head, class_id = "Unknown", "Not Assigned", "Not Assigned", None
    else:
        student_name, class_name, class_head, class_id = student_details

    students_list = [{'roll_no': row[0], 'name': row[1]} for row in students] if students else []
    subjects_list = [{'id': row[0], 'subject_name': row[1], 'subject_head': row[2], 'email': row[3]} for row in subjects] if subjects else []
    announcements_list = [{'id': row[0], 'message': row[1], 'created_at': row[2]} for row in announcements] if announcements else []

    return render(request, 'students/student_class.html', {
        'student_name': student_name,
        'class_name': class_name,
        'class_head': class_head,
        'students': students_list,
        'student_count': len(students_list),
        'subjects': subjects_list,
        'announcements': announcements_list,  # Pass announcements to the template
    })


def student_chat(request):
    student_id = request.session.get('student_id')
    if not student_id:
        return redirect('student_login')

    try:
        student_obj = Students.objects.select_related('class_obj').get(id=student_id)
        student_name = student_obj.name
        student_class_id = student_obj.class_obj.id 
    except (Students.DoesNotExist, AttributeError):
        return redirect('student_login')

    # Get student's user_id
    student_user_id = student_obj.user_id

    # Query for users already in chat
    query_recent_chats = '''
        SELECT DISTINCT u.id, u.role
        FROM main_chat c
        JOIN main_users u ON (c.sender_id = u.id OR c.receiver_id = u.id)
        JOIN main_students s ON (s.user_id = c.sender_id OR s.user_id = c.receiver_id)
        WHERE s.id = %s
        AND u.role IN ('class_head', 'subject_head')
    '''
    
    # REMOVED the "NOT IN" subquery so they appear in Discovery even if active
    query_suggested_users = '''
        SELECT u.id, u.role, 
               CASE 
                   WHEN u.role = 'class_head' THEN c.class_head
                   WHEN u.role = 'subject_head' THEN s.subject_head
               END AS name
        FROM main_users u
        LEFT JOIN main_classes c ON u.id = c.user_id
        LEFT JOIN main_subjects s ON u.id = s.user_id
        WHERE u.role IN ('class_head', 'subject_head')
        AND (
            (u.role = 'class_head' AND c.id = %s) 
            OR 
            (u.role = 'subject_head' AND s.class_obj_id = %s)
        )
    '''

    with connection.cursor() as cursor:
        cursor.execute(query_recent_chats, [student_id])
        user_ids_in_chat = [user[0] for user in cursor.fetchall()]

        cursor.execute(query_suggested_users, [student_class_id, student_class_id])
        all_suggested = cursor.fetchall()

    suggested_class_heads = [u for u in all_suggested if u[1] == 'class_head']
    suggested_subject_heads = [u for u in all_suggested if u[1] == 'subject_head']

    # Fetch names for existing chat users
    class_heads_active = list(Classes.objects.filter(user_id__in=user_ids_in_chat).values_list('user_id', 'class_head'))
    subject_heads_active = list(Subjects.objects.filter(user_id__in=user_ids_in_chat).values_list('user_id', 'subject_head'))

    # Count unread messages for each chat user
    from django.db.models import Q, Count
    unread_counts = {}
    for user_id in user_ids_in_chat:
        unread_count = Chat.objects.filter(
            sender_id=user_id,
            receiver_id=student_user_id,
            is_read=False
        ).count()
        unread_counts[user_id] = unread_count

    # Add unread counts to class_heads and subject_heads
    class_heads_with_unread = []
    for user_id, name in class_heads_active:
        class_heads_with_unread.append((user_id, name, unread_counts.get(user_id, 0)))
    
    subject_heads_with_unread = []
    for user_id, name in subject_heads_active:
        subject_heads_with_unread.append((user_id, name, unread_counts.get(user_id, 0)))

    return render(request, 'students/student_chat.html', {
        'class_heads': class_heads_with_unread,  
        'subject_heads': subject_heads_with_unread,  
        'suggested_class_heads': suggested_class_heads,
        'suggested_subject_heads': suggested_subject_heads,
        'student_name': student_name
    })

def student_chat_user(request, user_id):
    student_id = request.session.get('student_id')
    if not student_id:
        return redirect('student_login')

    # Fetch the user_id for the logged-in student
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT s.user_id
            FROM main_students s
            WHERE s.id = %s
        """, [student_id])
        student_user_id = cursor.fetchone()

    if not student_user_id:
        messages.error(request, 'Student not found.')
        return redirect('student_dashboard')

    logged_in_user_id = student_user_id[0]

    # --- ADDED LOGIC TO FETCH STUDENT NAME ---
    # Fetching the name of the logged-in student from the Students model
    student_info = Students.objects.filter(id=student_id).values_list('name', flat=True).first()
    # -----------------------------------------

    # Mark all messages from this user as read
    Chat.objects.filter(
        sender_id=user_id,
        receiver_id=logged_in_user_id,
        is_read=False
    ).update(is_read=True)

    # Fetch messages between the logged-in student and the selected user
    query = '''
        SELECT message, sender_id, receiver_id, created_at
        FROM main_chat
        WHERE (sender_id = %s AND receiver_id = %s) OR (sender_id = %s AND receiver_id = %s)
        ORDER BY created_at ASC
    '''
    with connection.cursor() as cursor:
        cursor.execute(query, [logged_in_user_id, user_id, user_id, logged_in_user_id])
        messages_fetched = cursor.fetchall()

    # Get the selected user's role
    user = Users.objects.get(id=user_id)
    selected_user_role = user.role

    # Fetch the selected user's name based on their role
    if selected_user_role == 'class_head':
        selected_user_name = Classes.objects.filter(user_id=user_id).values_list('class_head', flat=True).first()
    elif selected_user_role == 'subject_head':
        selected_user_name = Subjects.objects.filter(user_id=user_id).values_list('subject_head', flat=True).first()
    else:
        selected_user_name = 'Unknown'

    # Handle message submission (POST request)
    if request.method == 'POST':
        message_text = request.POST.get('message')
        if message_text:
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO main_chat (sender_id, receiver_id, message, created_at, is_read)
                    VALUES (%s, %s, %s, %s, FALSE)
                """, [logged_in_user_id, user_id, message_text, now()])
            
            # if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            #     return JsonResponse({'status': 'success'})
                
            return redirect('student_chat_user', user_id=user_id)

    return render(request, 'students/student_chat_user.html', {
        'messages': messages_fetched,
        'user_id': user_id,
        'logged_in_user_id': logged_in_user_id,
        'selected_user_name': selected_user_name,
        'selected_user_role': selected_user_role,
        'student_name': student_info,  # Pass the student's name to the frontend
    })



def student_subject_detail(request, subject_id):
    if 'student_id' not in request.session:
        return redirect('student_login')

    student_id = request.session['student_id']

    with connection.cursor() as cursor:
        # Fetch subject details along with class information
        cursor.execute("""
            SELECT s.id, s.subject_name, s.subject_head, s.email, c.class_name, c.class_head
            FROM main_subjects s
            JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [subject_id])
        subject_data = cursor.fetchone()

        if not subject_data:
            return render(request, 'students/student_subject_detail.html', {'error': 'Subject not found'})

        # Fetch student name
        cursor.execute("SELECT name FROM main_students WHERE id = %s", [student_id])
        student_name_data = cursor.fetchone()
        student_name = student_name_data[0] if student_name_data else "Unknown"

        # Fetch all quizzes for this subject - latest first
        cursor.execute("SELECT id, name, created_at FROM main_quizzes WHERE subject_id = %s ORDER BY id DESC", [subject_id])
        quizzes_raw = cursor.fetchall()

        quizzes = []
        # Fixed: Added q_created to match the 3 columns in SELECT
        for q_id, q_name, q_created in quizzes_raw:
            # Count total questions in this quiz
            cursor.execute("SELECT COUNT(*) FROM main_quizquestions WHERE quiz_id = %s", [q_id])
            total_q = cursor.fetchone()[0] or 0
            
            # Check for student responses to determine if attended and calculate score
            cursor.execute("""
                SELECT COUNT(r.id), 
                    SUM(CASE WHEN r.student_response = qq.correct_option THEN 1 ELSE 0 END)
                FROM main_quizresponse r
                JOIN main_quizquestions qq ON r.question_id = qq.id
                WHERE qq.quiz_id = %s AND r.student_id = %s
            """, [q_id, student_id])
            resp_data = cursor.fetchone()
            
            attended_count = resp_data[0] or 0
            correct_count = resp_data[1] or 0
            
            is_attended = attended_count > 0
            score_pct = 0
            if is_attended and total_q > 0:
                score_pct = round((correct_count / total_q) * 100)
                
            quizzes.append({
                'id': q_id,
                'name': q_name,
                'is_attended': is_attended,
                'score_pct': score_pct,
                'created_at': q_created  # Fixed: Use the variable from the loop
            })

        # Fetch study materials for this subject
        cursor.execute("""
            SELECT id, file_url, created_at, announcement 
            FROM main_studymaterials 
            WHERE subject_id = %s
        """, [subject_id])
        studys = [
            {'id': row[0], 'file_url': row[1], 'created_at': row[2], 'announcement': row[3]}
            for row in cursor.fetchall()
        ]

    # Prepare subject dictionary
    subject = {
        'id': subject_data[0],
        'subject_name': subject_data[1],
        'subject_head': subject_data[2],
        'email': subject_data[3],
        'class_name': subject_data[4],
        'class_head': subject_data[5]
    }

    return render(request, 'students/student_subject_detail.html', {
        'subject': subject,
        'student_name': student_name,
        'quizzes': quizzes,  # Pass quizzes to the template
        'studys': studys  # Pass study materials to the template
    })


def student_quiz(request, subject_id, quiz_id):
    if 'student_id' not in request.session:
        return redirect('student_login')

    student_id = request.session['student_id']

    # 1. Fetch student and quiz metadata
    with connection.cursor() as cursor:
        cursor.execute("SELECT name FROM main_students WHERE id = %s", [student_id])
        student_row = cursor.fetchone()
        student_name = student_row[0] if student_row else "Unknown"

        cursor.execute("SELECT name FROM main_quizzes WHERE id = %s", [quiz_id])
        quiz_row = cursor.fetchone()
        quiz_name = quiz_row[0] if quiz_row else "Quiz"

    subject = get_object_or_404(Subjects, id=subject_id)

    # 2. Check for existing attempt
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(*) FROM main_quizresponse 
            WHERE student_id = %s AND question_id IN 
            (SELECT id FROM main_quizquestions WHERE quiz_id = %s)
        """, [student_id, quiz_id])
        quiz_attempted = cursor.fetchone()[0] > 0

    # 3. CASE: SHOW RESULTS (Already Attempted)
    if quiz_attempted:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT q.id, q.question, q.option_a, q.option_b, q.option_c, q.option_d, 
                       q.correct_option, r.student_response 
                FROM main_quizquestions q
                LEFT JOIN main_quizresponse r ON q.id = r.question_id AND r.student_id = %s
                WHERE q.quiz_id = %s
            """, [student_id, quiz_id])

            quiz_results = []
            for row in cursor.fetchall():
                quiz_results.append({
                    'id': row[0], 'question': row[1],
                    'options_list': [('A', row[2]), ('B', row[3]), ('C', row[4]), ('D', row[5])],
                    'correct_option': row[6], 'student_response': row[7],
                    'is_correct': str(row[6]).strip().upper() == str(row[7]).strip().upper()
                })

        score = sum(1 for q in quiz_results if q['is_correct'])
        percentage = (score / len(quiz_results)) * 100 if quiz_results else 0

        return render(request, 'students/student_quiz.html', {
            'student_name': student_name, 'subject': subject, 'quiz_name': quiz_name,
            'quiz_results': quiz_results, 'quiz_attempted': True,
            'score': score, 'total': len(quiz_results), 'percentage': round(percentage, 2)
        })

    # 4. CASE: ATTEND QUIZ (First time)
    if request.method == "POST":
        with connection.cursor() as cursor:
            for key, value in request.POST.items():
                if key.startswith("question_"):
                    q_id = key.split("_")[1]
                    cursor.execute("""
                        INSERT INTO main_quizresponse (student_response, question_id, student_id)
                        VALUES (%s, %s, %s)
                    """, [value, q_id, student_id])
        messages.success(request, 'Quiz submitted successfully! Analyzing your responses...')
        return redirect('student_quiz', subject_id=subject_id, quiz_id=quiz_id)

    # Fetch questions for the form
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, question, option_a, option_b, option_c, option_d FROM main_quizquestions WHERE quiz_id = %s", [quiz_id])
        quiz_questions = []
        for row in cursor.fetchall():
            quiz_questions.append({
                'id': row[0], 'question': row[1],
                'options_list': [('A', row[2]), ('B', row[3]), ('C', row[4]), ('D', row[5])]
            })

    return render(request, 'students/student_quiz.html', {
        'student_name': student_name, 'subject': subject, 'quiz_name': quiz_name,
        'quiz_questions': quiz_questions, 'quiz_attempted': False
    })

def student_performance(request):
    if 'student_id' not in request.session:
        print("[DEBUG] Redirecting to student login (no student_id in session).")
        return redirect('student_login')

    student_id = request.session['student_id']
    subject_id = request.GET.get('subject_id')

    print(f"\n[INFO] Student Performance View Loaded for Student ID: {student_id}")
    
    student_data = {}
    evaluations = {}
    quiz_percentage = 0 
    # attendance_rate = 0  # Initialize attendance_rate
    subjects = []
    selected_subject_name = None

    try:
        with connection.cursor() as cursor:
            # Fetch student details along with class info
            cursor.execute("""
                SELECT s.name, s.roll_no, c.class_abbreviation, s.email, c.id AS class_id
                FROM main_students s
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s
            """, [student_id])
            row = cursor.fetchone()
            
            if row:
                student_data = {
                    "name": row[0],
                    "roll_no": row[1],
                    "class_abbreviation": row[2],
                    "email": row[3],
                    "class_id": row[4]
                }
            else:
                return redirect('student_dashboard')

        # Fetch subjects for the student's class
        if student_data.get("class_id"):
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, subject_name FROM main_subjects WHERE class_obj_id = %s
                """, [student_data["class_id"]])
                subjects = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]

            selected_subject_name = next((s["name"] for s in subjects if str(s["id"]) == subject_id), None)
                
        # # --- ATTENDANCE RATE CALCULATION ---
        # if student_data.get("class_id"):
        #     with connection.cursor() as cursor:
        #         # Fixed: Changed [class_id] to [student_data["class_id"]]
        #         cursor.execute("""
        #             SELECT 
        #                 COALESCE(COUNT(CASE WHEN status = 'present' THEN 1 END) * 100.0 / NULLIF(COUNT(*), 0), 0)
        #             FROM main_attendance 
        #             WHERE student_id = %s
        #         """, [student_id]) # Usually attendance is calculated per student, not class-wide here
        #         attendance_rate = round(cursor.fetchone()[0] or 0, 1)

        # Fetch student evaluation details for the selected subject
        if subject_id:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT marks_percentage, attendance_percentage,
                           study_time_rating, sleep_time_rating, 
                           class_participation_rating, academic_activity_rating
                    FROM main_studentevaluation
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    evaluations = {
                        "marks_percentage": round(row[0] or 0, 2),
                        "attendance_percentage": round(row[1] or 0, 2),
                        "study_time_rating": round(row[2] or 0, 2),
                        "sleep_time_rating": round(row[3] or 0, 2),
                        "class_participation_rating": round(row[4] or 0, 2),
                        "academic_activity_rating": round(row[5] or 0, 2),
                    }

        # Fetch quiz performance for the selected subject
        if subject_id:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT COUNT(*) AS total_attempted,
                           SUM(CASE WHEN q.correct_option = r.student_response THEN 1 ELSE 0 END) AS correct_answers
                    FROM main_quizresponse r
                    JOIN main_quizquestions q ON r.question_id = q.id
                    JOIN main_quizzes mq ON q.quiz_id = mq.id
                    WHERE r.student_id = %s AND mq.subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    total_attempted, correct_answers = row[0], row[1] or 0
                    quiz_percentage = round((correct_answers / total_attempted) * 100, 2) if total_attempted > 0 else 0

        # Prepare data for Graph
        graph_data = [
            float(evaluations.get("marks_percentage", 0)),
            float(evaluations.get("attendance_percentage", 0)),
            float(evaluations.get("study_time_rating", 0)),
            float(evaluations.get("sleep_time_rating", 0)),
            float(evaluations.get("class_participation_rating", 0)),
            float(evaluations.get("academic_activity_rating", 0)),
            float(quiz_percentage)
        ]

    except Exception as e:
        print(f"[ERROR] An error occurred: {e}")
        return redirect('student_dashboard')

    context = {
        "student": student_data,
        "evaluations": evaluations,
        "quiz_percentage": quiz_percentage,
        "subjects": subjects,
        "selected_subject_id": int(subject_id) if subject_id else None,
        "graph_data": json.dumps(graph_data) # Stringified for Chart.js
    }

    return render(request, "students/student_performance.html", context)



def subject_head_quiz_responses(request, quiz_id):
    # Ensure Subject Head is logged in
    subject_id = request.session.get('subject_id')
    if not subject_id: return redirect('subject_head_login')

    with connection.cursor() as cursor:
        # Fetch Student Name, total correct answers, and total questions
        cursor.execute("""
            SELECT s.name, 
                   SUM(CASE WHEN r.student_response = q.correct_option THEN 1 ELSE 0 END) as score,
                   COUNT(q.id) as total_questions
            FROM main_students s
            JOIN main_quizresponse r ON s.id = r.student_id
            JOIN main_quizquestions q ON r.question_id = q.id
            WHERE q.quiz_id = %s
            GROUP BY s.id, s.name
        """, [quiz_id])
        
        results = cursor.fetchall()

    return render(request, 'subject_head/quiz_responses.html', {'results': results})



def student_view_attendance(request):
    if 'student_id' not in request.session:
        return redirect('student_login')
    
    student_id = request.session['student_id']
    # Default to today's date if no date is selected
    selected_date = request.GET.get('date') or timezone.now().strftime('%Y-%m-%d')
    
    student_data = {}
    subjects_list = []
    attendance_by_date = {} 
    
    try:
        with connection.cursor() as cursor:
            # 1. Get Student & Class Info
            cursor.execute("""
                SELECT s.name, s.roll_no, c.id 
                FROM main_students s
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s
            """, [student_id])
            row = cursor.fetchone()
            if not row: return redirect('student_dashboard')
            student_data = {"name": row[0], "roll_no": row[1], "class_id": row[2]}

            # 2. Get All Subjects for this class (Ensures even unmarked subjects show up)
            cursor.execute("SELECT id, subject_name FROM main_subjects WHERE class_obj_id = %s", [student_data["class_id"]])
            subjects_list = [{"id": r[0], "name": r[1]} for r in cursor.fetchall()]

            # 3. Get Attendance Records for the specific date
            cursor.execute("""
                SELECT subject_id, hour, status, attendance_date 
                FROM main_attendance 
                WHERE student_id = %s AND attendance_date = %s
                ORDER BY hour ASC
            """, [student_id, selected_date])
            attendance_rows = cursor.fetchall()

            # 4. Structure Data: Pre-fill ALL subjects for the selected date
            # This ensures subjects with no data show "--" (Nill)
            attendance_by_date[selected_date] = {
                str(s['id']): {
                    "subject_name": s['name'],
                    "h1": "--", "h2": "--", "h3": "--", "h4": "--", "h5": "--"
                } for s in subjects_list
            }

            # 5. Fill in the actual marked attendance
            for row in attendance_rows:
                sub_id, hour, status, date = row[0], row[1], row[2], row[3]
                hour_key = f"h{hour}"
                if str(sub_id) in attendance_by_date[selected_date]:
                    attendance_by_date[selected_date][str(sub_id)][hour_key] = status.capitalize() if status else status
                    
            # 6. Calculate Overall Attendance Percentage
            cursor.execute("""
                SELECT 
                    SUM(CASE WHEN status = 'present' THEN 1 ELSE 0 END) * 100.0 / NULLIF(COUNT(*), 0)
                FROM main_attendance
                WHERE student_id = %s
            """, [student_id])
            overall_att_row = cursor.fetchone()
            overall_attendance = round(overall_att_row[0], 2) if overall_att_row and overall_att_row[0] else 0

        print(f"\n{'='*70}")
        print(f"DEBUGGING ATTENDANCE FOR: {student_data['name']} | DATE: {selected_date}")
        print(f"{'-'*70}")
        for sub_id, data in attendance_by_date[selected_date].items():
            print(f"Sub ID: {sub_id:<4} | {data['subject_name']:<20} | H1:{data['h1']:<8} H2:{data['h2']:<8} H3:{data['h3']:<8} H4:{data['h4']:<8} H5:{data['h5']:<8}")
        print(f"{'='*70}\n")

    except Exception as e:
        print(f"Error: {e}")
        return redirect('student_dashboard')

    context = {
        "student": student_data,
        "attendance_by_date": attendance_by_date,
        "selected_date": selected_date,
        "overall_attendance": overall_attendance,
    }
    return render(request, "students/student_view_attendance.html", context)


# 🔹 Function to preprocess text: correct spelling, remove extra spaces
def preprocess_text(text):
    text = text.strip().lower()  # Remove extra spaces & lowercase text
    text = str(TextBlob(text).correct())  # Correct spelling
    return text

from groq import Groq
from dotenv import load_dotenv
load_dotenv()

print("GROQ_API_KEY:", os.getenv("GROQ_API_KEY")[:10] + "..." if os.getenv("GROQ_API_KEY") else "Not found")

client = Groq(api_key=os.getenv("GROQ_API_KEY"))

SYSTEM_PROMPT = """You are the Eduke Success Pal — a friendly, cool, and supportive AI buddy for students. 🚀

Your vibe is helpful and conversational. Keep it simple and direct. Use professional but friendly English.
Address the student directly as 'you'. 

Rules for your persona:
- BE BRIEF. No long paragraphs unless specifically asked for a detailed analysis.
- If the student just says 'Hi' or 'Hello', just say hi back in a friendly way! Don't dump their data immediately.
- Only use academic data if the student asks a question related to their performance, or if you're giving a quick, relevant tip.
- Keep total response length to 3-4 sentences maximum.
"""



def student_eduke_bot(request):
    if 'student_id' not in request.session:
        return redirect('student_login')

    student_id = request.session['student_id']
    
    # 1. Fetch Student Details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT s.name, s.roll_no, c.class_name, c.id as class_id
            FROM main_students s
            LEFT JOIN main_classes c ON s.class_obj_id = c.id
            WHERE s.id = %s
        """, [student_id])
        student_row = cursor.fetchone()
        if not student_row:
            return redirect('student_login')
        
        student_name, roll_no, class_name, class_id = student_row

    # Build Comprehensive Context
    student_context = f"--- STUDENT PROFILE ---\n"
    student_context += f"Name: {student_name}\n"
    student_context += f"Roll Number: {roll_no}\n"
    student_context += f"Class: {class_name}\n\n"

    # 2. Fetch Comprehensive Performance Evaluation from StudentEvaluation
    student_context += "--- ACADEMIC PERFORMANCE & EVALUATION ---\n"
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT sub.subject_name, e.marks_percentage, e.attendance_percentage,
                   e.study_time_rating, e.sleep_time_rating, 
                   e.class_participation_rating, e.academic_activity_rating
            FROM main_studentevaluation e
            JOIN main_subjects sub ON e.subject_id = sub.id
            WHERE e.student_id = %s
        """, [student_id])
        evals = cursor.fetchall()
        for ev in evals:
            student_context += f"Subject: {ev[0]}\n"
            student_context += f"  - Internal Marks: {ev[1]}%\n"
            student_context += f"  - Attendance Rate: {ev[2]}%\n"
            student_context += f"  - Study Time Rating: {ev[3]}/10\n"
            student_context += f"  - Sleep Time Rating: {ev[4]}/10\n"
            student_context += f"  - Class Participation: {ev[5]}/10\n"
            student_context += f"  - Academic Activity: {ev[6]}/10\n"
            student_context += "\n"

    # 3. Quiz Performance Analysis
    student_context += "--- QUIZ PERFORMANCE ANALYSIS ---\n"
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT mq.name as quiz_name, sub.subject_name,
                   COUNT(q.id) as total_questions,
                   SUM(CASE WHEN q.correct_option = r.student_response THEN 1 ELSE 0 END) as correct_answers
            FROM main_quizresponse r
            JOIN main_quizquestions q ON r.question_id = q.id
            JOIN main_quizzes mq ON q.quiz_id = mq.id
            JOIN main_subjects sub ON mq.subject_id = sub.id
            WHERE r.student_id = %s
            GROUP BY mq.id, mq.name, sub.subject_name
        """, [student_id])
        quizzes = cursor.fetchall()
        for q in quizzes:
            total_q = q[2]
            correct_a = q[3] or 0
            percentage = (correct_a / total_q) * 100 if total_q > 0 else 0
            student_context += f"Quiz: {q[0]} ({q[1]})\n"
            student_context += f"  - Performance: {correct_a}/{total_q} ({round(percentage, 2)}%)\n"

    # 4. Overall Attendance Analysis
    student_context += "\n--- OVERALL ATTENDANCE ANALYSIS ---\n"
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                COUNT(*) as total_records,
                SUM(CASE WHEN status = 'present' THEN 1 ELSE 0 END) as present_count
            FROM main_attendance
            WHERE student_id = %s
        """, [student_id])
        att_row = cursor.fetchone()
        if att_row and att_row[0] and att_row[0] > 0:
            total_classes = att_row[0]
            present_classes = att_row[1] or 0
            att_pct = round((present_classes / total_classes) * 100, 1)
            student_context += f"Total Classes Conducted: {total_classes}\n"
            student_context += f"Classes You Attended: {present_classes}\n"
            student_context += f"Aggregate Attendance Rate: {att_pct}%\n"
        else:
            student_context += "No attendance records found yet.\n"

    # 5. Recommended Study Materials
    student_context += "\n--- RECOMMENDED STUDY MATERIALS ---\n"
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT sub.subject_name, sm.announcement, sm.file_url
            FROM main_studymaterials sm
            JOIN main_subjects sub ON sm.subject_id = sub.id
            WHERE sm.class_obj_id = %s
            ORDER BY sm.created_at DESC
            LIMIT 10
        """, [class_id])
        materials = cursor.fetchall()
        for m in materials:
            student_context += f"Subject: {m[0]}\n"
            student_context += f"  - Resource: {m[1]}\n"
            student_context += f"  - Reference Link: {m[2]}\n"

    if request.method == "GET" and "query" in request.GET:
        user_message = request.GET.get("query", "").strip()

        if not user_message:
            return JsonResponse({"response": "Please enter a query for the Eduke Academic Assistant."})
        
        # Specialized handling for initial professional greeting
        is_greeting = user_message == "initiate_academic_greeting"
        
        # Initialize or fetch history
        if is_greeting:
            request.session['student_bot_history'] = []
            user_message = "Provide a very brief professional 'Welcome back' greeting. Acknowledge my name and state you are ready to assist. Keep it under 2 or 3 sentences."
        
        history = request.session.get('student_bot_history', [])

        api_key = os.getenv("GROQ_API_KEY")
        if not api_key:
            return JsonResponse({"response": "An internal server error occurred: GROQ_API_KEY is not configured."})

        # Construct Professional Personalized System Prompt
        professional_personalized_prompt = (
            f"{SYSTEM_PROMPT}\n\n"
            f"--- YOUR ACADEMIC CONTEXT ---\n"
            f"{student_context}\n\n"
            f"INSTRUCTIONS:\n"
            f"1. RESPOND BRIEFLY. 3-4 sentences max.\n"
            f"2. If the user query is a greeting like 'Hi', 'Hello', or 'How are you', do NOT mention academic data. Just be friendly.\n"
            f"3. Only mention specific data (marks, attendance) if the user asks about it or if it's highly relevant to their specific question.\n"
            f"4. Use a supportive, friendly 'Success Pal' tone. Avoid formal business language."
        )

        try:
            # Build messages list with history
            messages_payload = [{"role": "system", "content": professional_personalized_prompt}]
            
            # Add last 10 messages from history to keep context manageable
            for msg in history[-10:]:
                messages_payload.append(msg)
                
            messages_payload.append({"role": "user", "content": user_message})

            chat_completion = client.chat.completions.create(
                messages=messages_payload,
                model="llama-3.3-70b-versatile",
                temperature=0.7,
                max_tokens=1000,
            )
            answer = chat_completion.choices[0].message.content.strip()

            # Save to history (excluding the greeting trigger itself)
            if not is_greeting:
                history.append({"role": "user", "content": user_message})
            history.append({"role": "assistant", "content": answer})
            
            # Keep history to last 20 messages total
            request.session['student_bot_history'] = history[-20:]
            request.session.modified = True

        except Exception as e:
            print("Groq Error:", str(e))
            answer = "I apologize, but I am currently experiencing technical difficulties processing your request. Please try again momentarily."

        return JsonResponse({"response": answer})

    return render(request, "students/student_eduke_bot.html", {"student": {"name": student_name, "roll_no": roll_no, "class_name": class_name}})

import json
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.db import connection
from ml.predict import predict_performance  # Import the prediction function

def student_prediction(request):
    if 'student_id' not in request.session:
        print("\033[91m[DEBUG] Redirecting to student login (no student_id in session).\033[0m")
        return JsonResponse({"error": "Unauthorized access"}, status=403) if request.headers.get("X-Requested-With") == "XMLHttpRequest" else redirect('student_login')

    student_id = request.session['student_id']
    stud = Students.objects.get(id=student_id)
    stud_name = stud.name
    selected_subject_id = request.GET.get('subject_id')  
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"  
    

    print(f"\033[96m[DEBUG] Student ID:\033[0m {student_id}, \033[96mSelected Subject ID:\033[0m {selected_subject_id}")

    try:
        with connection.cursor() as cursor:
            # Fetch student details
            print("\033[94m[DEBUG] Fetching student details...\033[0m")
            cursor.execute("""
                SELECT s.name, s.roll_no, s.email, c.class_name 
                FROM main_students s 
                LEFT JOIN main_classes c ON s.class_obj_id = c.id
                WHERE s.id = %s
            """, [student_id])
            student = cursor.fetchone()

            if not student:
                print("\033[91m[ERROR] Student not found! Redirecting to login.\033[0m")
                return JsonResponse({"error": "Student not found"}, status=404) if is_ajax else redirect("student_login")

            student_name, roll_no, email, class_name = student

            # Fetch subjects for dropdown
            print("\033[94m[DEBUG] Fetching subjects for dropdown...\033[0m")
            cursor.execute("""
                SELECT id, subject_name FROM main_subjects 
                WHERE class_obj_id = (SELECT class_obj_id FROM main_students WHERE id = %s)
            """, [student_id])
            subjects = cursor.fetchall()
            subjects_list = [{"id": sub[0], "name": sub[1]} for sub in subjects]

            # Default values
            study_time = sleep_time = class_participation = academic_activity = attendance_percentage = marks_percentage = 0
            predicted_marks = None  # Initialize predicted marks as None

            # Fetch evaluation details if subject is selected
            if selected_subject_id:
                print(f"\033[94m[DEBUG] Fetching evaluation data for subject ID: {selected_subject_id}\033[0m")
                cursor.execute("""
                    SELECT study_time_rating, sleep_time_rating, class_participation_rating, 
                           academic_activity_rating, attendance_percentage, marks_percentage
                    FROM main_studentevaluation
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, selected_subject_id])
                evaluation = cursor.fetchone()

                if evaluation:
                    # Replace None values with 0 using a list comprehension
                    evaluation = [val if val is not None else 0 for val in evaluation]
                    study_time, sleep_time, class_participation, academic_activity, attendance_percentage, marks_percentage = evaluation

                    # Call prediction model
                    print("\033[94m[DEBUG] Calling mark prediction model...\033[0m")
                    predicted_marks = predict_performance(
                        attendance_percentage, marks_percentage, class_participation, 
                        academic_activity, sleep_time, study_time
                    )

                    print(f"\033[92m[DEBUG] Predicted Marks: {predicted_marks}\033[0m")

            # Prepare student data dictionary
            student_data = {
                "student_name": stud_name,
                "roll_no": roll_no,
                "email": email,
                "class_name": class_name if class_name else "N/A",
                "study_time_rating": study_time,
                "sleep_time_rating": sleep_time,
                "class_participation_rating": class_participation,
                "academic_activity_rating": academic_activity,
                "attendance_percentage": attendance_percentage,
                "internal_marks": marks_percentage,
                "predicted_marks": predicted_marks,  # Add predicted marks
                "subjects": subjects_list
            }

            # If AJAX request, return JSON response
            if is_ajax:
                return JsonResponse({
                    "predicted_marks": predicted_marks if predicted_marks is not None else "N/A"
                })

    except Exception as e:
        print(f"\033[91m[ERROR] Exception occurred: {e}\033[0m")
        return JsonResponse({"error": str(e)}, status=500) if is_ajax else redirect("student_login")

    print("\033[94m[DEBUG] Rendering student_prediction.html with student data and subjects.\033[0m")
    return render(
        request, 
        "students/student_prediction.html", 
        {
            "student_data": student_data,
            "subjects": subjects_list,
            "selected_subject_id": selected_subject_id
        }
    )



######################################################################################################################

def parent_login(request):
    form = ParentLoginForm()

    if request.method == "POST":
        form = ParentLoginForm(request.POST)

        if form.is_valid():
            roll_no = form.cleaned_data.get("roll_no").upper()
            password = form.cleaned_data.get("password")

            # Validate credentials with raw SQL query
            with connection.cursor() as cursor:
                print("🔹 Executing SQL query to authenticate parent...")
                cursor.execute("""
                    SELECT p.id, p.name, s.id AS student_id, s.class_obj_id, p.password
                    FROM main_parents p
                    JOIN main_students s ON p.student_id = s.id
                    JOIN main_users u ON p.user_id = u.id
                    JOIN main_classes c ON s.class_obj_id = c.id
                    WHERE s.roll_no = %s AND u.role = 'parent'
                """, [roll_no])
                parent = cursor.fetchone()

            if parent:
                parent_id, parent_name, student_id, class_id, db_password = parent
                
                # Check password using decryption for encrypted passwords
                if verify_password(password, db_password):
                    # Save the parent ID in session
                    request.session['parent_id'] = parent_id
                    
                    # Create audit log
                    try:
                        from .models import Institution, Parents, Students
                        client_ip = get_client_ip(request)
                        # Get institution through student's class
                        class_obj = Classes.objects.get(id=class_id)
                        institution = Institution.objects.get(institution_id=class_obj.institution_id)
                        parent_obj = Parents.objects.get(id=parent_id)
                        AuditLog.objects.create(
                            user=parent_obj.user,
                            institution=institution,
                            action='login',
                            ip_address=client_ip,
                            user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
                            details=f'Parent login: {parent_name or "Parent of " + roll_no}',
                            related_object_id=parent_id,
                            related_object_type='Parents'
                        )
                    except Exception as e:
                        print(f"Parent login audit logging failed: {e}")
                    
                    return redirect('parent_dashboard')  # Redirect to parent dashboard
            else:
                messages.error(request, "Invalid Roll number or Password!", extra_tags="p_login_error")

    return render(request, 'parents/parent_login.html', {'form': form})



def parent_dashboard(request):
    print("🔹 Parent Dashboard View Called")

    # Debug session data
    print(f"🔹 Session Data: {dict(request.session.items())}")

    parent_id = request.session.get('parent_id')
    if not parent_id:
        print("❌ No parent_id in session. Redirecting to login.")
        return redirect('parent_login')

    with connection.cursor() as cursor:
        print(f"🔹 Fetching parent details for parent_id: {parent_id}")
        
        cursor.execute("""
            SELECT p.id, p.name AS parent_name, p.student_id, 
                   s.id AS student_id, s.name AS student_name, s.class_obj_id
            FROM main_parents p
            JOIN main_students s ON p.student_id = s.id  
            WHERE p.id = %s
        """, [parent_id])
        
        parent_data = cursor.fetchone()
        
        if not parent_data:
            print(f"❌ No parent found with id {parent_id}. Redirecting to login.")
            return redirect('parent_login')

        print(f"✅ Parent Data Fetched: {parent_data}")

        # Correctly unpacking all 6 values
        parent_id, parent_name, student_roll_no, student_id, student_name, student_class_id = parent_data

        print(f"🔹 Fetching class details for class_id: {student_class_id}")
        cursor.execute("""
            SELECT class_name, class_head FROM main_classes
            WHERE id = %s
        """, [student_class_id])
        class_data = cursor.fetchone()

        if class_data:
            student_class, class_teacher = class_data
            print(f"✅ Class Data: {class_data}")
        else:
            student_class, class_teacher = "Class not assigned", "No teacher assigned"
            print("❌ Class data not found.")

        # Fetch subjects count
        print(f"🔹 Fetching subject count for class_id: {student_class_id}")
        cursor.execute("""
            SELECT COUNT(*) FROM main_subjects
            WHERE class_obj_id = %s
        """, [student_class_id])
        subjects_count_data = cursor.fetchone()
        subjects_count = subjects_count_data[0] if subjects_count_data else 0
        print(f"✅ Subjects Count: {subjects_count}")

    # Prepare context for the template
    context = {
        'parent_id': parent_id,
        'parent_name': parent_name,
        'student_roll_no': student_roll_no,  # Added
        'student_id': student_id,  # Added
        'student_name': student_name,
        'student_class': student_class,
        'class_teacher': class_teacher,
        'subjects_count': subjects_count
    }

    print("✅ Rendering Parent Dashboard Page")
    return render(request, 'parents/parent_dashboard.html', context)



def parent_teacher_profile(request):
    parent_id = request.session.get('parent_id')  # Ensure the parent is logged in
    if not parent_id:
        return redirect('parent_login')

    # Fetch class teacher details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT c.class_head AS teacher_name, c.email AS teacher_email, 
                   c.class_name, i.institution_name
            FROM main_parents p
            JOIN main_students s ON p.student_id = s.id
            LEFT JOIN main_classes c ON s.class_obj_id = c.id
            LEFT JOIN main_institution i ON c.institution_id = i.institution_id
            WHERE p.id = %s
        """, [parent_id])
        teacher_data = cursor.fetchone()

    # Handle case where teacher data is not found
    if not teacher_data:
        messages.error(request, "Unable to retrieve teacher details.")
        return redirect('parent_dashboard')

    # Prepare context data for the template
    context = {
        'teacher_name': teacher_data[0] or "N/A",
        'teacher_email': teacher_data[1] or "N/A",
        'class_name': teacher_data[2] or "N/A",
        'institution_name': teacher_data[3] or "N/A",
        'parent_name': Parents.objects.get(id=parent_id).name,
        'student_name' : Parents.objects.get(id=parent_id).student.name,
    }
    return render(request, 'parents/parent_teacher_profile.html', context)



def parent_class(request):
    print("🔹 Parent Class View Called")

    parent_id = request.session.get('parent_id')
    print(f"🔹 Session Data: {request.session.items()}")  # Debug session

    if not parent_id:
        print("❌ No parent session found. Redirecting to login.")
        return redirect('parent_login')

    with connection.cursor() as cursor:
        # Fetch parent and student details
        cursor.execute("""
            SELECT p.id, p.name AS parent_name, p.student_id, 
                   s.name AS student_name, s.class_obj_id 
            FROM main_parents p
            JOIN main_students s ON p.student_id = s.id
            WHERE p.id = %s
        """, [parent_id])
        parent_data = cursor.fetchone()
        print(f"🔹 Parent Data Fetched: {parent_data}")

        if not parent_data:
            print("❌ Parent data not found. Redirecting to login.")
            return redirect('parent_login')

        # Extract details
        _, parent_name, student_roll_no, student_name, student_class_id = parent_data

        # Fetch class details
        cursor.execute("""
            SELECT class_name, class_head FROM main_classes
            WHERE id = %s
        """, [student_class_id])
        class_data = cursor.fetchone()
        student_class = class_data[0] if class_data else "Class not assigned"
        class_teacher = class_data[1] if class_data else "No teacher assigned"
        print(f"🔹 Class Data: {class_data}")

        # Fetch subjects
        cursor.execute("""
            SELECT subject_name, subject_head, email FROM main_subjects
            WHERE class_obj_id = %s
        """, [student_class_id])
        subjects = cursor.fetchall()
        print(f"🔹 Subjects Found: {len(subjects)}")

        # Fetch students
        cursor.execute("""
            SELECT roll_no, name, email FROM main_students
            WHERE class_obj_id = %s
        """, [student_class_id])
        students = cursor.fetchall()
        print(f"🔹 Students Found: {len(students)}")

    # Prepare context for the template
    context = {
        'parent_name': parent_name,
        'student_name': student_name,
        'student_class': student_class,
        'class_teacher': class_teacher,
        'subjects': subjects,
        'students': students,
        'student_count': len(students)
    }

    print("✅ Rendering parent_class.html")
    return render(request, 'parents/parent_class.html', context)



def parent_profile(request):
    parent_id = request.session.get('parent_id')  # Ensure the parent is logged in
    if not parent_id:
        return redirect('parent_login')

    if request.method == 'POST':
        # Handle profile update
        parent_name = request.POST.get('parent_name')
        password = request.POST.get('password')
        parent_email = request.POST.get('parent_email')

        # Update name, password, and email directly
        with connection.cursor() as cursor:
            if password:
                # Encrypt the password before updating
                encrypted_password = encrypt_password(password)
                cursor.execute("""
                    UPDATE main_parents 
                    SET name = %s, password = %s, email = %s 
                    WHERE id = %s
                """, [parent_name, encrypted_password, parent_email, parent_id])
            else:
                cursor.execute("""
                    UPDATE main_parents 
                    SET name = %s, email = %s 
                    WHERE id = %s
                """, [parent_name, parent_email, parent_id])
            messages.success(request, "Profile updated successfully!")

        return redirect('parent_profile')

    # Fetch parent's profile details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT p.name AS parent_name, p.password, p.email AS parent_email, 
                   s.name AS student_name, s.roll_no, 
                   c.class_name, 
                   c.class_head AS teacher_name, c.email AS teacher_email
            FROM main_parents p
            JOIN main_students s ON p.student_id = s.id
            LEFT JOIN main_classes c ON s.class_obj_id = c.id
            WHERE p.id = %s
        """, [parent_id])
        parent_data = cursor.fetchone()

    # Handle case where profile data is not found
    if not parent_data:
        messages.error(request, "Unable to retrieve profile details.")
        return redirect('parent_dashboard')

    # Prepare context data for the template
    context = {
        'parent_name': parent_data[0] or "",
        'password': decrypt_password(parent_data[1]) if parent_data[1] else "",  # Decrypt password
        'parent_email': parent_data[2] or "",  # Add the email to the context
        'student_name': parent_data[3] or "N/A",
        'student_roll_no': parent_data[4] or "N/A",
        'class_name': parent_data[5] or "N/A",
        'teacher_name': parent_data[6] or "N/A",
        'teacher_email': parent_data[7] or "N/A",
    }
    return render(request, 'parents/parent_profile.html', context)



def parent_chat(request):
    parent_id = request.session.get('parent_id') 
    if not parent_id:
        return redirect('parent_login')  

    with connection.cursor() as cursor:
        # 1. Get the parent's user ID and their child's class ID
        cursor.execute("""
            SELECT p.user_id, s.class_obj_id, s.name as student_name
            FROM main_parents p
            JOIN main_students s ON p.student_id = s.id
            WHERE p.id = %s
        """, [parent_id])
        parent_info = cursor.fetchone()

        if not parent_info:
            return render(request, 'parents/parent_chat.html', {'chat_users': [], 'suggested_users': []})

        user_id, student_class_id, student_name = parent_info

        # 2. Get recent chat users (already messaged) with unread counts
        cursor.execute("""
            SELECT DISTINCT u.id, 
                            COALESCE(c.class_head, s.subject_head) AS name, 
                            u.role,
                            COUNT(CASE WHEN ch.receiver_id = %s AND ch.is_read = FALSE THEN 1 END) as unread_count
            FROM main_chat ch
            JOIN main_users u ON (ch.sender_id = u.id OR ch.receiver_id = u.id)
            LEFT JOIN main_classes c ON u.id = c.user_id
            LEFT JOIN main_subjects s ON u.id = s.user_id
            WHERE (ch.sender_id = %s OR ch.receiver_id = %s)
            AND u.role IN ('class_head', 'subject_head')
            AND u.id != %s
            GROUP BY u.id, c.class_head, s.subject_head, u.role
        """, [user_id, user_id, user_id, user_id])
        chat_users = cursor.fetchall()

        # 3. Get Suggested Users (Class Head and Subject Heads of THIS class - including active ones) with unread counts
        # We filter by matching c.id = student_class_id or sub.class_obj_id = student_class_id
        cursor.execute("""
            SELECT DISTINCT u.id, 
                CASE 
                    WHEN u.role = 'class_head' THEN c.class_head
                    WHEN u.role = 'subject_head' THEN sub.subject_head
                END AS name,
                u.role,
                COUNT(CASE WHEN ch.receiver_id = %s AND ch.is_read = FALSE THEN 1 END) as unread_count
            FROM main_users u
            LEFT JOIN main_classes c ON u.id = c.user_id
            LEFT JOIN main_subjects sub ON u.id = sub.user_id
            LEFT JOIN main_chat ch ON (ch.sender_id = u.id AND ch.receiver_id = %s)
            WHERE (
                (u.role = 'class_head' AND c.id = %s) 
                OR 
                (u.role = 'subject_head' AND sub.class_obj_id = %s)
            )
            GROUP BY u.id, c.class_head, sub.subject_head, u.role
            ORDER BY u.role DESC, name ASC
        """, [user_id, user_id, student_class_id, student_class_id])

        suggested_users = cursor.fetchall()

    return render(request, 'parents/parent_chat.html', {
        'chat_users': chat_users,
        'suggested_users': suggested_users,
        'parent_name': Parents.objects.get(id=parent_id).name,
        'student_name' : Parents.objects.get(id=parent_id).student.name,
    })

def parent_chat_user(request, user_id):
    parent_id = request.session.get('parent_id')
    if not parent_id:
        return redirect('parent_login')

    # Fetch the parent's user_id
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT p.user_id
            FROM main_parents p
            WHERE p.id = %s
        """, [parent_id])
        parent_user_id = cursor.fetchone()

    if not parent_user_id:
        messages.error(request, 'Parent not found.')
        return redirect('parent_dashboard')

    # Debugging: Print the parent's user_id
    print(f"Parent's user_id for chat: {parent_user_id[0]}")

    # Mark all messages from this user as read
    Chat.objects.filter(
        sender_id=user_id,
        receiver_id=parent_user_id[0],
        is_read=False
    ).update(is_read=True)

    # Fetch messages between the logged-in parent and the selected user
    query = '''
        SELECT message, sender_id, receiver_id, created_at
        FROM main_chat
        WHERE (sender_id = %s AND receiver_id = %s) OR (sender_id = %s AND receiver_id = %s)
        ORDER BY created_at ASC
    '''
    with connection.cursor() as cursor:
        cursor.execute(query, [parent_user_id[0], user_id, user_id, parent_user_id[0]])
        messages_fetched = cursor.fetchall()

    # Debugging: Print the fetched messages
    print(f"Messages fetched: {messages_fetched}")

    # Get the selected user's role
    user = Users.objects.get(id=user_id)
    selected_user_role = user.role

    # Fetch the selected user's name based on their role
    if selected_user_role == 'class_head':
        selected_user_name = Classes.objects.filter(user_id=user_id).values_list('class_head', flat=True).first()
    elif selected_user_role == 'subject_head':
        selected_user_name = Subjects.objects.filter(user_id=user_id).values_list('subject_head', flat=True).first()
    else:
        selected_user_name = 'Unknown'

    # Fetch the teacher's user_id for message sender identification
    if selected_user_role == 'class_head':
        teacher_user_id = Classes.objects.filter(class_head=user_id).values_list('user_id', flat=True).first()
    else:
        teacher_user_id = Subjects.objects.filter(subject_head=user_id).values_list('user_id', flat=True).first()

    # Handle message submission via AJAX
    if request.method == 'POST':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.utils import timezone
            message_text = request.POST.get('message', '').strip()
            if message_text:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        INSERT INTO main_chat (sender_id, receiver_id, message, created_at, is_read)
                        VALUES (%s, %s, %s, %s, FALSE)
                    """, [parent_user_id[0], user_id, message_text, timezone.now()])
                return JsonResponse({'success': True, 'message': 'Message sent'})
            return JsonResponse({'success': False, 'error': 'Empty message'})
        else:
            # Fallback to regular form submission
            message_text = request.POST.get('message')
            if message_text:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        INSERT INTO main_chat (sender_id, receiver_id, message, created_at, is_read)
                        VALUES (%s, %s, %s, CURRENT_TIMESTAMP, FALSE)
                    """, [parent_user_id[0], user_id, message_text])
                return redirect('parent_chat_user', user_id=user_id)

    return render(request, 'parents/parent_chat_user.html', {
        'messages': messages_fetched, 
        'user_id': user_id,
        'logged_in_user_id': parent_user_id[0],  # Pass this for sender alignment
        'selected_user_name': selected_user_name,
        'selected_user_role': selected_user_role,
        'teacher_user_id': teacher_user_id,
        'parent_name':Parents.objects.get(id=parent_id).name
    })


def parent_fetch_messages(request, user_id):
    """API endpoint for parents to fetch new messages via AJAX"""
    if request.method == 'GET':
        parent_id = request.session.get('parent_id')
        if not parent_id:
            return JsonResponse({'error': 'Unauthorized'}, status=401)
        
        # Get logged-in parent user ID
        with connection.cursor() as cursor:
            cursor.execute("SELECT user_id FROM main_parents WHERE id = %s", [parent_id])
            parent_user = cursor.fetchone()
            if not parent_user:
                return JsonResponse({'error': 'Unauthorized'}, status=401)
            logged_in_user_id = parent_user[0]
        
        # Fetch messages
        with connection.cursor() as cursor:
            cursor.execute(""" 
                SELECT message, sender_id, receiver_id, created_at 
                FROM main_chat 
                WHERE (sender_id = %s AND receiver_id = %s) 
                OR (sender_id = %s AND receiver_id = %s) 
                ORDER BY created_at;
            """, [user_id, logged_in_user_id, logged_in_user_id, user_id])  
            messages = cursor.fetchall() or []
        
        # Mark messages as read
        Chat.objects.filter(
            sender_id=user_id,
            receiver_id=logged_in_user_id,
            is_read=False
        ).update(is_read=True)
        
        # Format messages for JSON response
        formatted_messages = []
        for msg in messages:
            formatted_messages.append({
                'message': msg[0],
                'sender_id': msg[1],
                'receiver_id': msg[2],
                'created_at': msg[3].isoformat() if msg[3] else ''  # Send ISO format for JS to convert to local time
            })
        
        return JsonResponse({'messages': formatted_messages})
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

from django.db import connection
from django.shortcuts import render, redirect
from django.contrib import messages

def parent_evaluation(request):
    parent_id = request.session.get('parent_id')
    if not parent_id:
        return redirect('parent_login')

    # 1. Fetch Student Details
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT s.id, s.name, s.roll_no, s.class_obj_id
            FROM main_students s
            JOIN main_parents p ON s.id = p.student_id
            WHERE p.id = %s
        """, [parent_id])
        student = cursor.fetchone()

    if not student:
        return render(request, 'parents/parent_evaluation.html', {'student': None, 'subjects': []})

    student_id, student_name, roll_no, class_id = student

    # 2. Fetch Subjects
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, subject_name FROM main_subjects WHERE class_obj_id = %s", [class_id])
        subjects_raw = cursor.fetchall()

    # 3. Fetch Existing Evaluation Data
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT subject_id, study_time_rating, sleep_time_rating 
            FROM main_studentevaluation 
            WHERE student_id = %s
        """, [student_id])
        # Map by subject_id for easy lookup
        eval_map = {row[0]: {'study': row[1], 'sleep': row[2]} for row in cursor.fetchall()}

    # 4. Prepare structured data for the sliders
    # This ensures the 'value' attribute in your HTML range input is populated
    structured_subjects = []
    for sub in subjects_raw:
        sub_id, sub_name = sub
        structured_subjects.append({
            'id': sub_id,
            'name': sub_name,
            # Default to 30 (Standard) if no record exists
            'study_time_rating': eval_map.get(sub_id, {}).get('study', 30),
            'sleep_time_rating': eval_map.get(sub_id, {}).get('sleep', 30),
        })

    # 5. Handle Form Submission
    if request.method == "POST":
        for subject in structured_subjects:
            s_id = subject['id']
            # Get values from sliders (Matches 'name' attribute in HTML)
            study_val = request.POST.get(f"study_time_rating_{s_id}")
            sleep_val = request.POST.get(f"sleep_time_rating_{s_id}")

            if study_val and sleep_val:
                with connection.cursor() as cursor:
                    # Check if record exists to decide UPDATE vs INSERT
                    cursor.execute("SELECT id FROM main_studentevaluation WHERE student_id = %s AND subject_id = %s", [student_id, s_id])
                    exists = cursor.fetchone()

                    if exists:
                        cursor.execute("""
                            UPDATE main_studentevaluation 
                            SET study_time_rating = %s, sleep_time_rating = %s 
                            WHERE student_id = %s AND subject_id = %s
                        """, [study_val, sleep_val, student_id, s_id])
                    else:
                        cursor.execute("""
                            INSERT INTO main_studentevaluation (student_id, subject_id, study_time_rating, sleep_time_rating)
                            VALUES (%s, %s, %s, %s)
                        """, [student_id, s_id, study_val, sleep_val])

        messages.success(request, "Behavioral metrics synchronized with AI Core.")
        return redirect('parent_evaluation')

    return render(request, 'parents/parent_evaluation.html', {
        'student_name': student_name,
        'subjects': structured_subjects,
        'parent_name': Parents.objects.get(id=parent_id).name
    })



from decimal import Decimal, ROUND_HALF_UP

def parent_student_performance(request):
    print("🔹 Starting parent_student_performance view...")

    parent_id = request.session.get('parent_id')  # Get parent ID from session
    if not parent_id:
        print("❌ Parent ID not found in session. Redirecting to login.")
        return redirect('parent_login')

    student_data = {}
    evaluations = {}
    quiz_percentage = 0  # Default to 0 instead of "--" for easier float conversion
    subjects = []
    selected_subject_name = None

    subject_id = request.GET.get('subject_id')  # Get selected subject from query params

    print(f"ℹ️ Parent ID from session: {parent_id}")

    try:
        # Fetch student linked to this parent
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT s.id, s.name, s.roll_no, c.class_name, s.email, c.id AS class_id
                FROM main_parents p
                JOIN main_students s ON p.student_id = s.id
                JOIN main_classes c ON s.class_obj_id = c.id
                WHERE p.id = %s
            """, [parent_id])
            row = cursor.fetchone()

            if row:
                student_id = row[0]
                student_data = {
                    "name": row[1],
                    "roll_no": row[2],
                    "class_name": row[3],
                    "email": row[4],
                    "class_id": row[5],
                }
                print(f"✅ Student Data Retrieved: {student_data}")
            else:
                print("⚠️ No student linked to this parent.")
                return render(request, "parents/parent_student_performance.html", {"error": "No student data available."})

        # Fetch subjects for the student's class
        if student_data.get("class_id"):
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, subject_name FROM main_subjects WHERE class_obj_id = %s
                """, [student_data["class_id"]])
                subjects = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]

            # Get the name of the selected subject
            selected_subject_name = next((s["name"] for s in subjects if str(s["id"]) == subject_id), None)
            
            if selected_subject_name:
                print(f"🔹 Selected Subject: {selected_subject_name} (ID: {subject_id})")
            else:
                print("⚠️ Selected subject not found in student's class subjects.")

        # Fetch Student Evaluations for the selected subject
        if subject_id:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT marks_percentage, attendance_percentage,
                           study_time_rating, sleep_time_rating, class_participation_rating, academic_activity_rating
                    FROM main_studentevaluation
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    evaluations = {
                        "marks_percentage": round(float(row[0] or 0), 2),
                        "attendance_percentage": round(float(row[1] or 0), 2),
                        "study_time_rating": round(float(row[2] or 0), 2),
                        "sleep_time_rating": round(float(row[3] or 0), 2),
                        "class_participation_rating": round(float(row[4] or 0), 2),
                        "academic_activity_rating": round(float(row[5] or 0), 2),
                    }
                    print(f"📊 Evaluations Retrieved: {evaluations}")
                else:
                    print(f"⚠️ No evaluation data found for {selected_subject_name}.")

        # Fetch Quiz Performance for the selected subject
        if subject_id:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT COUNT(*) AS total_attempted,
                           SUM(CASE WHEN q.correct_option = r.student_response THEN 1 ELSE 0 END) AS correct_answers
                    FROM main_quizresponse r
                    JOIN main_quizquestions q ON r.question_id = q.id
                    JOIN main_quizzes mq ON q.quiz_id = mq.id
                    WHERE r.student_id = %s AND mq.subject_id = %s
                """, [student_id, subject_id])
                row = cursor.fetchone()

                if row:
                    total_attempted, correct_answers = row[0], row[1] or 0
                    quiz_percentage = round((correct_answers / total_attempted) * 100, 2) if total_attempted > 0 else 0
                    print(f"📝 Quiz Performance: Attempted = {total_attempted}, Correct = {correct_answers}, Percentage = {quiz_percentage}%")
                else:
                    print(f"⚠️ No quiz data found for {selected_subject_name}.")

        # Prepare data for Graph
        graph_data = [
            float(evaluations.get("marks_percentage", 0)),
            float(evaluations.get("attendance_percentage", 0)),
            float(evaluations.get("study_time_rating", 0)),
            float(evaluations.get("sleep_time_rating", 0)),
            float(evaluations.get("class_participation_rating", 0)),
            float(evaluations.get("academic_activity_rating", 0)),
            float(quiz_percentage)
        ]

    except Exception as e:
        print(f"❌ An error occurred: {e}")
        return redirect('error_page')

    context = {
        "student": student_data,
        "evaluations": evaluations,
        "quiz_percentage": quiz_percentage,
        "subjects": subjects,
        "selected_subject_id": int(subject_id) if subject_id else None,
        "graph_data": graph_data,  # Pass graph data for chart rendering
        "parent_name":Parents.objects.get(id=parent_id).name,
        'student_name' : Parents.objects.get(id=parent_id).student.name,
    }

    print("✅ Rendering 'parent_student_performance.html' template.")
    return render(request, "parents/parent_student_performance.html", context)





"""
Handles chatbot page rendering and processing chatbot queries for parents.

The chatbot can respond to basic queries like "Hi, how are you%s" or "What is your name%s" and also process more complex queries like "What are my child's weak subjects%s" or "How can I improve my child's attendance%s"
"""

from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.db import connection


print("GROQ_API_KEY (parent):", os.getenv("GROQ_API_KEY")[:10] + "..." if os.getenv("GROQ_API_KEY") else "Not found")

client = Groq(api_key=os.getenv("GROQ_API_KEY"))

PARENT_SYSTEM_PROMPT = """You are Eduke Parent Bot — a professional, supportive, and knowledgeable AI assistant for parents of students at Eduke. 🏫👔

You speak in clear, professional, and respectful English. Your goal is to provide accurate information and helpful guidance based on the data provided for each student.
Stay helpful and encouraging, but maintain a professional distance.

Rules for your persona:
- NEVER use the word "beta", "baccha", or any other overly familiar or casual terms when addressing parents or referring to students.
- Avoid excessive emojis; use only a few to remain friendly but professional.
- Use the student's name respectrufully.
- Do not be overly "warm" or "elderly"; act as a professional school counsellor.

Your main role:
- Help parents understand their child's actual academic performance using the real data provided.
- Provide objective motivation and practical study advice.
- Explain performance trends clearly and professionally.

---
STUDENT DATA (use this to answer academic queries):
{student_data_context}
---

Always base your answers on the provided data. If a parent asks for something not in the data, state clearly that the information is currently unavailable. 🌟"""

def parent_eduke_bot(request):
    """LLM-powered Eduke Bot for Parents — fetches real student data and passes it to the AI."""

    if 'parent_id' not in request.session:
        return redirect('parent_login')

    parent_id = request.session.get("parent_id")

    # ── 1. Fetch parent & linked student basic info ────────────────────────────
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT main_parents.name,
                   main_students.id,
                   main_students.name,
                   main_students.roll_no,
                   main_classes.class_name
            FROM main_parents
            LEFT JOIN main_students ON main_parents.student_id = main_students.id
            LEFT JOIN main_classes ON main_students.class_obj_id = main_classes.id
            WHERE main_parents.id = %s
        """, [parent_id])
        parent_row = cursor.fetchone()

    if not parent_row:
        parent_details = {"parent_name": "Parent", "student_name": "your child", "student_id": None, "roll_no": "", "class": ""}
    else:
        parent_details = {
            "parent_name": parent_row[0] or " ",
            "student_id": parent_row[1],
            "student_name": parent_row[2] or "your child",
            "roll_no": parent_row[3] or "",
            "class": parent_row[4] or "—",
        }

    student_id = parent_details["student_id"]
    student_name = parent_details["student_name"]

    # ── 2. Build rich student data context (once, for both GET and POST) ───────
    student_data_context = f"Student: {student_name} | Roll No: {parent_details['roll_no']} | Class: {parent_details['class']}\n"

    if student_id:
        with connection.cursor() as cursor:

            # Internal Marks (from StudentEvaluation)
            cursor.execute("""
                SELECT s.subject_name, e.marks_percentage
                FROM main_studentevaluation e
                JOIN main_subjects s ON e.subject_id = s.id
                WHERE e.student_id = %s
                ORDER BY s.subject_name
            """, [student_id])
            eval_marks_rows = cursor.fetchall()
            if eval_marks_rows:
                student_data_context += "\nInternal Marks (Current Evaluation):\n"
                for row in eval_marks_rows:
                    student_data_context += f"  - {row[0]}: {round(float(row[1] or 0), 1)}%\n"
            else:
                student_data_context += "\nInternal Marks: No evaluation data available yet.\n"

            # Attendance calculation up to today
            cursor.execute("""
                SELECT 
                    COUNT(*) as total_records,
                    SUM(CASE WHEN status = 'present' THEN 1 ELSE 0 END) as present_count
                FROM main_attendance
                WHERE student_id = %s AND attendance_date <= DATE('now')
            """, [student_id])
            att_row = cursor.fetchone()
            if att_row and att_row[0] and att_row[0] > 0:
                total_classes = att_row[0]
                present_classes = att_row[1] or 0
                att_pct = round((present_classes / total_classes) * 100, 1)
                student_data_context += f"\nAttendance Data (up to today):\n"
                student_data_context += f"  - Classes Conducted: {total_classes}\n"
                student_data_context += f"  - Classes Attended: {present_classes}\n"
                student_data_context += f"  - Performance Percentage: {att_pct}%\n"
            else:
                student_data_context += "\nAttendance Data: No records found up to today.\n"

            # Quiz scores (per quiz, latest first)
            cursor.execute("""
                SELECT mq.name,
                       COUNT(r.id) AS attempted,
                       SUM(CASE WHEN q.correct_option = r.student_response THEN 1 ELSE 0 END) AS correct,
                       COUNT(q.id) AS total_questions
                FROM main_quizresponse r
                JOIN main_quizquestions q ON r.question_id = q.id
                JOIN main_quizzes mq ON q.quiz_id = mq.id
                WHERE r.student_id = %s
                GROUP BY mq.id, mq.name
                ORDER BY mq.id DESC
                LIMIT 10
            """, [student_id])
            quiz_rows = cursor.fetchall()
            if quiz_rows:
                student_data_context += "\nQuiz Scores (latest first):\n"
                for row in quiz_rows:
                    title, attempted, correct, total_q = row
                    pct = round((correct / attempted) * 100, 1) if attempted else 0
                    student_data_context += f"  - {title}: {correct}/{attempted} correct ({pct}%)\n"
            else:
                student_data_context += "\nQuiz Scores: No quizzes attempted yet.\n"

            # Subject-wise evaluation ratings
            cursor.execute("""
                SELECT s.subject_name,
                       e.study_time_rating,
                       e.sleep_time_rating,
                       e.class_participation_rating,
                       e.academic_activity_rating,
                       e.attendance_percentage,
                       e.marks_percentage
                FROM main_studentevaluation e
                JOIN main_subjects s ON e.subject_id = s.id
                WHERE e.student_id = %s
                ORDER BY s.subject_name
            """, [student_id])
            eval_rows = cursor.fetchall()
            if eval_rows:
                student_data_context += "\nTeacher Evaluation Ratings (per subject, scale 0-100):\n"
                for row in eval_rows:
                    subj, study, sleep, participation, activity, att_eval, marks_eval = row
                    student_data_context += (
                        f"  - {subj}: Study Time={round(float(study or 0),1)}, "
                        f"Sleep={round(float(sleep or 0),1)}, "
                        f"Class Participation={round(float(participation or 0),1)}, "
                        f"Academic Activity={round(float(activity or 0),1)}, "
                        f"Attendance%={round(float(att_eval or 0),1)}, "
                        f"Marks%={round(float(marks_eval or 0),1)}\n"
                    )

    # ── 3. Handle chat query ───────────────────────────────────────────────────
    if request.method == "GET" and "query" in request.GET:
        user_message = request.GET.get("query", "").strip()

        if not user_message:
            return JsonResponse({
                "response": f"Hello! 😊 How can I help you with {student_name}'s studies today? Feel free to ask about marks, attendance, quizzes, or anything else!"
            })

        # Fetch conversation history from session
        history = request.session.get('parent_bot_history', [])

        api_key = os.getenv("GROQ_API_KEY")
        if not api_key or len(api_key) < 15:
            return JsonResponse({
                "response": "Arre! Server side mein thodi dikkat aa gayi hai 😅 API key check karo zara... Please try again after some time."
            })

        try:
            # Inject real student data into the system prompt
            filled_system = PARENT_SYSTEM_PROMPT.replace("{student_data_context}", student_data_context)

            # Build messages list with history
            messages_payload = [{"role": "system", "content": filled_system}]
            
            # Add last 10 messages from history to keep context manageable
            for msg in history[-10:]:
                messages_payload.append(msg)
                
            messages_payload.append({"role": "user", "content": user_message})

            chat_completion = client.chat.completions.create(
                messages=messages_payload,
                model="llama-3.3-70b-versatile",
                temperature=0.7,
                max_tokens=600,
                top_p=0.9,
            )

            answer = chat_completion.choices[0].message.content.strip()
            
            # Save to history
            history.append({"role": "user", "content": user_message})
            history.append({"role": "assistant", "content": answer})
            
            # Keep history to last 20 messages total
            request.session['parent_bot_history'] = history[-20:]
            request.session.modified = True

            print(f"[Parent Bot] Response length: {len(answer)}")

        except Exception as e:
            error_str = str(e)
            print("Groq Error (parent):", error_str)
            if "decommissioned" in error_str.lower():
                answer = "Thodi der mein try karna please... model update chal raha hai 😅"
            elif "401" in error_str or "Invalid" in error_str:
                answer = "Authentication issue aa gaya... server team ko batate hain 🔑😓"
            elif "429" in error_str:
                answer = "Thoda zyada baat ho gayi server se... thodi der baad try karna yaar 😅"
            else:
                answer = f"Oops! Kuch gadbad ho gayi 😓 ({error_str[:60]}...) Thodi der baad try karenge okay?"

        return JsonResponse({"response": answer})

    # ── 4. Render page (initial load) ──────────────────────────────────────────
    return render(request, "parents/parent_eduke_bot.html", {
        "parent": parent_details,
        "parent_name": Parents.objects.get(id=parent_id).name,
        'student_name' : Parents.objects.get(id=parent_id).student.name,
    })

import json
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.db import connection
from ml.predict import predict_performance  # Import the prediction function

def parent_prediction(request):
    print("🔹 Starting parent_prediction view...")

    if 'parent_id' not in request.session:
        print("❌ Parent ID not found in session. Redirecting to login.")
        return JsonResponse({"error": "Unauthorized access"}, status=403) if request.headers.get("X-Requested-With") == "XMLHttpRequest" else redirect('parent_login')

    parent_id = request.session['parent_id']
    selected_subject_id = request.GET.get('subject_id')
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    print(f"📌 Parent ID: {parent_id}, 📌 Selected Subject ID: {selected_subject_id}")

    try:
        with connection.cursor() as cursor:
            # Fetch student's details associated with the parent
            print("🔎 Fetching student details linked to the parent...")
            cursor.execute("""
                SELECT s.id, s.name, s.roll_no, s.email, c.class_name 
                FROM main_students s
                INNER JOIN main_parents p ON p.student_id = s.id
                LEFT JOIN main_classes c ON s.class_obj_id = c.id
                WHERE p.id = %s
            """, [parent_id])
            student = cursor.fetchone()

            if not student:
                print("❌ No student found for this parent. Redirecting to login.")
                return JsonResponse({"error": "Student not found"}, status=404) if is_ajax else redirect("parent_login")

            student_id, student_name, roll_no, email, class_name = student

            print(f"""
                🎓 Student Details:
                - Student ID: {student_id}
                - Name: {student_name}
                - Roll No: {roll_no}
                - Email: {email}
                - Class Name: {class_name if class_name else "N/A"}
            """)

            # Fetch subjects for dropdown
            print("📘 Fetching subjects for dropdown...")
            cursor.execute("""
                SELECT id, subject_name FROM main_subjects 
                WHERE class_obj_id = (SELECT class_obj_id FROM main_students WHERE id = %s)
            """, [student_id])
            subjects = cursor.fetchall()
            subjects_list = [{"id": sub[0], "name": sub[1]} for sub in subjects]

            print(f"📚 Subjects Linked to Student: {subjects_list}")

            # Default values
            study_time = sleep_time = class_participation = academic_activity = attendance_percentage = marks_percentage = 0
            predicted_marks = None  # Initialize predicted marks as None

            # Fetch evaluation details if subject is selected
            if selected_subject_id:
                print(f"📊 Fetching evaluation data for subject ID: {selected_subject_id}")
                cursor.execute("""
                    SELECT study_time_rating, sleep_time_rating, class_participation_rating, 
                           academic_activity_rating, attendance_percentage, marks_percentage
                    FROM main_studentevaluation
                    WHERE student_id = %s AND subject_id = %s
                """, [student_id, selected_subject_id])
                evaluation = cursor.fetchone()

                if evaluation:
                    evaluation = [val if val is not None else 0 for val in evaluation]
                    study_time, sleep_time, class_participation, academic_activity, attendance_percentage, marks_percentage = evaluation

                    print(f"""
                        📊 Evaluation Data:
                        - Study Time Rating: {study_time}
                        - Sleep Time Rating: {sleep_time}
                        - Class Participation Rating: {class_participation}
                        - Academic Activity Rating: {academic_activity}
                        - Attendance Percentage: {attendance_percentage}
                        - Internal Marks: {marks_percentage}
                    """)

                    # Call prediction model
                    print("🧠 Calling mark prediction model...")
                    predicted_marks = predict_performance(
                        attendance_percentage, marks_percentage, class_participation, 
                        academic_activity, sleep_time, study_time
                    )
                    print(f"✅ Predicted Marks: {predicted_marks}")

            # Prepare student data dictionary
            student_data = {
                
                "study_time_rating": study_time,
                "sleep_time_rating": sleep_time,
                "class_participation_rating": class_participation,
                "academic_activity_rating": academic_activity,
                "attendance_percentage": attendance_percentage,
                "internal_marks": marks_percentage,
            }

            # Print final student data for debugging
            print(f"📌 Final Student Data: {json.dumps(student_data, indent=4)}")

            # If AJAX request, return JSON response
            if is_ajax:
                return JsonResponse({
                    "predicted_marks": predicted_marks if predicted_marks is not None else "N/A"
                })

    except Exception as e:
        print(f"❌ Exception occurred: {e}")
        return JsonResponse({"error": str(e)}, status=500) if is_ajax else redirect("parent_login")

    print("🔹 Rendering parent_prediction.html with student data and subjects.")
    return render(
        request, 
        "parents/parent_prediction.html", 
        {
            "student_data": student_data,
            "subjects": subjects_list,
            "selected_subject_id": selected_subject_id,
            "parent_name":Parents.objects.get(id=parent_id).name,
            'student_name' : Parents.objects.get(id=parent_id).student.name,

        }
    )



######################################################################################################################




# 🚀 Load Trained Model & Scaler
MODEL_PATH = os.path.join(settings.BASE_DIR, 'ml', 'final_model.pkl')
SCALER_PATH = os.path.join(settings.BASE_DIR, 'ml', 'scaler.pkl')

# ✅ Check if files exist before loading
if os.path.exists(MODEL_PATH) and os.path.exists(SCALER_PATH):
    model = joblib.load(MODEL_PATH)
    scaler = joblib.load(SCALER_PATH)
else:
    model = None
    scaler = None
    print(f"Error: Model or Scaler file missing. Check {MODEL_PATH} and {SCALER_PATH}.")

@csrf_exempt
def predict_marks(request):
    if request.method == "POST":
        if model is None or scaler is None:
            return JsonResponse({"error": "Model or scaler not found. Retrain and save them."}, status=500)

        try:
            # 📌 Parse JSON Request
            data = json.loads(request.body)
            sleep_time = float(data["sleep_time_rating"])
            study_time = float(data["study_time_rating"])
            class_participation = float(data["class_participation_rating"])
            academic_activity = float(data["academic_activity_rating"])
            attendance = float(data["attendance_percentage"])
            internal_marks = float(data["internal_marks"])

            # 📌 Scale Input Data
            input_data = np.array([[sleep_time, study_time, class_participation,
                                    academic_activity, attendance, internal_marks]])
            input_scaled = scaler.transform(input_data)

            # 📌 Predict Final Marks
            predicted_marks = model.predict(input_scaled)[0]

            return JsonResponse({"predicted_final_marks": round(predicted_marks, 2)})
        
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"message": "Send a POST request with student data."})


def export_class_report(request, class_id):
    # 1. Data Fetching
    try:
        class_obj = Classes.objects.get(id=class_id)
        # Sort by roll number to keep the registry organized
        students = Students.objects.filter(class_obj=class_obj).order_by('roll_no')
    except Classes.DoesNotExist:
        return HttpResponse("Class not found", status=404)

    # 2. Response Setup
    response = HttpResponse(content_type='application/pdf')
    filename = f"{class_obj.class_name}_Performance_Report.pdf".replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # 3. PDF Document Setup (A4)
    # A4 width is 8.27". We use ~7.2" for content after 40pt (~0.55") margins.
    doc = SimpleDocTemplate(
        response, 
        pagesize=A4, 
        rightMargin=40, 
        leftMargin=40, 
        topMargin=40, 
        bottomMargin=40
    )
    elements = []
    styles = getSampleStyleSheet()

    # --- Custom Styles ---
    title_style = ParagraphStyle(
        'TitleStyle', 
        parent=styles['Heading1'], 
        fontSize=20, 
        textColor=colors.HexColor("#4c1d95"), 
        spaceAfter=12
    )
    
    meta_style = ParagraphStyle(
        'MetaStyle',
        fontSize=10,
        textColor=colors.HexColor("#64748b"),
        leading=14
    )

    # Helper for colored percentage text
    def get_score_html(val):
        try:
            # Convert to float first to handle string decimals, then int
            val_int = int(float(val)) 
        except (ValueError, TypeError):
            val_int = 0
        color = "#10b981" if val_int >= 75 else ("#f59e0b" if val_int >= 40 else "#f43f5e")
        return f'<font color="{color}"><b>{val_int}%</b></font>'

    # --- Header Section ---
    elements.append(Paragraph("EDUKÉ PERFORMANCE PROTOCOL", styles['Normal'])) # [cite: 1]
    elements.append(Paragraph(f"Class Registry: {class_obj.class_name}", title_style)) # [cite: 2]
    
    # Metadata Table - Widened first column to ensure Class Head name is never cut off
    meta_data = [
        [Paragraph(f"<b>Class Head:</b> {class_obj.class_head}", meta_style), 
         Paragraph(f"<b>Total Students:</b> {students.count()}", meta_style)], # [cite: 3, 5]
        [Paragraph(f"<b>Date Generated:</b> 28 Feb 2026", meta_style), 
         Paragraph(f"<b>System Status:</b> ACTIVE", meta_style)] # [cite: 4, 6]
    ]
    meta_table = Table(meta_data, colWidths=[4.2*inch, 2.9*inch])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 20))

    # --- Main Data Table ---
    # Adjusted widths: Roll(1.1"), Identity(1.9"), Email(2.0"), Att(1.1"), Marks(1.1")
    data = [['ROLL NO', 'STUDENT IDENTITY', 'EMAIL', 'ATTENDANCE', 'AVG MARKS']] # 
    
    for s in students:
        # Accessing values safely. Ensure these field names match your models exactly.
        att_val = getattr(s, 'attendance', 0) or 0
        mark_val = getattr(s, 'marks', 0) or 0
        
        data.append([
            str(s.roll_no) if s.roll_no else "--",
            Paragraph(s.name, styles['Normal']), # Paragraph allows long names to wrap
            s.email,
            Paragraph(get_score_html(att_val), styles['Normal']),
            Paragraph(get_score_html(mark_val), styles['Normal'])
        ])

    # Table Geometry - Increased Roll No column to prevent text overlapping
    table = Table(data, colWidths=[1.1*inch, 1.9*inch, 2.0*inch, 1.1*inch, 1.1*inch], repeatRows=1)
    
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#7c3aed")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        
        # Body
        ('ALIGN', (0, 1), (0, -1), 'LEFT'), 
        ('ALIGN', (3, 1), (-1, -1), 'CENTER'), 
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))

    elements.append(table)
    
    # --- Footer ---
    elements.append(Spacer(1, 40))
    footer_text = "This is a computer-generated document. Identity verification required for official use." # [cite: 8]
    elements.append(Paragraph(footer_text, ParagraphStyle('Footer', fontSize=7, textColor=colors.grey, alignment=1)))

    doc.build(elements)
    return response

from django.db import connection # Ensure this is imported

def edit_quiz(request, quiz_id):
    # --- Fetch Subject Head Name and Allotted Class ---
    subject_id = request.session.get('subject_id')
    subject_head_name = "Unknown"
    allotted_classes = [] # Initialize as empty list
    
    if subject_id:
        # Fetch Subject Head Name
        with connection.cursor() as cursor:
            cursor.execute("SELECT subject_head, class_obj_id FROM main_subjects WHERE id = %s", [subject_id])
            row = cursor.fetchone()
            if row:
                subject_head_name = row[0]
                class_id = row[1]
                # Fetch ONLY the specific class allotted to this subject
                allotted_classes = Classes.objects.filter(id=class_id)
    # ---------------------------------------

    # 1. Fetch the actual Quiz object
    quiz = get_object_or_404(Quizzes, pk=quiz_id)
    
    # 2. Fetch questions belonging to this quiz
    questions = QuizQuestions.objects.filter(quiz=quiz)

    if request.method == "POST":
        # Update Main Quiz Registry
        quiz.name = request.POST.get('title')
        # Update the class from the form if allowed
        class_obj_id = request.POST.get('class_id')
        if class_obj_id:
            quiz.class_obj_id = class_obj_id
        quiz.save()

        # Synchronize Questions Matrix
        q_ids = request.POST.getlist('question_ids[]')
        q_texts = request.POST.getlist('questions[]')
        opts_a = request.POST.getlist('option_a[]')
        opts_b = request.POST.getlist('option_b[]')
        opts_c = request.POST.getlist('option_c[]')
        opts_d = request.POST.getlist('option_d[]')

        for i in range(len(q_texts)):
            # Get the correct option for this question using its index
            correct_option = request.POST.get(f'correct_options_{i + 1}[]')
            
            if q_ids[i] == 'new':
                QuizQuestions.objects.create(
                    quiz=quiz,
                    question=q_texts[i],
                    option_a=opts_a[i],
                    option_b=opts_b[i],
                    option_c=opts_c[i],
                    option_d=opts_d[i],
                    correct_option=correct_option or 'A'
                )
            else:
                q_obj = get_object_or_404(QuizQuestions, id=q_ids[i])
                q_obj.question = q_texts[i]
                q_obj.option_a = opts_a[i]
                q_obj.option_b = opts_b[i]
                q_obj.option_c = opts_c[i]
                q_obj.option_d = opts_d[i]
                q_obj.correct_option = correct_option or 'A'
                q_obj.save()

        messages.success(request, "Node Registry Synchronized Successfully.")
        
    
    return render(request, 'subject_head/edit_quiz.html', {
        'quiz': quiz,
        'questions': questions,
        'classes': allotted_classes, # CHANGED: Now only contains the allowed class
        'subject_head': subject_head_name
    })
    
    
def delete_quiz(request, quiz_id):
    # Ensure this action only happens via POST for security
    if request.method == "POST":
        try:
            # 1. Fetch the node
            quiz_obj = get_object_or_404(Quizzes, pk=quiz_id)
            
            # 2. Manual cleanup of related questions (Optional but safer)
            QuizQuestions.objects.filter(quiz=quiz_obj).delete()
            
            # 3. Terminate main registry node
            quiz_obj.delete()
            
            messages.success(request, f"REGISTRY_TERMINATED: Node #{quiz_id} purged from system.")
        except Exception as e:
            messages.error(request, f"TERMINATION_FAILED: Critical system error - {str(e)}")
            
        return redirect('subject_head_quiz') # Redirect back to the quiz list
    
    # If someone tries to access via URL directly (GET), just redirect them
    return redirect('subject_head_quiz')




def get_user_by_email(email, user_type):
    mapping = {'institution': Institution, 
                'student': Students,
                'class_head' : Classes,
                'subject_head' : Subjects,
                'parent' : Parents
            }
    model = mapping.get(user_type)
    if model:
        user = model.objects.filter(email=email).first()
        if user:
            return user, model
    return None, None


def user_portal(request):
    return render(request, 'user_portal.html')

def generate_otp():
    import random
    return random.randint(100000, 999999)

def forgot_password(request, user_type):
    if request.method == 'POST':
        email = request.POST.get('email')
        user, _ = get_user_by_email(email, user_type)
        print(user_type)
        
        if not user:
            return JsonResponse({'status': 'error', 'message': 'Email not found in our records.'})
        
        otp = generate_otp() # Your existing OTP generator
        print(f"Generated OTP: {otp} for email: {email}")
        cache.set(f"otp_{email}", otp, timeout=180)
        request.session['reset_email'] = email
        request.session['reset_user_type'] = user_type
        
        try:
            send_otp_via_email(email, otp)
            return JsonResponse({'status': 'success', 'message': 'OTP sent to your email.'})
        except Exception:
            return JsonResponse({'status': 'error', 'message': 'Failed to send email.'})
            
    return render(request, 'forgot_password.html', {'user_type': user_type})

def verify_otp(request, user_type):
    if request.method == 'POST':
        otp = request.POST.get('otp')
        email = request.session.get('reset_email')
        
        cached_otp = cache.get(f"otp_{email}")
        if cached_otp and str(cached_otp) == str(otp):
            # We set a 'verified' flag in session so they can proceed to reset
            request.session['otp_verified'] = True
            return JsonResponse({'status': 'success', 'message': 'OTP verified.'})
        
        return JsonResponse({'status': 'error', 'message': 'Invalid or expired OTP.'})

def reset_password(request, user_type):
    if request.method == 'POST':
        email = request.session.get('reset_email')
        is_verified = request.session.get('otp_verified')
        new_password = request.POST.get('new_password')
        
        if not email or not is_verified:
            return JsonResponse({'status': 'error', 'message': 'Session expired. Restart process.'})

        user, _ = get_user_by_email(email, user_type)
       
        if user:
            # Encrypt the new password before saving
            user.password = encrypt_password(new_password)
            user.save()
            
            # Cleanup session
            del request.session['reset_email']
            del request.session['otp_verified']
            return JsonResponse({'status': 'success', 'message': 'Password updated successfully!'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request.'})


def send_otp_via_email(email, otp):
    print(f"Sending OTP: {otp} to {email}")
    def _send():
        subject = f"{otp} is your Secure Recovery Code"
        
        # Clean plain-text body (used if HTML fails or for notifications)
        message = textwrap.dedent(f"""
            Hello,

            We received a request to reset your password for your Eduke account. 
            Use the following verification token to proceed:

            Verification Token: {otp}

            This code is valid for 3 minutes. For your security, do not share this code with anyone.

            If you did not request this reset, please ignore this email.

            Securely,
            The Eduke Security Team
        """).strip()
        
        # Styled HTML version
        html_message = f"""
        <div style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 500px; margin: auto; padding: 20px; border: 1px solid #e2e8f0; border-radius: 16px; background-color: #ffffff;">
            <div style="text-align: center; margin-bottom: 20px;">
                <h2 style="color: #7c3aed; margin-bottom: 5px;">Access Reset</h2>
                <p style="font-size: 14px; color: #64748b; text-transform: uppercase; letter-spacing: 1px;">Step 2: Authenticity Check</p>
            </div>
            
            <p style="color: #1e293b; line-height: 1.5;">Hello,</p>
            <p style="color: #475569; line-height: 1.5;">Use the verification token below to complete your password reset. This code is valid for <strong style="color: #7c3aed;">3 minutes</strong>.</p>
            
            <div style="background: #f5f3ff; padding: 30px; text-align: center; border-radius: 12px; margin: 25px 0; border: 1px dashed #c084fc;">
                <span style="font-size: 36px; font-weight: 800; letter-spacing: 8px; color: #2e1065; font-family: monospace;">{otp}</span>
            </div>
            
            <p style="font-size: 12px; color: #94a3b8; text-align: center; margin-top: 25px;">
                If you did not request this reset, please ignore this email or contact security support.<br>
                <strong>Do not share this code with anyone.</strong>
            </p>
        </div>
        """
        url = os.environ.get('EMAIL_URL')
    # Create and start the thread
    
        epayload = {
            "email": email,
            "subject": subject,
            "html_content": html_message
        }
        print(f"DEBUG: Payload created: {epayload}")
        
        # Send the request to Pipedream
        try:
            response = requests.post(url, json=epayload)
            response.raise_for_status()
        except requests.exceptions.RequestException as e:
            print(f"Error sending email: {e}")
        
        try:
            send_mail(
                subject=subject,
                message=message,  # This is the "body"
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
                html_message=html_message
            )
            print(f"OTP email successfully dispatched to {email}")
        except Exception as e:
            print(f"Failed to send OTP email: {str(e)}")

    # Start the background thread
    threading.Thread(target=_send, daemon=True).start()
 

def notify_students_new_quiz(email_list, subject_name, quiz_name):
    """
    Notifies a list of students about a new quiz via background thread.
    """
    def _send_notifications():
        for email in email_list:
            # Optional: Fetch student name if you want to personalize the email
            student_name = "Student"
            try:
                with connection.cursor() as cursor:
                    cursor.execute("SELECT name FROM main_students WHERE email = %s", [email])
                    row = cursor.fetchone()
                    if row:
                        student_name = row[0]
            except Exception:
                pass # Fallback to "Student" if query fails

            subject = f"New Quiz Available: {quiz_name} ({subject_name})"
            
            # Plain text version
            message = textwrap.dedent(f"""
                Hello {student_name},

                A new quiz has been posted for your subject: {subject_name}.

                Quiz Name: {quiz_name}

                How to attend:
                1. Login to your Student Portal.
                2. Select your 'Class' option.
                3. Choose the subject: {subject_name}.
                4. Go to the 'Subject Quiz' section.
                5. Locate '{quiz_name}' and click open to begin.

                Good luck!
                The Eduke Team
            """).strip()

            # Styled HTML version
            html_message = f"""
            <div style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 600px; margin: auto; padding: 25px; border: 1px solid #e2e8f0; border-radius: 16px; background-color: #ffffff;">
                <div style="text-align: center; margin-bottom: 25px;">
                    <h2 style="color: #7c3aed; margin-bottom: 5px;">New Quiz Released!</h2>
                    <p style="font-size: 14px; color: #64748b; text-transform: uppercase; letter-spacing: 1px;">{subject_name}</p>
                </div>
                
                <p style="color: #1e293b; font-size: 16px;">Hello <strong>{student_name}</strong>,</p>
                <p style="color: #475569; line-height: 1.6;">You have a new assessment waiting for you in <strong>{subject_name}</strong>. Please complete the quiz <strong>"{quiz_name}"</strong> at your earliest convenience.</p>
                
                <div style="background: #f8fafc; padding: 20px; border-radius: 12px; border-left: 4px solid #7c3aed; margin: 25px 0;">
                    <h4 style="margin-top: 0; color: #1e293b;">How to attend:</h4>
                    <ol style="color: #475569; line-height: 1.8; padding-left: 20px;">
                        <li>Login to your <strong>Student Portal</strong>.</li>
                        <li>Navigate to your <strong>Class</strong> menu.</li>
                        <li>Select <strong>{subject_name}</strong> from your subjects.</li>
                        <li>Open the <strong>Subject Quiz</strong> section.</li>
                        <li>Click on <strong>{quiz_name}</strong> to start.</li>
                    </ol>
                </div>
                
                <div style="text-align: center; margin-top: 30px;">
                    <p style="font-size: 12px; color: #94a3b8;">
                        This is an automated notification. Please do not reply to this email.<br>
                        &copy; {2026} Eduke Learning Management System
                    </p>
                </div>
            </div>
            """

            try:
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email], # Sent individually for privacy
                    fail_silently=False,
                    html_message=html_message
                )
            except Exception as e:
                print(f"Failed to notify {email}: {str(e)}")

    # Start the background thread
    threading.Thread(target=_send_notifications, daemon=True).start()
from .models import AuditLog

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def access_logs(request):
    institution_id = request.session.get('institution_id')
    if not institution_id:
         # Handle case where institution is not logged in, maybe redirect to login
         return redirect('login') # Or appropriate login page

    institution = get_object_or_404(Institution, institution_id=institution_id)
    
    # Filter logs for this institution
    logs = AuditLog.objects.filter(institution=institution).order_by('-timestamp')[:50]
    processed_logs = []
    
    for log in logs:
        # Default to institution info
        prefix = institution.institution_abbreviation or "INS"
        username = institution.institution_name
        
        # If there's a specific user, get their details
        if log.user:
            try:
                from .models import Students, Classes, Subjects, Parents
                
                # Check if it's a student
                if log.user.role == 'student':
                    student = Students.objects.filter(user=log.user).first()
                    if student:
                        prefix = student.roll_no[:2].upper() or "ST"
                        username = student.name
                
                # Check if it's a class head
                elif log.user.role == 'class_head':
                    class_obj = Classes.objects.filter(user=log.user).first()
                    if class_obj:
                        prefix = class_obj.class_abbreviation[:2].upper() or "CH"
                        username = class_obj.class_head
                
                # Check if it's a subject head
                elif log.user.role == 'subject_head':
                    subject_obj = Subjects.objects.filter(user=log.user).first()
                    if subject_obj:
                        prefix = subject_obj.subject_name.split()[0][:2].upper() if subject_obj.subject_name else "SH"
                        username = subject_obj.subject_head
                
                # Check if it's a parent
                elif log.user.role == 'parent':
                    parent_obj = Parents.objects.filter(user=log.user).first()
                    if parent_obj:
                        # Get parent name or use student info
                        if parent_obj.name:
                            prefix = "PT"
                            username = parent_obj.name
                        else:
                            # Use student roll number as fallback
                            student = Students.objects.filter(id=parent_obj.student_id).first()
                            if student:
                                prefix = f"P-{student.roll_no[:2]}" if student.roll_no else "PT"
                                username = f"Parent of {student.name}"
            except Exception as e:
                print(f"Error processing user details for audit log: {e}")
        
        local_time = localtime(log.timestamp)
        
        processed_logs.append({
            'user_prefix': prefix,
            'username': username,
            'ip': log.ip_address,
            'action': log.action,
            'date': local_time.strftime("%b %d"),
            'time': local_time.strftime("%I:%M %p"),
            'status': True
        })
        
    return render(request, 'admin/access.html', {'logs': processed_logs, 'institution': institution})

@csrf_exempt
def contact_email(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            name = data.get('name')
            institution = data.get('institution')
            email = data.get('email')
            message = data.get('message')

            if not all([name, institution, email, message]):
                return JsonResponse({'status': 'error', 'message': 'All fields are required.'}, status=400)

            # 1. Plain Text Version (Back-up)
            text_content = textwrap.dedent(f"""
                New Contact Form Submission:
                
                Name: {name}
                Institution: {institution}
                Email: {email}

                Project Brief:
                {message}

                Sent via Eduke Admin Portal.
            """).strip()

            # 2. Styled HTML Version (Branded)
            html_content = f"""
            <div style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 600px; margin: auto; padding: 20px; border: 1px solid #e2e8f0; border-radius: 16px; background-color: #ffffff;">
                <div style="text-align: center; margin-bottom: 25px; border-bottom: 2px solid #f5f3ff; padding-bottom: 15px;">
                    <h2 style="color: #7c3aed; margin-bottom: 5px;">New Contact Request</h2>
                    <p style="font-size: 14px; color: #64748b; text-transform: uppercase; letter-spacing: 1px;">Eduke Lead Generation</p>
                </div>
                
                <div style="background: #f8fafc; padding: 20px; border-radius: 12px; margin-bottom: 20px;">
                    <p style="margin: 5px 0;"><strong style="color: #475569;">From:</strong> {name}</p>
                    <p style="margin: 5px 0;"><strong style="color: #475569;">Institution:</strong> {institution}</p>
                    <p style="margin: 5px 0;"><strong style="color: #475569;">Email:</strong> <a href="mailto:{email}" style="color: #7c3aed; text-decoration: none;">{email}</a></p>
                </div>

                <div style="padding: 10px 5px;">
                    <h3 style="color: #1e293b; font-size: 16px; margin-bottom: 10px;">Project Brief:</h3>
                    <p style="color: #475569; line-height: 1.6; white-space: pre-wrap; background: #ffffff; border-left: 4px solid #7c3aed; padding-left: 15px;">{message}</p>
                </div>
                
                <div style="margin-top: 30px; padding-top: 15px; border-top: 1px solid #f1f5f9; text-align: center;">
                    <a href="mailto:{email}" style="display: inline-block; background-color: #7c3aed; color: #ffffff; padding: 12px 25px; border-radius: 8px; text-decoration: none; font-weight: bold;">Reply to Lead</a>
                </div>
                
                <p style="font-size: 11px; color: #94a3b8; text-align: center; margin-top: 25px;">
                    This is an automated notification from the Eduke Platform.
                </p>
            </div>
            """

            subject = f"🚀 New Lead: {name} from {institution}"

            def send_email_thread():
                try:
                    send_mail(
                        subject,
                        text_content,
                        settings.DEFAULT_FROM_EMAIL,
                        [settings.DEFAULT_FROM_EMAIL],
                        html_message=html_content,  # This adds the styled version
                        fail_silently=False,
                    )
                except Exception as e:
                    print(f"Error sending contact email: {e}")

            threading.Thread(target=send_email_thread, daemon=True).start()

            return JsonResponse({'status': 'success', 'message': 'Your message has been sent successfully!'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)